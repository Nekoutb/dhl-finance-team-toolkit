"""Sales — Revenue Analysis.

Feeds on the monthly IB434 revenue detail extract (one row per billing line,
56 columns). Each file carries its Billing Period; uploading a month again
REPLACES it, which is how the running month grows through the month.

Definitions (agreed with the finance lead):

* REVENUE RECOGNISED = LCU Total − LCU Taxes to Applicable Charges
  (column BD less column BC). VAT comes out; everything else the invoice
  carries stays in.
* The KPIs (per day, per shipment, per kilo) are built on the WEIGHT CHARGE
  alone (column AO / its LCU twin AX) — the carriage the country actually
  sells — not on the recognised total, which carries fuel, fees and duty.
* BILLABLE DAYS come from the DATA, not a calendar: a day with no billing
  is assumed non-billable (public holidays fall out automatically).
  Sundays never count; an active Saturday counts HALF a day ("two Saturdays
  count as one day"); an active weekday counts one. Counted on the INVOICE
  date — the same axis the revenue sits on, so a partial month never
  divides one span's revenue by another span's days.
* Credit-and-rebill pairs (negative reversal + positive rebill on the same
  airwaybill) are handled by SIGNED sums throughout — never de-duplicate.
* PRICING uses the Weight Charge alone ("the weight charge is the revenue
  for the country"): price/kg = weight charge ÷ billed kilos.
* FUEL SURCHARGE is measured as fuel ÷ weight charge, and only on the
  products that carry one: D, N, P, T and Y.
"""
import json
import math
import os
import re
import threading
import time
from datetime import date, datetime, timedelta
from pathlib import Path

from ..config import DATA_DIR
from .iro import _FileLock

STORE_PATH = DATA_DIR / "revenue" / "store.json"
UPLOAD_DIR = DATA_DIR / "revenue" / "uploads"

# A day is ACTIVE when it carries at least this many shipment lines — a lone
# stray shipment on a public holiday must not turn the holiday billable.
MIN_ACTIVE_ROWS = 3

_lock = threading.Lock()

# The columns the parser needs, matched on the EXACT IB434 header names
# (case/space tolerant). Parsing fails loudly if one is missing — a wrong
# file must never produce a silently empty month.
_REQUIRED = [
    "Billing Period", "Air waybill", "Bill To Account",
    "Bill To Account Name", "Shipment Date", "Invoice Date",
    "Billed Weight (Kilos)", "LCU Weight Charge", "LCU Fuel Surcharges",
    "LCU Other Charges", "LCU Discount", "LCU Imp/Exp Duties & Taxes",
    "LCU Taxes to Applicable Charges", "LCU Total",
    "Service Type", "Billing Type", "Orgn", "Dest",
    "Local Product Code",
]

# Lanes kept per direction, biggest first. The PRIOR months must keep more
# than the ten shown, or a lane that merely ranked lower last month comes
# back as "new" instead of being compared.
MAX_LANES = 400

# Lanes are COUNTRY to country, not city to city — the file routes on IATA
# city codes (DLA, YAO, BRU …) and the owner reports on countries. The map
# ships with the app (samples/reference/iata_country.json, generated from an
# open airport dataset with the corrections this file's own data proves —
# chiefly BAF = Bafoussam CM, which an airport-keyed dataset places in the
# USA). Anything unmapped renders as "?XXX" rather than guessed, and the
# page lists those codes so they can be filled in.
_IATA_PATH = (Path(__file__).resolve().parent.parent.parent
              / "samples" / "reference" / "iata_country.json")
_iata_cache = None

# Billing type T is duty / customs billing: it carries charges but ZERO
# weight, is billed to DHL itself, and its "origin" is an internal code
# (MHN, ZJF …) that is not a place. Left in, it invents phantom lanes and
# poisons every RPK it touches, so lanes exclude it. The money stays in the
# month's revenue — it simply is not lane traffic.
LANE_EXCLUDED_BILLING_TYPES = {"T"}

# Only these products carry a fuel surcharge. The file writes the code with
# trailing spaces ("P "), so it is stripped before comparison.
FUEL_PRODUCTS = {"D", "N", "P", "T", "Y"}


def iata_countries():
    """{IATA city code: ISO country}, config overrides applied last."""
    global _iata_cache
    if _iata_cache is None:
        try:
            base = json.loads(_IATA_PATH.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            base = {}
        from ..config import load_config
        override = (load_config().get("iata_country_overrides") or {})
        base.update({str(k).strip().upper()[:3]: str(v).strip().upper()[:2]
                     for k, v in override.items() if k and v})
        _iata_cache = base
    return _iata_cache


def invalidate_iata_cache():
    global _iata_cache
    _iata_cache = None


def country_of(code):
    """ISO country for an IATA city code, or '?CODE' when unmapped."""
    c = str(code or "").strip().upper()
    if not c:
        return ""
    hit = iata_countries().get(c)
    return hit if hit else f"?{c}"


def _norm_h(h):
    return re.sub(r"\s+", " ", str(h or "")).strip().lower()


def _num(v):
    """A finite float, or 0. A text 'NaN'/'inf' cell passes float() but a
    single NaN would poison every sum it touches AND survive the JSON round
    trip — so non-finite values are rejected here, at the door."""
    if v in (None, ""):
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    return f if math.isfinite(f) else 0.0


_EXCEL_EPOCH = datetime(1899, 12, 30)


def _iso_day(v):
    """A cell's date part as YYYY-MM-DD, or ''.

    Handles the shapes a re-exported file can carry: real datetimes, ISO
    text, Excel serial numbers (a stripped number format leaves 45808.0) and
    dd/mm/yyyy text. Everything is CALENDAR-validated — an ISO-shaped
    impossible date like 2026-06-31 must never reach the store, where it
    would break the billable-day arithmetic on every page view."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)) and 20000 <= v <= 80000:
        return (_EXCEL_EPOCH + timedelta(days=float(v))).strftime("%Y-%m-%d")
    s = str(v or "").strip()
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        candidate = m.group(0)
    else:
        m2 = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
        if not m2:
            return ""
        candidate = (f"{m2.group(3)}-{int(m2.group(2)):02d}"
                     f"-{int(m2.group(1)):02d}")
    try:
        date.fromisoformat(candidate)
    except ValueError:
        return ""
    return candidate


def _load():
    if STORE_PATH.exists():
        try:
            return json.loads(STORE_PATH.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"periods": {}}


def _save(data):
    STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = STORE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    os.replace(tmp, STORE_PATH)


def parse_file(path, source=""):
    """One IB434 file → one period record. Raises ValueError on a file that
    is not an IB434 revenue detail (missing columns, no period)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = [_norm_h(h) for h in next(rows_iter)]
        ix = {}
        for name in _REQUIRED:
            try:
                ix[name] = header.index(_norm_h(name))
            except ValueError:
                raise ValueError(
                    f"column '{name}' not found — is this the IB434 "
                    "revenue detail export?")
        totals = {k: 0.0 for k in ("weight", "fuel", "other", "discount",
                                   "duty", "tax", "net", "gross")}
        fuel_base = {"weight": 0.0, "fuel": 0.0, "rows": 0}
        fuel_cust = {}
        kilos = 0.0
        rows = 0
        period_votes = {}
        daily = {}          # invoice day -> {net, kilos, awb_net{}}
        ship_days = {}
        awb_net = {}
        customers = {}
        lanes = {"OB": {}, "IB": {}}
        seen_codes = set()
        for r in rows_iter:
            if not any(v not in (None, "") for v in r):
                continue
            rows += 1
            p = str(r[ix["Billing Period"]] or "").strip()
            if re.fullmatch(r"\d{4}-\d{2}", p):
                period_votes[p] = period_votes.get(p, 0) + 1
            w = _num(r[ix["LCU Weight Charge"]])
            f = _num(r[ix["LCU Fuel Surcharges"]])
            o = _num(r[ix["LCU Other Charges"]])
            d = _num(r[ix["LCU Discount"]])
            duty = _num(r[ix["LCU Imp/Exp Duties & Taxes"]])
            tax = _num(r[ix["LCU Taxes to Applicable Charges"]])
            # Revenue recognised = LCU Total less LCU Taxes (BD - BC).
            gross = _num(r[ix["LCU Total"]])
            net = gross - tax
            kg = _num(r[ix["Billed Weight (Kilos)"]])
            # Fuel surcharge is only levied on these products; measuring it
            # across the rest would dilute the percentage with rows that can
            # never carry one.
            product = str(r[ix["Local Product Code"]] or "").strip().upper()
            if product in FUEL_PRODUCTS:
                fuel_base["weight"] += w
                fuel_base["fuel"] += f
                fuel_base["rows"] += 1
                acct_name = str(r[ix["Bill To Account Name"]] or "").strip()
                fkey = acct_name.upper() or "?"
                fc = fuel_cust.get(fkey)
                if fc is None:
                    fc = fuel_cust[fkey] = {"name": acct_name or fkey,
                                            "weight": 0.0, "fuel": 0.0,
                                            "rows": 0}
                fc["weight"] += w
                fc["fuel"] += f
                fc["rows"] += 1
            totals["weight"] += w
            totals["fuel"] += f
            totals["other"] += o
            totals["discount"] += d
            totals["duty"] += duty
            totals["tax"] += tax
            totals["net"] += net
            totals["gross"] += gross
            kilos += kg
            inv_day = _iso_day(r[ix["Invoice Date"]])
            if inv_day:
                dd = daily.get(inv_day)
                if dd is None:
                    dd = daily[inv_day] = {"net": 0.0, "weight": 0.0,
                                           "kilos": 0.0, "awb_net": {}}
                dd["net"] += net
                dd["weight"] += w
                dd["kilos"] += kg
            ship_day = _iso_day(r[ix["Shipment Date"]])
            if ship_day:
                ship_days[ship_day] = ship_days.get(ship_day, 0) + 1
            awb = str(r[ix["Air waybill"]] or "").strip()
            if awb:
                awb_net[awb] = awb_net.get(awb, 0.0) + net
                if inv_day:
                    dd["awb_net"][awb] = dd["awb_net"].get(awb, 0.0) + net
            # Lanes: OB = leaving the country, IB = coming in, aggregated
            # COUNTRY to country. Duty-billing rows are not lane traffic.
            svc = str(r[ix["Service Type"]] or "").strip().upper()
            btype = str(r[ix["Billing Type"]] or "").strip().upper()
            orgn = str(r[ix["Orgn"]] or "").strip().upper()
            dest = str(r[ix["Dest"]] or "").strip().upper()
            if orgn:
                seen_codes.add(orgn)
            if dest:
                seen_codes.add(dest)
            if (svc in lanes and orgn and dest
                    and btype not in LANE_EXCLUDED_BILLING_TYPES):
                lane = lanes[svc].setdefault(
                    f"{country_of(orgn)}-{country_of(dest)}", {
                        "net": 0.0, "kilos": 0.0, "awb_net": {}})
                lane["net"] += net
                lane["kilos"] += kg
                if awb:
                    lane["awb_net"][awb] = lane["awb_net"].get(awb, 0.0) + net
            acct = str(r[ix["Bill To Account"]] or "").strip()
            if acct:
                c = customers.get(acct)
                if c is None:
                    c = customers[acct] = {"name": "", "net": 0.0,
                                           "weight": 0.0, "kilos": 0.0}
                name = str(r[ix["Bill To Account Name"]] or "").strip()
                if name:
                    c["name"] = name
                c["net"] += net
                c["weight"] += w
                c["kilos"] += kg
    finally:
        wb.close()
    if not period_votes:
        raise ValueError("no Billing Period values found in the file")
    period = max(period_votes, key=period_votes.get)
    for k in totals:
        totals[k] = round(totals[k], 2)
    for c in customers.values():
        for k in ("net", "weight", "kilos"):
            c[k] = round(c[k], 2)
    daily_out = {}
    for day, dd in sorted(daily.items()):
        awbs = sorted(a for a, v in dd["awb_net"].items() if v > 0.005)
        daily_out[day] = {
            "net": round(dd["net"], 2), "weight": round(dd["weight"], 2),
            "kilos": round(dd["kilos"], 2),
            "shipments": len(awbs),
            # the AWBs themselves, so a partial-month window counts each
            # shipment ONCE even when it is invoiced across two days
            "awbs": awbs}
    lanes_out = {}
    for svc, table in lanes.items():
        top = sorted(table.items(), key=lambda kv: -kv[1]["net"])[:MAX_LANES]
        lanes_out[svc] = {
            k: {"net": round(v["net"], 2), "kilos": round(v["kilos"], 2),
                "shipments": sum(1 for x in v["awb_net"].values()
                                 if x > 0.005)}
            for k, v in top}
    return {
        "period": period,
        "source": str(source or Path(path).name),
        "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rows": rows,
        "totals": totals,
        "kilos": round(kilos, 2),
        # A shipment = an airwaybill whose SIGNED net is positive this
        # period; a fully reversed billing nets to zero and is not one.
        "shipments": sum(1 for v in awb_net.values() if v > 0.005),
        "daily": daily_out,
        "lanes": lanes_out,
        # city codes this file used that the country map does not know —
        # surfaced on the page instead of being silently guessed
        "unmapped_codes": sorted(c for c in seen_codes
                                 if country_of(c).startswith("?")),
        "fuel": {"weight": round(fuel_base["weight"], 2),
                 "fuel": round(fuel_base["fuel"], 2),
                 "rows": fuel_base["rows"],
                 "customers": {k: {"name": v["name"],
                                   "weight": round(v["weight"], 2),
                                   "fuel": round(v["fuel"], 2),
                                   "rows": v["rows"]}
                               for k, v in fuel_cust.items()}},
        "ship_days": dict(sorted(ship_days.items())),
    }, {acct: c for acct, c in customers.items()}


def store_period(record, customers):
    """Insert/replace one period atomically."""
    record = dict(record)
    record["customers"] = customers
    with _lock, _FileLock(STORE_PATH):
        data = _load()
        data["periods"][record["period"]] = record
        _save(data)
    return record


def delete_period(period):
    with _lock, _FileLock(STORE_PATH):
        data = _load()
        gone = data["periods"].pop(str(period), None)
        _save(data)
    return gone is not None


# --------------------------------------------------------------------------- #
# Background ingest (same contract as the BIT/Cash AR upload: the request
# returns at once, the page polls `processing` until the month lands).
# --------------------------------------------------------------------------- #
def ingest_async(jobs):
    """``jobs`` = [(stored_path, source_name)] — parse on a thread."""
    with _lock, _FileLock(STORE_PATH):
        data = _load()
        data["processing"] = {
            "started": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "beat": time.time(),
            "files": [s for _p, s in jobs]}
        data.pop("processing_error", None)
        _save(data)

    def _runner():
        errors = []
        for path, source in jobs:
            # heartbeat per file, so a long multi-file ingest is never
            # mistaken for an interrupted one by the stale-beat cleanup
            with _lock, _FileLock(STORE_PATH):
                data = _load()
                if data.get("processing"):
                    data["processing"]["beat"] = time.time()
                    _save(data)
            try:
                record, customers = parse_file(path, source)
                store_period(record, customers)
            except Exception as exc:  # noqa: BLE001 — surfaced on the page
                errors.append(f"{source}: {exc}")
            finally:
                # the spooled copy has served its purpose — the store keeps
                # the aggregates; keeping every upload would grow forever
                Path(path).unlink(missing_ok=True)
        with _lock, _FileLock(STORE_PATH):
            data = _load()
            data.pop("processing", None)
            if errors:
                data["processing_error"] = {
                    "at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "message": " · ".join(errors)}
            _save(data)

    threading.Thread(target=_runner, daemon=True).start()


def status():
    data = _load()
    proc = data.get("processing")
    if proc and time.time() - (proc.get("beat") or 0) > 900:
        with _lock, _FileLock(STORE_PATH):
            data = _load()
            if data.get("processing"):
                data.pop("processing", None)
                data["processing_error"] = {
                    "at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "message": "Processing was interrupted before it "
                               "finished — upload the file again."}
                _save(data)
    return {"processing": data.get("processing"),
            "processing_error": data.get("processing_error")}


# --------------------------------------------------------------------------- #
# Metrics
# --------------------------------------------------------------------------- #
def _union_ship_days(periods):
    """Billing activity per calendar day across EVERY uploaded file.

    Counted on the INVOICE date — the owner's rule is "where you don't see
    any billing, it was not a billable day", and the revenue being divided
    is invoice-dated, so the divisor must sit on the same axis or a partial
    month compares two different spans. Records written before that detail
    existed fall back to their shipment dates.
    """
    union = {}
    for rec in periods.values():
        daily = rec.get("daily")
        if daily:
            for day, dd in daily.items():
                union[day] = union.get(day, 0) + (dd.get("shipments") or 0)
        else:
            for day, n in (rec.get("ship_days") or {}).items():
                union[day] = union.get(day, 0) + n
    return union


def billable_days(month_key, union_days):
    """Billable days of ``month_key`` (YYYY-MM) under the agreed rule:
    only days the data shows activity on count (>= MIN_ACTIVE_ROWS lines);
    Sundays never count; Saturdays count half; weekdays count one."""
    total = 0.0
    for day, n in union_days.items():
        if not day.startswith(month_key) or n < MIN_ACTIVE_ROWS:
            continue
        try:
            wd = date.fromisoformat(day).weekday()  # Mon=0 … Sun=6
        except ValueError:                  # stored before validation existed
            continue
        if wd == 6:
            continue
        total += 0.5 if wd == 5 else 1.0
    return total


def month_metrics(rec, union_days):
    """Headline revenue is BD − BC; the three KPIs are built on the WEIGHT
    CHARGE, which is the carriage the country sells — fuel, fees and duty
    ride on top of it and would flatter every ratio."""
    net = rec["totals"]["net"]
    weight = rec["totals"]["weight"]
    days = billable_days(rec["period"], union_days)
    shipments = rec.get("shipments") or 0
    kilos = rec.get("kilos") or 0.0
    return {
        "period": rec["period"],
        "source": rec.get("source", ""),
        "uploaded": rec.get("uploaded", ""),
        "rows": rec.get("rows", 0),
        "net": net,
        "weight": weight,
        "gross": rec["totals"]["gross"],
        "totals": rec["totals"],
        "billable_days": days,
        "rev_per_day": (weight / days) if days else None,
        "shipments": shipments,
        "rev_per_shipment": (weight / shipments) if shipments else None,
        "kilos": kilos,
        "rev_per_kg": (weight / kilos) if kilos > 0 else None,
    }


def _month_billable_sequence(month_key, union_days):
    """The month's billable days in date order as [(iso_day, weight)]."""
    seq = []
    for day, n in sorted(union_days.items()):
        if not day.startswith(month_key) or n < MIN_ACTIVE_ROWS:
            continue
        try:
            wd = date.fromisoformat(day).weekday()
        except ValueError:
            continue
        if wd == 6:
            continue
        seq.append((day, 0.5 if wd == 5 else 1.0))
    return seq


def same_days_window(rec, union_days, target_days):
    """A month's position after its FIRST ``target_days`` billable days —
    the fair yardstick for an ongoing month ("compare it to the same number
    of days in the prior month", not to a full month's average).

    Uses the invoice-dated daily detail; None when the record predates that
    detail or the window cannot be built."""
    daily = rec.get("daily")
    if not daily or not target_days:
        return None
    seq = _month_billable_sequence(rec["period"], union_days)
    if not seq:
        return None
    cum, cutoff = 0.0, None
    for day, w in seq:
        cum += w
        cutoff = day
        if cum >= target_days - 1e-9:
            break
    net = kilos = weight = 0.0
    seen = set()
    counted = 0
    for day, dd in daily.items():
        if day.startswith(rec["period"]) and day <= cutoff:
            net += dd.get("net", 0.0)
            weight += dd.get("weight", 0.0)
            kilos += dd.get("kilos", 0.0)
            awbs = dd.get("awbs")
            if awbs is None:            # pre-v11.19 detail: best effort
                counted += dd.get("shipments", 0)
            else:
                seen.update(awbs)
    ships = len(seen) + counted
    return {"days": cum, "through": cutoff, "net": round(net, 2),
            "weight": round(weight, 2),
            "kilos": round(kilos, 2), "shipments": ships,
            "rev_per_day": (weight / cum) if cum else None,
            "rev_per_shipment": (weight / ships) if ships else None,
            "rev_per_kg": (weight / kilos) if kilos > 0 else None}


def like_for_like(months, periods, union_days):
    """Every month cut back to the SAME number of billable days the running
    month has reached, so the KPI lines compare equal spans.

    Without this a 3-day August sits beside a 25-day July and the chart
    reads as a collapse that is really just a shorter month. Months whose
    stored detail predates this feature keep their full-month figures and
    are marked ``partial=False`` so the page can say so.
    """
    ongoing = next((m for m in months if m["ongoing"]), None)
    if not ongoing or not ongoing["billable_days"]:
        return months, None
    target = ongoing["billable_days"]
    out = []
    for m in months:
        if m["ongoing"]:
            out.append({**m, "clipped": True})
            continue
        w = same_days_window(periods[m["period"]], union_days, target)
        if not w or not w["net"]:
            out.append({**m, "clipped": False})
            continue
        out.append({**m, "clipped": True, "billable_days": w["days"],
                    "net": w["net"], "weight": w["weight"],
                    "kilos": w["kilos"],
                    "shipments": w["shipments"],
                    "rev_per_day": w["rev_per_day"],
                    "rev_per_shipment": w["rev_per_shipment"],
                    "rev_per_kg": w["rev_per_kg"],
                    "through": w["through"]})
    return out, target


def landing_estimate(ongoing, months):
    """Where the running month lands if it keeps its current daily rate,
    projected over the typical billable days of the completed months."""
    if not ongoing or not ongoing["billable_days"] or not ongoing["net"]:
        return None
    full = [m["billable_days"] for m in months
            if not m["ongoing"] and m["billable_days"]]
    if not full:
        return None
    typical = sum(full) / len(full)
    rate = ongoing["net"] / ongoing["billable_days"]
    net = rate * typical
    weight = ongoing["weight"] / ongoing["billable_days"] * typical
    ship_rate = ongoing["shipments"] / ongoing["billable_days"]
    kilo_rate = ongoing["kilos"] / ongoing["billable_days"]
    shipments = ship_rate * typical
    kilos = kilo_rate * typical
    return {"label": ongoing["label"] + " landing",
            "period": ongoing["period"] + "-landing",
            "typical_days": round(typical, 1),
            "elapsed_days": ongoing["billable_days"],
            "net": net, "weight": weight,
            "billable_days": round(typical, 1),
            "shipments": shipments, "kilos": kilos,
            "rev_per_day": ongoing["rev_per_day"],
            "rev_per_shipment": ongoing["rev_per_shipment"],
            "rev_per_kg": ongoing["rev_per_kg"],
            "totals": {"tax": ongoing["totals"]["tax"] / ongoing["billable_days"] * typical,
                       "duty": ongoing["totals"]["duty"] / ongoing["billable_days"] * typical}}


def _graph_series(months):
    """Pixel-ready polylines for the three KPIs: a SOLID line through the
    complete months and a DASHED closing segment to the ongoing month, so
    the reader sees at a glance that the last point is provisional."""
    W, H, PAD = 460, 132, 30
    out = []
    for key, label in (("rev_per_day", "Revenue / day"),
                       ("rev_per_shipment", "Revenue / shipment"),
                       ("rev_per_kg", "Revenue / kg")):
        pts = [(m["label"], m[key], m["ongoing"])
               for m in months if m.get(key)]
        if len(pts) < 2:
            continue
        vals = [v for _l, v, _o in pts]
        lo, hi = min(vals), max(vals)
        flat = (hi - lo) < 1e-9
        span = (hi - lo) or 1.0
        step = (W - 2 * PAD) / (len(pts) - 1)
        coords = []
        for i, (lab, v, ongoing) in enumerate(pts):
            # A flat series would otherwise pin every point to the floor of
            # the chart, reading as a collapse — draw it mid-height instead.
            frac = 0.5 if flat else (v - lo) / span
            coords.append({
                "x": round(PAD + i * step, 1),
                "y": round(H - PAD - frac * (H - 2 * PAD), 1),
                "v": v, "label": lab, "ongoing": ongoing})
        solid = [c for c in coords if not c["ongoing"]]
        dashed = coords[-2:] if coords[-1]["ongoing"] else []
        out.append({
            "key": key, "label": label, "coords": coords, "w": W, "h": H,
            "solid_points": " ".join(f"{c['x']},{c['y']}" for c in solid),
            "dash_points": " ".join(f"{c['x']},{c['y']}" for c in dashed)})
    return out


LANE_FLAT_PCT = 5.0             # within ±5% reads as unchanged


def lanes_for(period, top_n=10):
    """Top outbound + inbound lanes of a month by net revenue with their
    RPK (revenue per kilo), each compared against the SAME lane's RPK
    averaged over the three preceding months on record."""
    data = _load()
    periods = data.get("periods") or {}
    rec = periods.get(str(period))
    if not rec:
        return None
    prior_keys = [k for k in sorted(periods) if k < str(period)][-3:]
    out = {"prior_months": prior_keys}
    for svc, label in (("OB", "outbound"), ("IB", "inbound")):
        rows = []
        for lane, v in sorted((rec.get("lanes") or {}).get(svc, {}).items(),
                              key=lambda kv: -kv[1]["net"])[:top_n]:
            rpk = (v["net"] / v["kilos"]) if v["kilos"] > 0 else None
            # the same lane's RPK in each prior month, averaged
            past = []
            for pk in prior_keys:
                pv = ((periods[pk].get("lanes") or {})
                      .get(svc, {}).get(lane))
                if pv and pv.get("kilos", 0) > 0:
                    past.append(pv["net"] / pv["kilos"])
            prior_rpk = (sum(past) / len(past)) if past else None
            delta = (100.0 * (rpk - prior_rpk) / prior_rpk)                 if rpk is not None and prior_rpk and prior_rpk > 0 else None
            if delta is None:
                trend = "new"
            elif abs(delta) <= LANE_FLAT_PCT:
                trend = "flat"
            else:
                trend = "up" if delta > 0 else "down"
            rows.append({
                "lane": lane.replace("-", " → "), "net": v["net"],
                "kilos": v["kilos"], "shipments": v["shipments"],
                "rpk": rpk, "prior_rpk": prior_rpk,
                "delta_pct": delta, "trend": trend,
                "prior_n": len(past)})
        out[label] = rows
    return out


# Matching a trader against the credit-stop register. The register is
# keyed on the AR customer name, which is frequently TRUNCATED there
# ("SOCIETE ANONYME DES BOISSONS DU" for "…DU CAMEROUN"), so exact
# comparison silently misses real stops. Truncation is prefix-shaped, which
# is what separates a genuine miss from a coincidence: RGSTTC SARL and STBC
# SARL score 0.80 on similarity but are NOT prefixes of one another. On a
# credit control a false flag is as damaging as a missed one, so anything
# short of exact is reported as a LIKELY match, never as fact.
_STOP_MIN_PREFIX = 15


def _stop_key(name):
    return re.sub(r"\s+", " ", str(name or "").strip().upper())


def match_stopped(trader_key, stop_keys):
    """('exact'|'likely'|'', matched_name) for one trader."""
    k = _stop_key(trader_key)
    if not k:
        return "", ""
    if k in stop_keys:
        return "exact", k
    for s in stop_keys:
        if len(s) < _STOP_MIN_PREFIX and len(k) < _STOP_MIN_PREFIX:
            continue
        if k.startswith(s) or s.startswith(k):
            if min(len(k), len(s)) >= _STOP_MIN_PREFIX:
                return "likely", s
    return "", ""


def active_customers(top_n=60, now=None):
    """The top traders BY WEIGHT over the last three complete months, and
    what they are moving in the current month — the point is spotting a big
    trader who has gone quiet (or is trading while on credit stop)."""
    data = _load()
    periods = data.get("periods") or {}
    now = now or datetime.now()
    this_month = now.strftime("%Y-%m")
    complete = [p for p in sorted(periods) if p != this_month]
    last3 = complete[-3:]
    if not last3:
        return None
    ongoing_rec = periods.get(this_month)

    def by_name(rec):
        table = {}
        for acct, c in (rec.get("customers") or {}).items():
            key = (c.get("name") or acct).strip().upper()
            g = table.setdefault(key, {"name": c.get("name") or acct,
                                      "kilos": 0.0})
            g["kilos"] += c.get("kilos", 0.0)
        return table

    maps = {pkey: by_name(periods[pkey]) for pkey in last3}
    cur = by_name(ongoing_rec) if ongoing_rec else {}
    keys = set()
    for m in maps.values():
        keys |= set(m)
    rows = []
    for k in keys:
        kgs = {pkey: round(maps[pkey].get(k, {}).get("kilos", 0.0), 1)
               for pkey in last3}
        avg = round(sum(kgs.values()) / len(last3), 1)
        name = next((maps[pkey][k]["name"] for pkey in last3
                     if k in maps[pkey]), k)
        rows.append({"key": k, "name": name, "avg_kilos": avg,
                     "months": kgs,
                     "current": round(cur.get(k, {}).get("kilos", 0.0), 1)
                     if ongoing_rec else None})
    rows.sort(key=lambda r: -r["avg_kilos"])
    return {"months": last3, "current_period": this_month
            if ongoing_rec else "", "rows": rows[:top_n]}


def fuel_ranking(period, top_n=30):
    """Top customers by FUEL SURCHARGE as a percentage of their weight
    charge, on the products that carry one (D, N, P, T, Y).

    Grouped by customer name, and only customers with a real weight charge
    are ranked — a customer with fuel but no carriage would show an
    infinite percentage."""
    data = _load()
    rec = (data.get("periods") or {}).get(str(period))
    if not rec:
        return None
    fuel = rec.get("fuel") or {}
    base_w, base_f = fuel.get("weight") or 0.0, fuel.get("fuel") or 0.0
    overall = (100.0 * base_f / base_w) if base_w > 0 else None
    rows = []
    for c in (fuel.get("customers") or {}).values():
        if (c.get("weight") or 0) <= 0:
            continue
        pct = 100.0 * c["fuel"] / c["weight"]
        rows.append({"name": c["name"], "weight": c["weight"],
                     "fuel": c["fuel"], "rows": c["rows"], "pct": pct,
                     "delta_pts": (pct - overall) if overall is not None
                     else None})
    rows.sort(key=lambda r: -r["pct"])
    return {"rows": rows[:top_n], "overall": overall,
            "weight": base_w, "fuel": base_f,
            "products": sorted(FUEL_PRODUCTS),
            "customers_total": len(fuel.get("customers") or {})}


def pricing_top(rec, top_n=10):
    """The month's top customers by net revenue (internal DHL included, per
    the owner's choice) with price/kg = WEIGHT CHARGE / kilos, against the
    whole file's average, variance in %.

    Grouped by CUSTOMER NAME, not billing account — one customer (chiefly
    the internal DHL entity) bills under many account numbers and must be
    ONE line, not six."""
    customers = rec.get("customers") or {}
    by_name = {}
    for acct, c in customers.items():
        key = (c.get("name") or acct).strip().upper()
        g = by_name.get(key)
        if g is None:
            g = by_name[key] = {"name": c.get("name") or acct,
                                "accounts": [], "net": 0.0,
                                "weight": 0.0, "kilos": 0.0}
        g["accounts"].append(acct)
        g["net"] += c["net"]
        g["weight"] += c["weight"]
        g["kilos"] += c["kilos"]
    tot_weight = sum(c["weight"] for c in customers.values())
    tot_kilos = sum(c["kilos"] for c in customers.values())
    file_avg = (tot_weight / tot_kilos) if tot_kilos > 0 else None
    top = sorted(by_name.values(), key=lambda g: -g["net"])[:top_n]
    out = []
    for g in top:
        per_kg = (g["weight"] / g["kilos"]) if g["kilos"] > 0 else None
        variance = (100.0 * (per_kg - file_avg) / file_avg) \
            if per_kg is not None and file_avg and file_avg > 0 else None
        out.append({"account": ", ".join(g["accounts"][:3])
                    + (" …" if len(g["accounts"]) > 3 else ""),
                    "name": g["name"], "net": round(g["net"], 2),
                    "kilos": round(g["kilos"], 2),
                    "weight": round(g["weight"], 2), "per_kg": per_kg,
                    "variance_pct": variance})
    return {"rows": out, "file_avg": file_avg}


def dashboard(now=None):
    """Everything the page needs, months ascending. The month equal to the
    CURRENT calendar month is flagged ongoing — its per-day figure divides
    by the billable days elapsed in the data, i.e. a live run-rate."""
    from ..services.ctp_rules import EUR_RATES
    data = _load()
    periods = data.get("periods") or {}
    union_days = _union_ship_days(periods)
    now = now or datetime.now()
    this_month = now.strftime("%Y-%m")
    months = [month_metrics(rec, union_days)
              for _k, rec in sorted(periods.items())]
    for m in months:
        m["ongoing"] = m["period"] == this_month
        try:
            m["label"] = datetime.strptime(m["period"],
                                           "%Y-%m").strftime("%B %Y")
        except ValueError:
            m["label"] = m["period"]
    complete = [m for m in months if not m["ongoing"]]
    ongoing = next((m for m in months if m["ongoing"]), None)
    # The three KPI boxes: the ongoing month against the PRIOR month over
    # the SAME number of elapsed billable days ("compare it to the same
    # number of days in the prior month" — a full-month average would make
    # a two-day-old month look like a collapse). Falls back to the prior
    # month's full average when its daily detail predates this feature.
    kpis = []
    window = None
    prior = complete[-1] if complete else None
    if ongoing and prior:
        window = same_days_window(periods[prior["period"]], union_days,
                                  ongoing["billable_days"])
        # A window that caught no billing is not a baseline — fall back to
        # the prior month's full average rather than divide by nothing.
        if window and not window["net"]:
            window = None
        for key, label in (("rev_per_day", "Revenue / day"),
                           ("rev_per_shipment", "Revenue / shipment"),
                           ("rev_per_kg", "Revenue / kg")):
            cur = ongoing[key]
            base = (window or prior)[key]
            kpis.append({
                "key": key, "label": label, "value": cur, "baseline": base,
                # a negative baseline would invert the sign — suppress it
                "delta_pct": (100.0 * (cur - base) / base)
                if cur is not None and base and base > 0 else None})
    # Pricing defaults to the latest COMPLETE month — a run-rate month a few
    # days in would put noise in the price/kg column. The page offers every
    # month (the ongoing one included) as a selector.
    # The KPI chart compares EQUAL spans: every month cut back to the same
    # billable-day count the running month has reached.
    lfl_months, lfl_target = like_for_like(months, periods, union_days)
    landing = landing_estimate(ongoing, months)
    default = complete[-1] if complete else (months[-1] if months else None)
    # Months stored before the lane/daily detail existed still show their
    # headline figures, but cannot feed the lanes panel or the same-days
    # comparison. Name them so the user knows a re-upload unlocks those.
    stale = [m["label"] for m in months
             if not periods[m["period"]].get("daily")
             or not periods[m["period"]].get("lanes")]
    unmapped = sorted({c for rec in periods.values()
                       for c in (rec.get("unmapped_codes") or [])})
    return {"months": months, "ongoing": ongoing, "kpis": kpis,
            "needs_reupload": stale,
            "unmapped_codes": unmapped,
            "compare": {"prior_label": prior["label"] if prior else "",
                        "window": window},
            "landing": landing,
            "lfl_days": lfl_target,
            "lfl_partial": [m["label"] for m in lfl_months
                            if not m.get("clipped")],
            "graphs": _graph_series(lfl_months),
            "eur_rate": EUR_RATES["XAF"],
            "pricing_period": default["period"] if default else "",
            "pricing": pricing_top(periods[default["period"]])
            if default else {"rows": [], "file_avg": None},
            "lanes": lanes_for(default["period"]) if default else None,
            "fuel": fuel_ranking(default["period"]) if default else None}


def pricing_for(period):
    data = _load()
    rec = (data.get("periods") or {}).get(str(period))
    return pricing_top(rec) if rec else None
