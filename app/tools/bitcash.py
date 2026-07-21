"""BIT & Cash AR — Bank In Transit and Cash Account Receivables tracking,
payment-statement reconciliation and CM01 journal-entry generation.

Two uploads: the BIT file and the Cash AR file (Excel). Each upload replaces
the previous file of its kind; the number of OPEN items counted at upload is
snapshotted per day so the section can disclose today's position and plot the
daily series. The parsed rows are also kept so customer payment statements
(pick-up sheets / remittance evidence) can be reconciled:

  - the statement's detailed invoice references (AWBs) are looked up in the
    Cash AR (Assignment column carries the AWB),
  - its TOTAL is looked up among the BIT amounts (the amount actually paid),
  - each statement gets a sandbox (auto-matches highlighted; the user adjusts
    selections, approves/unapproves),
  - approved sandboxes are written into the CM01_PC journal-entry template
    (one document per sandbox: a 40 line from the BIT + a 15 line per AWB).

"Open" items: when the file carries a status-like column (status / statut /
état / state / open …), rows whose value reads open/ouvert/o/blank count as
open; without such a column every row is an open item (these files are
typically extracts of open items only).
"""
import json
import os
import re
import shutil
import tempfile
import threading
import time
import unicodedata
import uuid
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from threading import Lock

from ..config import BASE_DIR, DATA_DIR, UPLOAD_DIR
from ..services import ai_ocr, excel_reader

STORE_PATH = DATA_DIR / "bitcash.json"
ROWS_PATH = DATA_DIR / "bitcash_rows.json"
FILES_DIR = DATA_DIR / "bitcash" / "files"
RECON_DIR = DATA_DIR / "bitcash" / "recons"
TEMPLATE_PATH = BASE_DIR / "samples" / "cm01_pc_template.xlsx"
KINDS = {"bit": "BIT (Bank In Transit)", "cash": "Cash AR"}
_lock = Lock()

_STATUS_CANDIDATES = ["status", "statut", "etat", "état", "state",
                      "open/closed", "situation"]
_OPEN_WORDS = {"", "open", "ouvert", "ouverte", "o", "en cours", "pending",
               "outstanding", "not cleared", "non solde", "non soldé"}


def _norm(s):
    s = "".join(c for c in unicodedata.normalize("NFKD", str(s or ""))
                if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _counts_from(parsed):
    """Open-item counts from an already-parsed workbook."""
    header = parsed["header"]
    status_col = next(
        (h for h in header
         if any(cand in _norm(h) for cand in _STATUS_CANDIDATES)), None)
    rows = parsed["rows"]
    if not status_col:
        return {"items": len(rows), "open_items": len(rows), "status_col": None}
    open_items = sum(
        1 for r in rows
        if _norm(r["data"].get(status_col)) in _OPEN_WORDS
        or _norm(r["data"].get(status_col)).startswith(("open", "ouvert")))
    return {"items": len(rows), "open_items": open_items,
            "status_col": status_col}


def count_items(path):
    """Parse one BIT / Cash AR file → {items, open_items, status_col}."""
    return _counts_from(excel_reader.read_transactions(path))


def _load():
    if STORE_PATH.exists():
        try:
            return json.loads(STORE_PATH.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"current": {}, "history": {}}


def _save(data):
    STORE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp = STORE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    tmp.replace(STORE_PATH)


def record_upload(kind, path, source, progress=None):
    """Store the newly uploaded file's counts as the CURRENT position for its
    kind (the new file REPLACES the old one as the reference for every
    check), snapshot today's open counts into the daily history — prior-day
    points are never touched, so the graph keeps its evolution — and persist
    the parsed rows for reconciliation + journal-entry generation.

    The workbook is parsed ONCE; ``progress`` (a kwargs callback) receives
    stage / rows_done / rows_total / pct updates for the loading tracker."""
    if kind not in KINDS:
        raise ValueError(f"unknown kind: {kind}")
    tick = progress or (lambda **kw: None)
    tick(stage="reading the workbook", pct=3)
    parsed = excel_reader.read_transactions(path)
    counts = _counts_from(parsed)
    tick(stage="storing the rows", pct=40,
         rows_done=0, rows_total=len(parsed["rows"]))
    _persist_rows(kind, path, parsed=parsed, progress=tick)
    with _lock:
        data = _load()
        prev = (data["current"].get(kind) or {}).get("stored", "")
        data["current"][kind] = {
            **counts, "source": source,
            "stored": Path(path).name,
            "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M")}
        today = datetime.now().strftime("%Y-%m-%d")
        entry = data["history"].get(today, {})
        entry[f"{kind}_open"] = counts["open_items"]
        data["history"][today] = entry
        for key in sorted(data["history"])[:-180]:     # keep ~6 months
            del data["history"][key]
        _save(data)
    if prev and prev != Path(path).name:               # replaced — clean up
        (UPLOAD_DIR / Path(prev).name).unlink(missing_ok=True)
    tick(stage="done", pct=100, rows_done=len(parsed["rows"]),
         rows_total=len(parsed["rows"]))
    return counts


def status():
    """{current: {bit, cash}, history: [(date, {bit_open, cash_open}), …],
    processing: {...}|None, processing_error: {...}|None}.

    A processing flag whose heartbeat stopped (server restarted mid-parse)
    is cleared after 15 minutes and surfaced as an error instead of leaving
    the page refreshing forever."""
    data = _load()
    proc = data.get("processing")
    if proc:
        beat = proc.get("beat") or 0
        if not beat:
            try:
                beat = datetime.strptime(proc.get("started", ""),
                                         "%Y-%m-%d %H:%M").timestamp()
            except ValueError:
                beat = 0
        if time.time() - beat > 900:
            with _lock:
                data = _load()
                if data.get("processing"):
                    data.pop("processing", None)
                    data["processing_error"] = {
                        "at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "message": "Processing was interrupted (the server "
                                   "restarted?) before it finished"}
                    _save(data)
    return {"current": data.get("current", {}),
            "history": sorted(data.get("history", {}).items()),
            "processing": data.get("processing"),
            "processing_error": data.get("processing_error")}


def process_uploads_async(jobs):
    """Parse the uploaded BIT / Cash AR file(s) on a background thread — big
    extracts take longer than a web request may (proxy timeouts), so the
    upload responds instantly and the page shows a LIVE loading tracker
    (stage, rows ingested, % done and estimated time left).

    ``jobs`` = [(kind, stored_path, source_name)].
    """
    with _lock:
        data = _load()
        data["processing"] = {
            "started": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "beat": time.time(),
            "files": [{"kind": k, "label": KINDS[k], "source": s}
                      for k, _p, s in jobs],
            "stage": "queued", "current": "", "pct": 0,
            "rows_done": 0, "rows_total": 0, "eta_seconds": None}
        data.pop("processing_error", None)
        _save(data)

    started = time.time()
    n_files = len(jobs)
    last_write = [0.0]

    def _tracker(file_idx, label):
        def tick(stage=None, pct=None, rows_done=None, rows_total=None):
            # Overall % = this file's share of the whole upload.
            overall = None
            if pct is not None:
                overall = round((file_idx * 100 + min(pct, 100)) / n_files)
            now = time.time()
            # Throttle store writes; never skip stage changes or completion.
            if stage is None and (pct or 0) < 100 \
                    and now - last_write[0] < 0.4:
                return
            last_write[0] = now
            with _lock:
                data = _load()
                proc = data.get("processing")
                if not proc:
                    return
                proc["current"] = label
                proc["beat"] = now
                if stage is not None:
                    proc["stage"] = stage
                if overall is not None:
                    proc["pct"] = overall
                    elapsed = now - started
                    if overall and elapsed > 1:
                        proc["eta_seconds"] = max(
                            0, round(elapsed * (100 - overall) / overall))
                if rows_done is not None:
                    proc["rows_done"] = rows_done
                if rows_total is not None:
                    proc["rows_total"] = rows_total
                _save(data)
        return tick

    def _runner():
        errors = []
        for i, (kind, path, source) in enumerate(jobs):
            try:
                record_upload(kind, path, source,
                              progress=_tracker(i, KINDS[kind]))
            except Exception as exc:  # noqa: BLE001 — surfaced on the page
                errors.append(f"Could not read {source or kind}: "
                              f"{type(exc).__name__}: {exc}")
        with _lock:
            data = _load()
            data.pop("processing", None)
            if errors:
                data["processing_error"] = {
                    "at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "message": " · ".join(errors)}
            _save(data)

    threading.Thread(target=_runner, daemon=True).start()


# --------------------------------------------------------------------------- #
# Parsed rows kept for reconciliation + the journal entry
# --------------------------------------------------------------------------- #
def _atomic(path, text):
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(tmp, path)


def _digits(v):
    return re.sub(r"\D", "", str(v or ""))


def _cellstr(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_float(v):
    try:
        return round(float(str(v).replace(",", "").replace(" ", "") or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _persist_rows(kind, path, parsed=None, progress=None):
    """Parse an uploaded BIT / Cash AR file into the row store used by the
    reconciliation sandboxes and the journal entry. ``parsed`` lets the
    caller reuse an already-read workbook (single parse); ``progress``
    receives rows_done/rows_total/pct ticks every few hundred rows."""
    parsed = parsed or excel_reader.read_transactions(path)
    tick = progress or (lambda **kw: None)
    total_rows = max(len(parsed["rows"]), 1)

    def _row_tick(done):
        if done % 250 == 0 or done == total_rows:
            tick(rows_done=done, rows_total=total_rows,
                 pct=40 + round(58 * done / total_rows))
    header = parsed["header"]
    low = {h: _norm(h) for h in header}

    def col(*cands):
        for cand in cands:
            for h in header:
                if low[h] == cand:
                    return h
        for cand in cands:
            for h in header:
                if cand in low[h]:
                    return h
        return None

    with _lock:
        store = {}
        if ROWS_PATH.exists():
            try:
                store = json.loads(ROWS_PATH.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                store = {}
        if kind == "bit":
            c_gl = col("g/l account")
            c_asg = col("assignment")
            c_amt = col("company code currency value")
            c_date = col("posting date")
            c_ref = col("reference")
            c_txt = col("text")
            c_doc = col("document number")
            c_key = col("posting key")
            rows = []
            for i, r in enumerate(parsed["rows"]):
                d = r["data"]
                rows.append({
                    "id": i,
                    "gl_account": _cellstr(d.get(c_gl)),
                    "assignment": _cellstr(d.get(c_asg)),
                    "amount": _to_float(d.get(c_amt)),
                    "posting_date": _cellstr(d.get(c_date))[:10],
                    "reference": _cellstr(d.get(c_ref)),
                    "text": _cellstr(d.get(c_txt))[:60],
                    "doc_no": _cellstr(d.get(c_doc)),
                    "posting_key": _cellstr(d.get(c_key)),
                    "raw": [_cellstr(d.get(h)) for h in header],
                })
                _row_tick(i + 1)
            store["bit_header"] = list(header)
            store["bit"] = rows
        else:
            c_acct = col("sap acct", "sap account")
            c_asg = col("assignment")
            c_ref = col("reference")
            c_amt = col("amount")
            c_name = col("customer account: name", "customer name", "name")
            c_doc = col("doc. no.", "doc no")
            # Ageing is aged from the DOCUMENT / POSTING date — those candidates
            # come first so col() (first match wins) never falls back onto a
            # due/baseline date. Net-due/due date is deliberately excluded.
            c_date = col("posting date", "pstng date", "pstng dt",
                         "posting dt", "document date", "doc. date",
                         "doc date", "doc dt", "invoice date", "billing date",
                         "bill date", "baseline date", "bline date",
                         "value date", "entry date", "date piece",
                         "date de piece", "date document", "date comptable")
            rows = []
            for i, r in enumerate(parsed["rows"]):
                d = r["data"]
                rows.append({
                    "id": i,
                    "sap_acct": _cellstr(d.get(c_acct)),
                    "awb": _digits(d.get(c_asg)),
                    "assignment": _cellstr(d.get(c_asg)),
                    "reference": _cellstr(d.get(c_ref)),
                    "amount": _to_float(d.get(c_amt)),
                    "customer": _cellstr(d.get(c_name))[:40],
                    "doc_no": _cellstr(d.get(c_doc)),
                    "date": _date_cell_to_str(d.get(c_date)) if c_date else "",
                })
                _row_tick(i + 1)
            store["cash"] = rows
            # Remember which header fed the ageing date (or "" when none was
            # recognised) so the panel can show it — turns a silent
            # everything-undated into a legible diagnostic.
            store["cash_date_col"] = _cellstr(c_date) if c_date else ""
        # Every (re)upload gets a fresh generation stamp: sandboxes remember
        # the generation they matched against, so replaced files can never
        # silently re-point their selected row ids at different invoices.
        store[f"gen_{kind}"] = uuid.uuid4().hex[:12]
        _atomic(ROWS_PATH, json.dumps(store, ensure_ascii=False))


def rows_store():
    empty = {"bit_header": [], "bit": [], "cash": [],
             "gen_bit": "", "gen_cash": "", "cash_date_col": ""}
    if not ROWS_PATH.exists():
        return empty
    try:
        data = json.loads(ROWS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return empty
    for key, default in empty.items():
        data.setdefault(key, default)
    return data


def rows_generation(store=None):
    """The generation stamp of the files currently on record."""
    store = store or rows_store()
    return {"bit": store.get("gen_bit", ""), "cash": store.get("gen_cash", "")}


def _date_cell_to_str(v):
    """Normalise a raw date cell to a stable string for storage so ageing works
    regardless of how the workbook presented it: a real date/datetime cell ->
    ISO; an Excel serial number -> ISO; otherwise the trimmed text. _row_date()
    then parses the stored string when the buckets are computed."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    # A bare Excel serial date (days since 1899-12-30) in a plausible range.
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            n = float(v)
            if 20000 <= n <= 80000:            # ~1954-08 … 2119
                return (datetime(1899, 12, 30)
                        + timedelta(days=int(n))).date().isoformat()
        except (ValueError, OverflowError):
            pass
    return _cellstr(v)[:10]


def _row_date(value):
    """A stored date string -> date, else None. Accepts day-first dot/slash/dash
    ('15.07.2026', '15/07/2026', '15-07-2026'), ISO/year-first with any of
    -./ ('2026-07-15', '2026/07/15', '2026.07.15'), and a bare Excel serial."""
    s = str(value or "").strip()[:10]
    if not s:
        return None
    # Bare Excel serial that reached storage as a number-string.
    if re.fullmatch(r"\d{5}", s):
        try:
            return (datetime(1899, 12, 30)
                    + timedelta(days=int(s))).date()
        except (ValueError, OverflowError):
            return None
    m = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})[./-](\d{4})", s)   # day-first
    y = re.fullmatch(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})", s)   # year-first
    try:
        if m:
            return datetime(int(m.group(3)), int(m.group(2)),
                            int(m.group(1))).date()
        if y:
            return datetime(int(y.group(1)), int(y.group(2)),
                            int(y.group(3))).date()
    except ValueError:
        return None
    return None


def cash_ageing(today=None):
    """The Cash AR classified per ACCOUNT into ageing buckets as of today:
    0-30 days, 30-60 days, above 60 days (item posting date vs today).
    Items whose date could not be read land in 'undated'."""
    today = today or datetime.now().date()
    store = rows_store()
    accounts = {}
    dated = False
    for r in store["cash"]:
        acct = (r.get("sap_acct") or "").strip() or "—"
        g = accounts.setdefault(acct, {
            "account": acct, "customer": "", "b0": 0.0, "b31": 0.0,
            "b61": 0.0, "undated": 0.0, "total": 0.0})
        if not g["customer"] and (r.get("customer") or "").strip():
            g["customer"] = r["customer"].strip()
        amt = r.get("amount") or 0
        d = _row_date(r.get("date"))
        if d is None:
            g["undated"] = round(g["undated"] + amt, 2)
        else:
            dated = True
            age = (today - d).days
            if age <= 30:
                g["b0"] = round(g["b0"] + amt, 2)
            elif age <= 60:
                g["b31"] = round(g["b31"] + amt, 2)
            else:
                g["b61"] = round(g["b61"] + amt, 2)
        g["total"] = round(g["total"] + amt, 2)
    rows = sorted(accounts.values(), key=lambda g: -g["total"])
    totals = {k: round(sum(g[k] for g in rows), 2)
              for k in ("b0", "b31", "b61", "undated", "total")}
    return {"as_of": today.strftime("%d/%m/%Y"), "rows": rows,
            "totals": totals, "has_dates": dated,
            # Which header fed the ageing date ("" = none recognised) — shown on
            # the panel so an all-undated result is explained, not silent.
            "date_col": store.get("cash_date_col", "")}


def search_cash(query, limit=20):
    """Find Cash AR invoices by AWB / reference digits or customer name — used
    by the sandbox to add invoices for clearance manually."""
    q = str(query or "").strip()
    if not q:
        return []
    qd = _digits(q)
    ql = q.lower()
    out = []
    for r in rows_store()["cash"]:
        if (qd and (qd in r["awb"] or qd in _digits(r["reference"]))) \
                or (not qd and ql in r["customer"].lower()):
            out.append(r)
            if len(out) >= limit:
                break
    return out


# --------------------------------------------------------------------------- #
# Reconciliation sandboxes (one per uploaded payment statement)
# --------------------------------------------------------------------------- #
def _recon_path(token):
    return RECON_DIR / f"{token}.json"


def save_recon(rec):
    _atomic(_recon_path(rec["token"]), json.dumps(rec, ensure_ascii=False))
    return rec["token"]


def load_recon(token):
    if not re.fullmatch(r"[0-9a-f]+", token or ""):
        return None
    p = _recon_path(token)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def list_recons():
    """All sandboxes, open/reading first then approved, newest first inside."""
    if not RECON_DIR.exists():
        return []
    out = []
    for p in RECON_DIR.glob("*.json"):
        rec = load_recon(p.stem)
        if rec:
            out.append(rec)
    out.sort(key=lambda r: r.get("uploaded", ""), reverse=True)
    out.sort(key=lambda r: r.get("status") == "approved")
    return out


# A BIT payment rarely lands to the franc (rounded deposits, small bank
# charges) — when no exact match exists, candidates are searched within this
# margin of the evidence total (closest first).
BIT_MATCH_MARGIN = 1000.0

# Short/excess payments beyond this trigger the clarification email to the
# account holder (debit position → payment expected; credit → refund given).
VARIANCE_EMAIL_THRESHOLD = 5000.0


def automatch(statement, slip_total=None, payment_refs=None,
              slip_refs=None):
    """Match a parsed statement against the stored Cash AR + BIT rows.

    The statement TOTAL is always recomputed arithmetically from its lines
    (evidence files often carry a broken formula or a text like
    "TOTAL: 523.571" in the total cell — never trusted); any differing total
    stated in the document is kept as ``stated_total`` for disclosure.

    BIT anchors, strongest first: (1) ``payment_refs`` — bank references
    submitted with the evidence (operator returns) are looked up INSIDE the
    BIT's reference/text/assignment; a unique hit auto-selects; (2) the
    deposit-slip total (what was actually banked); (3) the evidence total,
    within ±BIT_MATCH_MARGIN, closest first.

    Returns (lines, ar_selected, bit_candidates, bit_selected).
    """
    data = rows_store()
    by_awb = {}
    for r in data["cash"]:
        if r["awb"]:
            by_awb.setdefault(r["awb"], []).append(r["id"])
    lines, ar_selected = [], []
    for ln in statement.get("lines", []):
        ref = _digits(ln.get("reference"))
        ids = list(by_awb.get(ref, []))
        if not ids and ref:
            ids = [r["id"] for r in data["cash"]
                   if ref and ref in _digits(r["reference"])]
        lines.append({**ln, "matched_ids": ids})
        ar_selected.extend(ids)

    # Arithmetic total of the evidences (authoritative).
    if lines:
        arith = round(sum(_to_float(ln.get("amount")) or 0 for ln in lines), 2)
        stated = _to_float(statement.get("total"))
        if arith:
            if stated and abs(stated - arith) > 0.005:
                statement["stated_total"] = stated
            statement["total"] = arith

    total = _to_float(statement.get("total"))
    slip = abs(_to_float(slip_total)) if slip_total else 0.0
    refs = list(payment_refs or []) + \
        [ln.get("payment_reference") for ln in lines
         if ln.get("payment_reference")]
    strong, agree, trail = _anchor_hits(refs, slip_refs, data["bit"], slip)
    ref_hits = strong + agree
    targets = [t for t in (slip, abs(total or 0)) if t]
    cands, bit_selected = [], None
    if targets:
        def _dist(row):
            return min(abs(abs(row["amount"]) - t) for t in targets)

        def _rank(row):
            # ZERO variance is the principle: the EXACT banked amount on
            # the deposit slip is the primary suggestion, the exact
            # evidence total second; the ±margin window is only a fallback.
            a = abs(row["amount"])
            if slip and a == slip:
                return (0, 0.0)
            if total and a == abs(total):
                return (1, 0.0)
            return (2, _dist(row))
        near = sorted((r for r in data["bit"]
                       if _dist(r) <= BIT_MATCH_MARGIN),
                      key=lambda r: (*_rank(r), r["id"]))
        cands = [r["id"] for r in near]
    # Reference hits outrank everything: submitted references first, then
    # slip anchors that AGREE on the banked amount; disagreeing slip hits
    # trail at the bottom. Auto-select: a unique submitted-reference hit,
    # else a unique agreeing slip hit.
    if ref_hits:
        cands = ref_hits + [i for i in cands if i not in ref_hits]
        if len(strong) == 1:
            bit_selected = strong[0]
        elif not strong and len(agree) == 1:
            bit_selected = agree[0]
    if trail:
        cands = [i for i in cands if i not in trail] + trail
    if bit_selected is None and targets:
        amounts = {r["id"]: abs(r["amount"]) for r in data["bit"]}
        for t in targets:
            exact = [i for i in cands if amounts.get(i) == t]
            if len(exact) == 1:
                bit_selected = exact[0]
                break
            if exact:          # the same amount more than once — user picks
                break
        if bit_selected is None and len(cands) == 1 \
                and cands[0] not in trail:
            bit_selected = cands[0]     # a lone TRAIL hit is a suggestion,
    return lines, sorted(set(ar_selected)), cands, bit_selected


def _norm_ref(value):
    """Reference normalisation for BIT lookups: alphanumerics, uppercased."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _reference_hits(refs, bit_rows):
    """BIT rows whose reference / text / assignment carry one of the
    submitted payment references (normalised containment, refs >= 4 chars).
    The bank reference is the strongest anchor there is — these hits rank
    before any amount-based candidate."""
    wanted = [_norm_ref(r) for r in (refs or [])]
    wanted = [w for w in wanted if len(w) >= 4]
    if not wanted:
        return []
    hits = []
    for row in bit_rows:
        hay = _norm_ref(f"{row.get('reference', '')} {row.get('text', '')} "
                        f"{row.get('assignment', '')}")
        if any(w in hay for w in wanted):
            hits.append(row["id"])
    return hits


def _slip_refs(info):
    """Machine-read anchors off the deposit slip (its reference + the
    account credited)."""
    info = info or {}
    return [r for r in (info.get("slip_reference", ""),
                        info.get("account_credited", "")) if r]


def _anchor_hits(strong_refs, slip_refs, bit_rows, slip_amount):
    """Two anchor tiers: STRONG = references a person submitted (unique hit
    auto-selects); WEAK = machine-read slip anchors (>= 6 chars after
    normalisation — short serials substring-match unrelated rows). A weak
    hit only carries authority when the row's amount AGREES with the slip
    total; disagreeing weak hits trail as low-rank suggestions.
    Returns (strong, agree, trail)."""
    strong = _reference_hits(strong_refs, bit_rows)
    srcs = [r for r in (slip_refs or []) if len(_norm_ref(r)) >= 6]
    weak = [i for i in _reference_hits(srcs, bit_rows) if i not in strong]
    amounts = {r["id"]: abs(r["amount"]) for r in bit_rows}
    slip_amt = abs(_to_float(slip_amount) or 0)
    agree = [i for i in weak if slip_amt and amounts.get(i) == slip_amt]
    trail = [i for i in weak if i not in agree]
    return strong, agree, trail


def _read_slip_info(slip_path, slip_media, ai_cfg):
    """Full deposit-slip reading via the dedicated extractor → {amount,
    date, bank, depositor, beneficiary, account_credited, slip_reference}
    (Excel slips: amount only). None when unreadable — the slip then stays
    pure evidence."""
    try:
        if slip_media == "excel":
            return {"amount": _read_excel_statement(slip_path)["total"],
                    "date": "", "bank": "", "depositor": "",
                    "beneficiary": "", "account_credited": "",
                    "slip_reference": ""}
        info = ai_ocr.extract_deposit_slip(
            Path(slip_path).read_bytes(), slip_media, ai_cfg)
        return {"amount": _to_float(info.get("amount_xaf")) or None,
                "date": info.get("date", ""),
                "bank": info.get("bank", ""),
                "depositor": info.get("depositor", ""),
                "beneficiary": info.get("beneficiary", ""),
                "account_credited": info.get("account_credited", ""),
                "slip_reference": info.get("slip_reference", "")}
    except Exception:  # noqa: BLE001 — a slip that can't be read is fine
        return None


def _read_slip_total(slip_path, slip_media, ai_cfg):
    """Back-compat wrapper: just the banked amount."""
    return (_read_slip_info(slip_path, slip_media, ai_cfg) or {}) \
        .get("amount")


def create_recon_async(stored_path, media_type, source, uploaded_by,
                       slip_path=None, slip_media=None, slip_source="",
                       extra_slips=None, payment_refs=None,
                       slip_total=None, slip_info=None):
    """Create a sandbox stub and AI-read/match the statement on a background
    thread (a scanned PDF can take a minute — never block the request).
    Copies of the payment advice AND the bank deposit slip(s) are kept with
    the sandbox; the slip's total and any submitted payment references
    anchor the BIT search."""
    token = uuid.uuid4().hex[:12]
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    advice_name = ""
    try:
        advice_name = f"advice_{token}{Path(stored_path).suffix.lower()}"
        shutil.copyfile(stored_path, FILES_DIR / advice_name)
    except OSError:
        advice_name = ""
    slip_rec = None
    if slip_path:
        try:
            slip_name = f"slip_{token}{Path(slip_path).suffix.lower()}"
            shutil.copyfile(slip_path, FILES_DIR / slip_name)
            slip_rec = {"name": slip_name,
                        "source": slip_source or slip_name,
                        "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M")}
        except OSError:
            slip_rec = None
    extras = []
    for n, (p, src) in enumerate(extra_slips or [], start=2):
        try:
            ename = f"slip{n}_{token}{Path(p).suffix.lower()}"
            shutil.copyfile(p, FILES_DIR / ename)
            extras.append({"name": ename, "source": src or ename})
        except OSError:
            continue
    save_recon({"token": token, "status": "reading",
                "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "uploaded_by": uploaded_by or "", "source": source,
                "file": advice_name, "slip": slip_rec,
                "extra_slips": extras,
                "payment_refs": [str(r)[:40] for r in (payment_refs or [])
                                 if str(r).strip()][:20],
                "statement": None, "error": ""})
    from ..config import load_config

    ai_cfg = load_config().get("ai")

    def _runner():
        try:
            if media_type == "excel":
                statement = _read_excel_statement(stored_path)
            else:
                statement = ai_ocr.extract_payment_statement(
                    Path(stored_path).read_bytes(), media_type, ai_cfg)
            # A pre-read that FAILED (empty fields) must not suppress the
            # background read — only a slip_info with content is trusted.
            usable = slip_info if slip_info and any(
                slip_info.get(k) for k in ("amount", "slip_reference",
                                           "bank", "beneficiary")) else None
            info = usable or (_read_slip_info(slip_path, slip_media,
                                              ai_cfg)
                              if slip_path else None)
            resolved_slip_total = slip_total if slip_total \
                else (info or {}).get("amount")
            stub = load_recon(token) or {}
            refs = stub.get("payment_refs", [])
            lines, ar_sel, cands, bit_sel = automatch(
                statement, resolved_slip_total, payment_refs=refs,
                slip_refs=_slip_refs(info))
            statement["lines"] = lines
            all_refs = refs + [ln.get("payment_reference") for ln in lines
                               if ln.get("payment_reference")]
            strong_h, agree_h, _trail_h = _anchor_hits(
                all_refs, _slip_refs(info), rows_store()["bit"],
                resolved_slip_total)
            ref_hits = strong_h + agree_h
            # "Duplicate" = an EXACT amount appears more than once in the
            # BIT; near matches within the margin are choices, not duplicates.
            bit_amounts = {r["id"]: abs(r["amount"])
                           for r in rows_store()["bit"]}
            tot = abs(_to_float(statement.get("total")) or 0)
            slp = abs(_to_float(resolved_slip_total) or 0)
            exact_n = max(
                sum(1 for i in cands if bit_amounts.get(i) == t)
                for t in (tot, slp)) if (tot or slp) else 0
            rec = load_recon(token) or {}
            rec.update({
                "status": "open", "statement": statement,
                "slip_total": resolved_slip_total,
                "slip_info": info,
                "rows_gen": rows_generation(),
                "ar_selected": ar_sel, "bit_candidates": cands,
                "bit_selected": bit_sel, "ref_hits": ref_hits,
                "bit_duplicate": exact_n > 1, "error": ""})
            save_recon(rec)
        except Exception as exc:  # noqa: BLE001 — surface, never crash silently
            rec = load_recon(token) or {"token": token}
            rec.update({"status": "error",
                        "error": f"{type(exc).__name__}: {exc}"})
            save_recon(rec)

    threading.Thread(target=_runner, daemon=True).start()
    return token


def rematch(token):
    """Re-run the automatic matching of an OPEN sandbox against the files
    currently on record (used after the BIT / Cash AR files were replaced).
    Selections, candidates and the generation stamp are rebuilt from the
    stored statement; a stale manual plug is dropped."""
    rec = load_recon(token)
    if not rec or rec.get("status") != "open" or not rec.get("statement"):
        return None
    statement = rec["statement"]
    lines, ar_sel, cands, bit_sel = automatch(
        statement, rec.get("slip_total"),
        payment_refs=rec.get("payment_refs", []),
        slip_refs=_slip_refs(rec.get("slip_info")))
    statement["lines"] = lines
    all_refs = rec.get("payment_refs", []) + \
        [ln.get("payment_reference") for ln in lines
         if ln.get("payment_reference")]
    strong_h, agree_h, _trail_h = _anchor_hits(
        all_refs, _slip_refs(rec.get("slip_info")), rows_store()["bit"],
        rec.get("slip_total"))
    rec["ref_hits"] = strong_h + agree_h
    bit_amounts = {r["id"]: abs(r["amount"]) for r in rows_store()["bit"]}
    tot = abs(_to_float(statement.get("total")) or 0)
    slp = abs(_to_float(rec.get("slip_total")) or 0)
    exact_n = max(sum(1 for i in cands if bit_amounts.get(i) == t)
                  for t in (tot, slp)) if (tot or slp) else 0
    rec.update({"statement": statement, "ar_selected": ar_sel,
                "bit_candidates": cands, "bit_selected": bit_sel,
                "bit_duplicate": exact_n > 1,
                "rows_gen": rows_generation()})
    rec.pop("plug", None)          # computed for the old difference
    save_recon(rec)
    return rec


_REF_CANDS = ("awb", "waybill", "reference", "invoice", "ref")
_AMT_CANDS = ("amount", "montant", "value")


def _scan_statement_rows(path):
    """Raw fallback for evidence sheets the table reader mis-headers (e.g. an
    Excel table's phantom 'Column1/Column2' row above the real
    'Date / Waybill / Amount (Xaf)' header, decorative title rows, or a text
    'TOTAL: …' cell). Scans every sheet for the first row carrying BOTH a
    reference-like and an amount-like heading, then reads the lines below it
    (rows without a reference — like the TOTAL row — are skipped)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            header_at = ref_col = amt_col = None
            for i, row in enumerate(rows):
                normed = [_norm(v) for v in row]
                rc = next((j for j, v in enumerate(normed)
                           if v and any(c in v for c in _REF_CANDS)), None)
                ac = next((j for j, v in enumerate(normed)
                           if v and any(c in v for c in _AMT_CANDS)), None)
                if rc is not None and ac is not None and rc != ac:
                    header_at, ref_col, amt_col = i, rc, ac
                    break
            if header_at is None:
                continue
            lines = []
            for row in rows[header_at + 1:]:
                ref = _digits(row[ref_col]) if ref_col < len(row) else ""
                amt = _to_float(row[amt_col]) if amt_col < len(row) else 0.0
                if ref and amt:
                    lines.append({"reference": ref, "amount": amt,
                                  "description": ""})
            if lines:
                return lines
        return []
    finally:
        wb.close()


def _read_excel_statement(path):
    """A typed (Excel) payment statement: reference + amount columns. Falls
    back to a raw row scan when the table reader finds no usable lines."""
    parsed = excel_reader.read_transactions(path)
    header = parsed["header"]
    low = {h: _norm(h) for h in header}

    def col(*cands):
        for cand in cands:
            for h in header:
                if cand in low[h]:
                    return h
        return None

    c_ref = col(*_REF_CANDS)
    c_amt = col(*_AMT_CANDS)
    c_desc = col("consignor", "customer", "client", "name", "description")
    c_pref = col("payment reference", "payment ref", "pay ref", "bank ref")
    lines = []
    for r in parsed["rows"]:
        d = r["data"]
        ref = _digits(d.get(c_ref)) if c_ref else ""
        amt = _to_float(d.get(c_amt)) if c_amt else 0.0
        if ref and amt:
            lines.append({"reference": ref, "amount": amt,
                          "payment_reference": _cellstr(d.get(c_pref))[:40]
                          if c_pref else "",
                          "description": _cellstr(d.get(c_desc))[:40]
                          if c_desc else ""})
    if not lines:
        lines = _scan_statement_rows(path)
    if not lines:
        raise ValueError(
            "no AWB/amount lines could be read — the file needs a "
            "'Waybill / AWB / Reference' column and an 'Amount' column")
    return {"label": Path(path).stem, "date": "",
            "total": round(sum(l["amount"] for l in lines), 2), "lines": lines}


def recon_view(recon):
    """Resolve a sandbox's rows for the template, plus the reconciliation
    difference, the manual plug and the residual left after it.

    Reconciliations are FIXED IN TIME: an approved sandbox reads from the
    row snapshot frozen at approval, never from the live files. An open
    sandbox whose BIT / Cash AR files were replaced since it was matched is
    flagged ``stale`` and resolves NOTHING (re-run the matching) — row ids
    from an old file must never be read against a new one."""
    data = rows_store()
    statement = recon.get("statement") or {}
    total = abs(_to_float(statement.get("total")) or 0)
    frozen = recon.get("frozen")
    stale = False
    matched_ids = {i for ln in statement.get("lines", [])
                   for i in ln.get("matched_ids", [])}
    ref_hit_ids = set(recon.get("ref_hits", []))
    if frozen:
        ar_rows = frozen.get("ar_rows", [])
        bit_row = frozen.get("bit_row")
        cands = [{**bit_row,
                  "delta": round(abs(bit_row["amount"]) - total, 2),
                  "by_ref": bit_row.get("id") in ref_hit_ids}] \
            if bit_row else []
    else:
        rec_gen = recon.get("rows_gen") or {"bit": "", "cash": ""}
        stale = recon.get("status") in ("open", "approved") \
            and rec_gen != rows_generation(data)
        if stale:
            ar_rows, cands, bit_row = [], [], None
        else:
            bit_by_id = {r["id"]: r for r in data["bit"]}
            cash_by_id = {r["id"]: r for r in data["cash"]}
            ar_rows = [cash_by_id[i] for i in recon.get("ar_selected", [])
                       if i in cash_by_id]
            cands = [{**bit_by_id[i],
                      "delta": round(abs(bit_by_id[i]["amount"]) - total, 2),
                      "by_ref": i in ref_hit_ids}
                     for i in recon.get("bit_candidates", [])
                     if i in bit_by_id]
            sel = recon.get("bit_selected")
            bit_row = bit_by_id.get(sel) if sel is not None else None
    ar_total = round(sum(r["amount"] for r in ar_rows), 2)
    bit_amount = abs(bit_row["amount"]) if bit_row else 0.0
    plug = recon.get("plug") or None
    plug_amount = round(_to_float((plug or {}).get("amount")) or 0, 2)
    difference = round(ar_total - bit_amount, 2)
    return {"ar_rows": ar_rows, "auto_matched_ids": matched_ids,
            "bit_candidates": cands, "bit_row": bit_row,
            "ar_total": ar_total, "bit_amount": bit_amount,
            "difference": difference, "plug": plug,
            "plug_amount": plug_amount,
            "residual": round(difference - plug_amount, 2),
            "margin": BIT_MATCH_MARGIN,
            "stale": stale, "frozen": bool(frozen),
            "frozen_at": (frozen or {}).get("at", ""),
            "slip_total": recon.get("slip_total"),
            # Variance escalation: only meaningful once a BIT line is chosen.
            "needs_email": bool(bit_row)
            and abs(difference) > VARIANCE_EMAIL_THRESHOLD,
            "email_threshold": VARIANCE_EMAIL_THRESHOLD,
            "variance_email": recon.get("variance_email"),
            "slip": recon.get("slip"), "advice": recon.get("file", ""),
            "slip_info": recon.get("slip_info"),
            "extra_slips": recon.get("extra_slips", []),
            "payment_refs": recon.get("payment_refs", [])}


def set_slip(token, stored_name, source, total=None, info=None):
    """Attach the bank deposit slip to a sandbox (any non-error state) —
    the evidence that pins the exact payment on the bank statement. When
    its total could be read it is stored too, so a re-match anchors the
    BIT search on the banked amount; the full slip reading (bank,
    depositor, beneficiary, account, reference) rides along."""
    rec = load_recon(token)
    if not rec or rec.get("status") not in ("open", "approved"):
        return None
    rec["slip"] = {"name": stored_name, "source": source or stored_name,
                   "uploaded": datetime.now().strftime("%Y-%m-%d %H:%M")}
    if total:
        rec["slip_total"] = round(_to_float(total), 2)
    if info:
        rec["slip_info"] = info
    save_recon(rec)
    return rec


def set_variance_email(token, to_addr, mode, eml_name=""):
    """Remember that the clarification email for this sandbox's variance was
    sent (smtp) or prepared (eml)."""
    rec = load_recon(token)
    if not rec:
        return None
    rec["variance_email"] = {
        "to": to_addr, "mode": mode, "eml": eml_name,
        "at": datetime.now().strftime("%Y-%m-%d %H:%M")}
    save_recon(rec)
    return rec


def set_plug(token, amount, account="", note=""):
    """Record (or clear, when the amount is 0/blank) the manual plug that
    absorbs a remaining reconciliation difference so the journal entry
    balances. Only while the sandbox is open."""
    rec = load_recon(token)
    if not rec or rec.get("status") not in ("open",):
        return None
    amt = _to_float(amount)
    if not amt:
        rec.pop("plug", None)
    else:
        rec["plug"] = {"amount": round(amt, 2),
                       "account": str(account or "").strip(),
                       "note": str(note or "").strip()}
    save_recon(rec)
    return rec


def set_selection(token, ar_ids, bit_id):
    """Returns the updated record, None when not open, or "stale" when the
    BIT / Cash AR files were replaced since this sandbox was matched — ids
    posted against the old files must not be applied to the new ones."""
    rec = load_recon(token)
    if not rec or rec.get("status") not in ("open",):
        return None
    data = rows_store()
    if (rec.get("rows_gen") or {"bit": "", "cash": ""}) \
            != rows_generation(data):
        return "stale"
    valid_cash = {r["id"] for r in data["cash"]}
    valid_bit = {r["id"] for r in data["bit"]}
    rec["ar_selected"] = sorted({int(i) for i in ar_ids
                                 if str(i).lstrip("-").isdigit()
                                 and int(i) in valid_cash})
    rec["bit_selected"] = (int(bit_id) if str(bit_id).lstrip("-").isdigit()
                           and int(bit_id) in valid_bit else None)
    save_recon(rec)
    return rec


def set_status(token, approved, by):
    """Approve (freezing the selected rows in time) or reopen a sandbox.
    Returns the record, None when not possible, or "stale" when approval is
    refused because the files were replaced since matching."""
    rec = load_recon(token)
    if not rec or rec.get("status") not in ("open", "approved"):
        return None
    if approved:
        data = rows_store()
        if (rec.get("rows_gen") or {"bit": "", "cash": ""}) \
                != rows_generation(data):
            return "stale"
        # Freeze the resolved rows: the approval is fixed in time and can
        # never change when newer BIT / Cash AR files are uploaded.
        cash_by_id = {r["id"]: r for r in data["cash"]}
        bit_by_id = {r["id"]: r for r in data["bit"]}
        sel = rec.get("bit_selected")
        rec["frozen"] = {
            "ar_rows": [dict(cash_by_id[i])
                        for i in rec.get("ar_selected", [])
                        if i in cash_by_id],
            "bit_row": dict(bit_by_id[sel]) if sel in bit_by_id else None,
            "rows_gen": rows_generation(data),
            "at": datetime.now().strftime("%Y-%m-%d %H:%M")}
        rec["status"] = "approved"
        rec["approved_by"] = by or ""
        rec["approved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    else:
        rec["status"] = "open"
        rec.pop("approved_by", None)
        rec.pop("approved_at", None)
        rec.pop("frozen", None)     # editing resumes against the live files
    save_recon(rec)
    return rec


def delete_recon(token):
    rec = load_recon(token)
    if rec and rec.get("status") != "approved":
        _recon_path(token).unlink(missing_ok=True)
        return True
    return False


# --------------------------------------------------------------------------- #
# CM01 journal-entry generation (approved sandboxes only)
# --------------------------------------------------------------------------- #
_MANUAL = "MANUAL RECIEPT "        # the template's exact wording (sic)


def _compact_date(d):
    """23.06.2026 -> 23626, exactly as the team's template writes dates."""
    return int(f"{d.day}{d.month}{d.strftime('%y')}")


def build_journal(out_path, now=None):
    """Write every APPROVED sandbox into the CM01 template as one document
    each (LI. 1..N), mirroring the AWB detail on BOTH sides: per selected
    Cash AR invoice a posting-key-40 bank line AND a posting-key-15 customer
    line with the SAME individual amount, the AWB in the Ref/Doc.Nr column
    of both. A manual plug adds a balanced pair (BIT G/L ↔ plug G/L) that
    returns the BIT G/L to the actually banked amount. Returns
    {count, lines, name, tokens} or None when nothing is approved/complete."""
    import openpyxl

    from openpyxl.styles import Font

    approved = [r for r in list_recons() if r.get("status") == "approved"
                and r.get("bit_selected") is not None and r.get("ar_selected")]
    if not approved:
        return None
    gen_now = rows_generation()
    arial = Font(name="Arial", size=10)
    now = now or datetime.now()
    data = rows_store()
    bit_by_id = {r["id"]: r for r in data["bit"]}
    cash_by_id = {r["id"]: r for r in data["cash"]}

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    old_tab = next(n for n in wb.sheetnames if n.startswith("CM01_"))
    new_tab = f"CM01_{now.strftime('%d.%m.%Y')}_PC_001"
    ws = wb[old_tab]
    ws.title = new_tab
    # Renaming does NOT rewrite formulas — patch every reference by hand.
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = cell.value.replace(old_tab, new_tab)

    compact = _compact_date(now)
    r = 4
    lines = 0
    li = 0
    bank_rows = []
    written = []
    skipped = 0
    for rec in approved:
        frozen = rec.get("frozen") or {}
        if frozen.get("bit_row") and frozen.get("ar_rows"):
            # Approved = fixed in time: always write the frozen snapshot.
            bit = frozen["bit_row"]
            ars = frozen["ar_rows"]
        elif (rec.get("rows_gen") or {"bit": "", "cash": ""}) == gen_now:
            bit = bit_by_id.get(rec["bit_selected"])
            ars = [cash_by_id[i] for i in rec["ar_selected"]
                   if i in cash_by_id]
        else:
            # No snapshot and the files were replaced since matching — row
            # ids would point at the WRONG invoices. Never write those.
            skipped += 1
            continue
        if not bit or not ars:
            skipped += 1
            continue
        li += 1
        written.append(rec)
        bank_rows.append(bit)
        # AWB split: the bank side mirrors the Cash AR one-for-one — a 40
        # line and a 15 line per AWB, each pair carrying the SAME individual
        # amount and the AWB in the Ref/Doc.Nr column.
        entries = []
        for a in ars:
            awb = a["awb"] or a["assignment"]
            entries.append((bit["gl_account"], 40, a["amount"],
                            bit["assignment"], awb))
            entries.append((a["sap_acct"], 15, a["amount"], awb, awb))
        # Manual plug: a balanced PAIR — short payment (plug > 0) credits
        # the BIT G/L back down to the amount actually banked and debits the
        # plug G/L; an excess payment does the reverse.
        plug = rec.get("plug") or {}
        plug_amt = _to_float(plug.get("amount"))
        if plug_amt and plug.get("account"):
            note = plug.get("note") or "DIFFERENCE PLUG"
            entries.append((bit["gl_account"], 50 if plug_amt > 0 else 40,
                            abs(plug_amt), bit["assignment"], _MANUAL))
            entries.append((plug["account"], 40 if plug_amt > 0 else 50,
                            abs(plug_amt), note, _MANUAL))
        for account, key, amount, assignment, doc_nr in entries:
            ws.cell(row=r, column=1, value=li)                       # LI.
            ws.cell(row=r, column=2, value="CM01")                   # Comp code
            ws.cell(row=r, column=3, value="DZ")                     # Doc type
            ws.cell(row=r, column=4, value=compact)                  # Post date
            ws.cell(row=r, column=5, value=compact)                  # Doc date
            ws.cell(row=r, column=6, value="XAF")                    # Currency
            ref = ws.cell(row=r, column=7)                           # Doc. Nr
            ref.value = str(doc_nr)
            ref.number_format = "@"
            ws.cell(row=r, column=8, value=_MANUAL)                  # Header txt
            # column 9 (Tax calculation) stays empty
            acct = ws.cell(row=r, column=10)                         # Account
            acct.value = int(account) if str(account).isdigit() else account
            ws.cell(row=r, column=11, value=key)                     # Post key
            ws.cell(row=r, column=12, value=amount)                  # Amount
            asg = ws.cell(row=r, column=14)                          # Assignment
            asg.value = str(assignment)
            asg.number_format = "@"
            for col in range(1, 15):        # uniform Arial 10, no stray colours
                ws.cell(row=r, column=col).font = arial
            r += 1
            lines += 1

    if not written:
        return {"count": 0, "lines": 0, "name": "", "tokens": [],
                "skipped": skipped}

    # BANK DETAILS: the raw BIT line of each payment in the journal.
    bd_name = next((n for n in wb.sheetnames
                    if n.strip().upper() == "BANK DETAILS"), None)
    if bd_name:
        bd = wb[bd_name]
        for i, bit in enumerate(bank_rows, start=2):
            for c, val in enumerate(bit.get("raw", []), start=1):
                cell = bd.cell(row=i, column=c,
                               value=(float(val)
                                      if re.fullmatch(r"-?\d+\.?\d*", val)
                                      and c not in (9,) else val) if val
                               else None)
                cell.font = arial

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    name = f"CM01_PC TEMPLATE_{now.strftime('%d.%m.%y')}_PC.xlsx"
    for rec in written:
        rec["generated_at"] = now.strftime("%Y-%m-%d %H:%M")
        save_recon(rec)
    return {"count": len(bank_rows), "lines": lines, "name": name,
            "tokens": [rec["token"] for rec in written],
            "skipped": skipped}


# --------------------------------------------------------------------------- #
# Journal pack: advice copies + evidences Excel + the journal, zipped
# --------------------------------------------------------------------------- #
def _safe_name(s):
    return re.sub(r"[^A-Za-z0-9._ -]+", "_", str(s or "")).strip() or "item"


def _evidence_xlsx(out_path, rec):
    """The parsed evidence lines of one sandbox as a clean Excel sheet —
    the 'Excel file copy' of the journal pack."""
    import xlsxwriter

    view = recon_view(rec)
    s = rec.get("statement") or {}
    wb = xlsxwriter.Workbook(str(out_path))
    ws = wb.add_worksheet("Evidences")
    f_t = wb.add_format({"bold": True, "font_size": 12, "font_name": "Arial"})
    f_b = wb.add_format({"font_name": "Arial", "font_size": 10})
    f_h = wb.add_format({"font_name": "Arial", "font_size": 10, "bold": True,
                         "bg_color": "#0b2545", "font_color": "white",
                         "border": 1})
    f_c = wb.add_format({"font_name": "Arial", "font_size": 10, "border": 1})
    f_n = wb.add_format({"font_name": "Arial", "font_size": 10, "border": 1,
                         "num_format": "#,##0"})
    f_tot = wb.add_format({"font_name": "Arial", "font_size": 10,
                           "bold": True, "border": 1, "num_format": "#,##0",
                           "bg_color": "#eef3fb"})
    for i, w in enumerate((22, 34, 16, 18)):
        ws.set_column(i, i, w)
    ws.write("A1", f"Payment evidences — {s.get('label') or rec.get('source')}",
             f_t)
    ws.write("A2", f"Uploaded {rec.get('uploaded')} by "
                   f"{rec.get('uploaded_by') or '—'} · source: "
                   f"{rec.get('source')}", f_b)
    row = 3
    for c, h in enumerate(("Reference (AWB)", "Description",
                           "Amount (XAF)", "Found in Cash AR?")):
        ws.write(row, c, h, f_h)
    for ln in s.get("lines", []):
        row += 1
        ws.write(row, 0, ln.get("reference", ""), f_c)
        ws.write(row, 1, ln.get("description", "") or "—", f_c)
        ws.write_number(row, 2, float(ln.get("amount") or 0), f_n)
        ws.write(row, 3, "yes" if ln.get("matched_ids") else "NO", f_c)
    row += 1
    ws.write(row, 0, "TOTAL OF EVIDENCES", f_tot)
    ws.write(row, 1, "", f_tot)
    ws.write_number(row, 2, float(s.get("total") or 0), f_tot)
    ws.write(row, 3, "", f_tot)
    row += 2
    bit = view.get("bit_row") or {}
    facts = [
        ("BIT payment selected",
         f"G/L {bit.get('gl_account')} · {bit.get('assignment')} · "
         f"{abs(bit.get('amount') or 0):,.0f} XAF · "
         f"{bit.get('posting_date') or ''}" if bit else "— none selected —"),
        ("Reconciliation difference", f"{view['difference']:,.0f} XAF"),
    ]
    if s.get("stated_total"):
        facts.insert(0, ("Total stated in the file (ignored)",
                         f"{s['stated_total']:,.0f} XAF"))
    plug = view.get("plug")
    if plug:
        facts.append(("Manual plug", f"{plug['amount']:,.0f} XAF on G/L "
                      f"{plug['account']} ({plug.get('note') or 'no note'})"))
    ve = rec.get("variance_email")
    if ve:
        facts.append(("Variance clarification email",
                      f"{ve.get('mode')} to {ve.get('to')} on {ve.get('at')}"))
    for label, value in facts:
        ws.write(row, 0, label, f_b)
        ws.write(row, 1, value, f_b)
        row += 1
    wb.close()
    return Path(out_path)


def build_pack(out_zip, journal_path, journal_name, tokens):
    """Zip the audit pack: the CM01 journal at the root plus, per approved
    sandbox, a folder with the payment-advice copy, the deposit slip (when
    attached) and the parsed-evidences Excel."""
    out_zip = Path(out_zip)
    out_zip.parent.mkdir(parents=True, exist_ok=True)
    tmp = Path(tempfile.mkdtemp(prefix="cm01pack_"))
    try:
        with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(journal_path, arcname=journal_name)
            for token in tokens:
                rec = load_recon(token)
                if not rec:
                    continue
                s = rec.get("statement") or {}
                folder = _safe_name(f"{s.get('label') or rec.get('source')}"
                                    )[:40] + f"_{token[:6]}"
                advice = rec.get("file", "")
                if advice and (FILES_DIR / Path(advice).name).exists():
                    zf.write(FILES_DIR / Path(advice).name,
                             arcname=f"{folder}/Payment advice - "
                                     f"{_safe_name(rec.get('source') or advice)}")
                slip = (rec.get("slip") or {}).get("name", "")
                if slip and (FILES_DIR / Path(slip).name).exists():
                    zf.write(FILES_DIR / Path(slip).name,
                             arcname=f"{folder}/Deposit slip - "
                                     f"{_safe_name((rec.get('slip') or {}).get('source') or slip)}")
                ev = tmp / f"Evidences_{token}.xlsx"
                _evidence_xlsx(ev, rec)
                zf.write(ev, arcname=f"{folder}/Evidences_{token}.xlsx")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return out_zip
