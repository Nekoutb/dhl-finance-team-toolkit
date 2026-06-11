"""Read an Excel export and locate the transaction table inside it.

Bank / telco exports usually have a few title/metadata rows above the real
table header. This module auto-detects the header row, captures the rows above
it as ``metadata``, and returns each transaction as a dict keyed by column name.

Refine ``KEY_CANDIDATES`` once we see a real Orange Cameroun file — the rest of
the heuristics adapt automatically.
"""
from pathlib import Path

import openpyxl

# Column-name fragments we treat as the "customer key" (the value we remember a
# customer's name against). First match wins; ordered roughly by usefulness.
KEY_CANDIDATES = [
    "msisdn", "numero", "numéro", "telephone", "téléphone", "phone",
    "compte", "account", "wallet", "payer", "sender", "expediteur",
    "expéditeur", "beneficiaire", "bénéficiaire", "client", "customer",
    "mobile", "tel", "number",
]


def _clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _is_blank_row(row):
    return all(cell in (None, "") for cell in row)


def _detect_header(grid, max_scan=40):
    """Return the index of the most likely header row."""
    scan = min(len(grid), max_scan)
    for i in range(scan):
        cells = [c for c in grid[i] if c not in (None, "")]
        if len(cells) < 3:
            continue
        text_ratio = sum(1 for c in cells if isinstance(c, str)) / len(cells)
        nxt = grid[i + 1] if i + 1 < len(grid) else []
        nxt_cells = [c for c in nxt if c not in (None, "")]
        # A header row is mostly text and is followed by a populated data row.
        if text_ratio >= 0.6 and len(nxt_cells) >= max(3, int(len(cells) * 0.5)):
            return i
    if not grid:
        return 0
    # Fallback: the row with the most populated cells.
    return max(range(scan), key=lambda i: sum(1 for c in grid[i] if c not in (None, "")))


def _build_header(header_row):
    last = -1
    for j, cell in enumerate(header_row):
        if cell not in (None, ""):
            last = j
    raw_names = []
    for j in range(last + 1):
        name = _clean(header_row[j])
        raw_names.append(str(name) if name not in (None, "") else f"Column {j + 1}")
    # De-duplicate so identical headers don't collide as dict keys.
    counts, header = {}, []
    for name in raw_names:
        if name in counts:
            counts[name] += 1
            header.append(f"{name} ({counts[name]})")
        else:
            counts[name] = 1
            header.append(name)
    return header


def _read_xls_grid(path):
    """Read a legacy Excel 97-2003 (.xls) file into the same grid shape."""
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    grid = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                except (ValueError, OverflowError):
                    row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                # Excel stores ints as floats; restore clean integers.
                row.append(int(cell.value) if cell.value == int(cell.value)
                           else cell.value)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                row.append(None)
            else:
                row.append(cell.value)
        grid.append(row)
    return grid, sheet.name


def read_transactions(path, max_scan=40):
    """Parse an .xlsx/.xlsm/.xls export into header, metadata and rows."""
    path = Path(path)
    if path.suffix.lower() == ".xls":
        grid, sheet_name = _read_xls_grid(path)
    else:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        grid = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheet_name = sheet.title
        workbook.close()

    while grid and _is_blank_row(grid[-1]):
        grid.pop()
    if not grid:
        return {"sheet": sheet_name, "header": [], "metadata": [], "rows": []}

    header_idx = _detect_header(grid, max_scan)
    header = _build_header(grid[header_idx])

    metadata = []
    for i in range(header_idx):
        cells = [str(_clean(c)) for c in grid[i] if c not in (None, "")]
        if cells:
            metadata.append("   ".join(cells))

    rows = []
    for i in range(header_idx + 1, len(grid)):
        raw = grid[i]
        if _is_blank_row(raw):
            continue
        record = {}
        for j, name in enumerate(header):
            record[name] = _clean(raw[j]) if j < len(raw) else ""
        rows.append({"index": i, "data": record})

    return {
        "sheet": sheet_name,
        "header": header,
        "metadata": metadata,
        "rows": rows,
        "header_index": header_idx,
    }


def guess_key_column(header):
    """Pick the column whose value identifies the customer (phone/account)."""
    lowered = [h.lower() for h in header]
    for candidate in KEY_CANDIDATES:
        for j, name in enumerate(lowered):
            if candidate in name:
                return header[j]
    return header[0] if header else None
