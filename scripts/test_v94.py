"""v9.4 — cheque↔AR columns, CtP daily stats + graphs, dashboard KPIs,
Orange Money AR columns + upload history, Account Stop.

All stores are isolated to a temp dir; the CtP store itself is pointed at the
temp dir so the fabricated analysis is the only one the app sees."""
import json
import re
import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from testutil import smtp_guard  # noqa: E402
smtp_guard()

from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402
from app.services import auth, cheques, customers  # noqa: E402
from app.tools import account_stop, bank, bitcash  # noqa: E402
from app.tools import ongoing_ctp as ctp  # noqa: E402
from app.tools import orange_cameroun as orange  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
SAMPLE = ROOT / "samples" / "orange_statement_sample.xlsx"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_tmp = Path(tempfile.mkdtemp(prefix="v94_"))

# Hermetic stores — nothing in data/ is read or written by this test.
ctp.STORE_DIR = _tmp / "ctp"
ctp.STATS_PATH = _tmp / "ctp_daily_stats.json"
cheques.BATCH_DIR = _tmp / "cheq_batches"
cheques.STATS_PATH = _tmp / "cheque_daily_stats.json"
bank.STORE_DIR = _tmp / "bank"
orange.HISTORY_DIR = _tmp / "orange"
customers._PATH = _tmp / "customers.json"
bitcash.RECON_DIR = _tmp / "recons"
bitcash.STORE_PATH = _tmp / "bitcash.json"
bitcash.ROWS_PATH = _tmp / "bitcash_rows.json"
bitcash.FILES_DIR = _tmp / "bitfiles"
bitcash.UPLOAD_DIR = _tmp / "bit_uploads"
bitcash.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def check(label, cond):
    print(f"[{'OK ' if cond else 'FAIL'}] {label}")
    if not cond:
        raise SystemExit(1)


client = TestClient(main.app, follow_redirects=False)

# === 1. CtP daily stats maths ================================================
tiny = {"customers": [], "summary": {"total_ar": 150.0},
        "invoices": [
            {"kind": "invoice", "amount": 100.0, "days_overdue": 70},
            {"kind": "invoice", "amount": 100.0, "days_overdue": 10},
            {"kind": "credit", "amount": -50.0, "days_overdue": 0}]}
s = ctp.compute_daily_stats(tiny)
check("pct>60 = 50.0 (credits excluded from the base)",
      s["pct_over_60"] == 50.0 and s["pct_over_90"] == 0.0)
check("empty result -> 0.0 not division error",
      ctp.compute_daily_stats({"customers": [], "invoices": [],
                               "summary": {}})["pct_over_60"] == 0.0)

hist = ctp.record_daily_stats(tiny, day="2026-07-01")
check("day override recorded", "2026-07-01" in hist)
for i in range(200):                       # prune keeps ~180 days
    hist = ctp.record_daily_stats(tiny, day=f"2027-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}")
check("history pruned to 180 days", len(hist) <= 180)
ctp.STATS_PATH.unlink()                    # clean slate for the page tests

# === 2. Account Stop empty state (no analysis on file yet) ==================
r = client.get("/tools/account-stop")
check("account-stop page 200 (empty)", r.status_code == 200)
check("no-analysis empty state shown", "No AR analysis on file yet" in r.text)

# === 3. Fabricated latest analysis (ACME stopped, BETA ok) ==================
RESULT = {
    "as_of": date(2026, 7, 16), "mapping": {}, "header": [],
    "default_rank": 60,
    "invoices": [
        {"kind": "invoice", "customer": "ACME LOGISTICS SARL",
         "account": "1004000001", "invoice_no": "INV-1", "amount": 300000.0,
         "days_overdue": 75, "invoice_date": None, "due_date": None},
        {"kind": "invoice", "customer": "ACME LOGISTICS SARL",
         "account": "1004000001", "invoice_no": "INV-2", "amount": 100000.0,
         "days_overdue": 95, "invoice_date": None, "due_date": None},
        {"kind": "invoice", "customer": "BETA TRADING SA",
         "account": "1004000002", "invoice_no": "INV-3", "amount": 100000.0,
         "days_overdue": 5, "invoice_date": None, "due_date": None},
        {"kind": "credit", "customer": "BETA TRADING SA",
         "account": "1004000002", "invoice_no": "CR-1", "amount": -50000.0,
         "days_overdue": 0, "invoice_date": None, "due_date": None},
    ],
    "customers": [
        {"key": "1004000001", "customer": "ACME LOGISTICS SARL",
         "total_ar": 400000.0, "overdue": 400000.0, "gross_ar": 400000.0,
         "status_key": "stop", "status": "Stop credit", "master_hold": True},
        {"key": "1004000002", "customer": "BETA TRADING SA",
         "total_ar": 50000.0, "overdue": 0.0, "gross_ar": 100000.0,
         "status_key": "ok", "status": "Within terms", "master_hold": False},
    ],
    "controls_summary": [],
    "summary": {"total_ar": 450000.0, "overdue_total": 400000.0,
                "credits_total": -50000.0, "actions_due": 0,
                "invoice_count": 4, "customer_count": 2},
    "total_check": None, "tb_check": None, "reconciliation": None,
    "master_used": False, "master_stats": None, "unmapped": [],
    "currencies": ["XAF"],
    "source": "ar_test_v94.xlsx", "created_at": "2026-07-17 08:00",
}
ctp.save_result("94feedbeefab", RESULT)
s = ctp.compute_daily_stats(json.loads(json.dumps(RESULT, default=str)))
check("held honours master_hold (1 of 2)", s["held"] == 1)
check("pct>60 = 80.0 / pct>90 = 20.0",
      s["pct_over_60"] == 80.0 and s["pct_over_90"] == 20.0)
ctp.record_daily_stats(RESULT)

r = client.get("/tools/ongoing-ctp-monitoring")
check("CtP portal shows the held-accounts daily graph",
      r.status_code == 200 and "Accounts on credit hold" in r.text)

# === 4. Cheque register AR columns ==========================================
bdir = cheques.BATCH_DIR / "aa11bb22cc33"
bdir.mkdir(parents=True)
(bdir / "results.json").write_text(json.dumps({
    "status": "done", "uploaded_by": "tester",
    "created_at": "2026-07-17 08:30", "banks": [], "bank_line_count": 0,
    "cheques": [
        {"idx": 0, "filename": "acme.png", "stored": "000.png",
         "media": "image/png", "status": "done",
         "result": {"cheque_number": "5550001",
                    "customer_name": "ACME LOGISTICS SARL",
                    "amount": 150000.0, "amount_currency": "XAF",
                    "cheque_date": "2026-07-01", "drawee_bank": "SCB",
                    "readability": "good"},
         "appearances": [{"bank": "SGBC", "amount": 150000.0,
                          "date": "2026-07-10",
                          "text": "REMISE CHQ 5550001 ACME"}], "error": ""},
        {"idx": 1, "filename": "who.png", "stored": "001.png",
         "media": "image/png", "status": "done",
         "result": {"cheque_number": "5550002",
                    "customer_name": "TOTALLY UNKNOWN ZZZQX",
                    "amount": 42000.0, "amount_currency": "XAF",
                    "cheque_date": "2026-07-02", "drawee_bank": "UBA",
                    "readability": "good"},
         "appearances": [], "error": ""},
    ]}, ensure_ascii=False), encoding="utf-8")

r = client.get("/tools/cheque-processing")
check("register shows the AR columns", r.status_code == 200
      and "AR customer (latest analysis)" in r.text
      and "Overdue balance" in r.text)
check("ACME matched to its AR account + balances",
      "1004000001" in r.text and "400,000" in r.text)
check("unknown name shows — (no match)", ">TOTALLY UNKNOWN ZZZQX<" in r.text)

rx = client.get("/tools/cheque-processing/register/export")
check("register Excel downloads", rx.status_code == 200)
import io  # noqa: E402

import openpyxl  # noqa: E402
wsx = openpyxl.load_workbook(io.BytesIO(rx.content)).active
head = [c.value for c in wsx[1]]
check("Excel carries the 4 AR columns in place",
      head[4:9] == ["Client name", "AR customer (latest analysis)",
                    "AR account", "AR balance", "Overdue balance"]
      and "Scan file" in head and head.index("Scan file") == 11)
acme_row = next(row for row in wsx.iter_rows(min_row=2, values_only=True)
                if row[4] == "ACME LOGISTICS SARL")
check("Excel AR values populated (account + balances)",
      acme_row[6] == "1004000001" and acme_row[7] == 400000
      and acme_row[8] == 400000)

# === 5. Dashboard ============================================================
r = client.get("/")
check("dashboard shows the unpresented-cheques KPI",
      r.status_code == 200 and "Unpresented cheques" in r.text)
check("dashboard shows the ageing trend", "&gt; 60 days %" in r.text
      or "> 60 days %" in r.text)
check("dashboard shows the trends section",
      "Receivables ageing" in r.text and "open items per day" in r.text)
check("dashboard shows the credit-hold daily graph",
      "Accounts on credit hold" in r.text)

# === 6. Orange Money — AR columns, history, comparison ======================
customers.set_name(orange.TOOL_SLUG, "699559325", "ACME LOGISTICS SARL")
customers.set_name(orange.ACCT_SLUG, "699559325", "1004000001")

with open(SAMPLE, "rb") as fh:
    r = client.post("/tools/orange-cameroun/upload",
                    files={"file": ("orange_mai.xlsx", fh, XLSX)})
check("review page renders", r.status_code == 200)
check("review shows the AR-ledger columns",
      "Account per AR ledger" in r.text and "Balance per AR ledger" in r.text)
check("saved account joined to the AR ledger (400 000 balance)",
      "400 000" in r.text)
m = re.search(r'name="htoken" value="([0-9a-f]{12})"', r.text)
check("history token embedded for generate", bool(m))
htoken = m.group(1)
check("upload summary persisted", (orange.HISTORY_DIR / f"{htoken}.json").exists())

with open(SAMPLE, "rb") as fh:
    client.post("/tools/orange-cameroun/upload",
                files={"file": ("orange_mai_again.xlsx", fh, XLSX)})
check("re-upload of the same month replaces (stable token)",
      len(list(orange.HISTORY_DIR.glob("*.json"))) == 1)

u = orange.mark_generated(htoken, {"699559325": {
    "name": "ACME LOGISTICS SARL", "account": "1004000001"}})
check("generated stamp + identities stored on the history entry",
      u and u["generated_at"]
      and any(c["account"] == "1004000001" for c in u["correspondants"]))

r = client.get("/tools/orange-cameroun")
check("processed files listed on the tool page",
      "Processed files" in " ".join(r.text.split())
      and "Compare vs AR" in r.text)

r = client.get(f"/tools/orange-cameroun/history/{htoken}")
check("history comparison page renders", r.status_code == 200
      and "Balance per latest AR TB" in r.text and "400 000" in r.text
      and "1004000001" in r.text)
r = client.get("/tools/orange-cameroun/history/ffffffffffff")
check("unknown history token -> friendly redirect",
      r.status_code == 303 and "error=" in r.headers["location"])

# === 7. Account Stop =========================================================
check("account-stop registered", any(t["slug"] == "account-stop"
                                     for t in __import__("app.tools.registry",
                                                          fromlist=["TOOLS"]).TOOLS))
bank.STORE_DIR.mkdir(parents=True, exist_ok=True)
(bank.STORE_DIR / "beefcafe0001.json").write_text(json.dumps({
    "token": "beefcafe0001", "created_at": "2026-07-15 10:00",
    "unmatched": [
        {"payer": "ACME LOGISTICS", "description": "VIR ACME LOGISTICS SARL",
         "credit": 250000.0, "debit": 0.0, "date": "2026-07-14",
         "text": "VIR ACME LOGISTICS SARL REF 778", "bank": "SGBC"},
        {"payer": "", "description": "FRAIS TENUE DE COMPTE",
         "credit": 0.0, "debit": 5000.0, "date": "2026-07-14",
         "text": "FRAIS", "bank": "SGBC"}],
    "matched": []}, ensure_ascii=False), encoding="utf-8")

lines = bank.all_credit_lines()
check("all_credit_lines keeps credits only (payer preserved)",
      len(lines) == 1 and lines[0]["payer"] == "ACME LOGISTICS"
      and lines[0]["credit"] == 250000.0)

view = account_stop.build_view()
check("one stopped account in the view",
      view["counts"]["stopped"] == 1
      and view["accounts"][0]["key"] == "1004000001")
a = view["accounts"][0]
check("bank payment traced", a["bank"]
      and a["bank"][0]["amount"] == 250000.0 and a["bank"][0]["bank"] == "SGBC")
check("cheque payment traced (cleared bank)", a["cheques"]
      and a["cheques"][0]["bank"] == "SGBC"
      and a["cheques"][0]["amount"] == 150000.0)
check("mobile money traced from the orange upload", a["momo"]
      and a["momo"][0]["operator"] == "Orange Money"
      and all(p["amount"] > 0 for p in a["momo"]))
momo_before = len(a["momo"])
view2 = account_stop.build_view()
check("momo payments deduped across re-reads",
      len(view2["accounts"][0]["momo"]) == momo_before)

r = client.get("/tools/account-stop")
check("account-stop page renders the three groups", r.status_code == 200
      and "Bank payments" in r.text and "Cheque payments" in r.text
      and "Mobile money" in r.text)
check("Account Stop present in the sidebar menu",
      'href="/tools/account-stop"' in r.text
      and 'nav-label">Account Stop' in r.text)
check("stopped account row rendered with payments",
      "1004000001" in r.text and "ACME LOGISTICS" in r.text
      and "Orange Money" in r.text and "250,000" in r.text)

# === 8. Access rights cover the new slug =====================================
acfg = {"enabled": True, "users": {"a": "x", "e": "x"}, "admins": ["a"],
        "access": {"e": {"cheque-processing": "read"}}}
check("non-admin without grant -> none",
      auth.area_level(acfg, "e", "account-stop") == "none")
acfg["access"]["e"]["account-stop"] = "read"
check("non-admin with grant -> read; admin always modify",
      auth.area_level(acfg, "e", "account-stop") == "read"
      and auth.area_level(acfg, "a", "account-stop") == "modify")

# === 9. BIT & Cash AR banner — no age bracket ================================
r = client.get("/tools/bit-cash-ar")
check("banner carries no '(today)' bracket",
      r.status_code == 200 and "(today)" not in r.text
      and "day old" not in r.text)

# === 10. BIT recon: arithmetic total, ±1000 window, manual plug =============
AWBS = [("4095823454", 53639.0), ("3751239391", 53639.0),
        ("6057894780", 53639.0), ("1616021735", 53639.0),
        ("1107827641", 56100.0), ("8457401400", 59600.0),
        ("8457378215", 73800.0), ("7177996162", 119515.0)]
FAKE_ROWS = {
    "bit_header": [],
    "bit": [{"id": 0, "amount": -523500.0, "gl_account": "1263001293",
             "assignment": "0548407000018", "posting_date": "07.07.2026",
             "reference": "CMBUE", "text": "CASH DEPOSIT BUEA",
             "doc_no": "410001", "raw": []},
            {"id": 1, "amount": -1000000.0, "gl_account": "1263001293",
             "assignment": "0548407000019", "posting_date": "07.07.2026",
             "reference": "X", "text": "OTHER", "doc_no": "410002",
             "raw": []}],
    "cash": [{"id": i, "awb": awb, "reference": awb, "amount": amt,
              "sap_acct": "4003025929", "assignment": awb,
              "customer": "DHL SERVICE POINT BUEA", "doc_no": f"90000{i}"}
             for i, (awb, amt) in enumerate(AWBS)],
}
bitcash.rows_store = lambda: FAKE_ROWS

st = {"label": "DHL BUEA CASHCMBUE", "date": "07.07.2026",
      "total": 523500.0,        # the (wrong) figure a document might state
      "lines": [{"reference": awb, "amount": amt, "description": ""}
                for awb, amt in AWBS]}
lines, ar_sel, cands, bit_sel = bitcash.automatch(st)
check("evidence total recomputed arithmetically (523,571)",
      st["total"] == 523571.0 and st["stated_total"] == 523500.0)
check("BIT candidate found within the window (523,500) and auto-selected",
      cands == [0] and bit_sel == 0)
check("all 8 AWBs matched to the Cash AR",
      len(ar_sel) == 8 and all(ln["matched_ids"] for ln in lines))
check("margin is a constant of 1000", bitcash.BIT_MATCH_MARGIN == 1000.0)

near = {"total": 0, "lines": [{"reference": "999", "amount": 524000.0,
                               "description": ""}]}
_l, _a, near_cands, near_sel = bitcash.automatch(near)
check("a 500-off amount is now a candidate (524,000 finds 523,500)",
      near_cands == [0] and near_sel == 0)

far = {"total": 0, "lines": [{"reference": "999", "amount": 525000.0,
                              "description": ""}]}
_l, _a, far_cands, _s = bitcash.automatch(far)
check("amounts beyond ±1000 are not candidates (525,000 vs 523,500)",
      far_cands == [])

st["lines"] = lines
rec = {"token": "feedbead0001", "status": "open",
       "uploaded": "2026-07-17 10:00", "uploaded_by": "tester",
       "source": "DHL BUEA CASHCMBUE 07.07.26.xlsx", "statement": st,
       "ar_selected": ar_sel, "bit_candidates": cands, "bit_selected": bit_sel,
       "bit_duplicate": False, "error": ""}
bitcash.save_recon(rec)
view = bitcash.recon_view(rec)
check("difference disclosed (71 = 523,571 - 523,500)",
      view["ar_total"] == 523571.0 and view["bit_amount"] == 523500.0
      and view["difference"] == 71.0 and view["residual"] == 71.0)
check("candidate carries its delta vs the evidences (-71)",
      view["bit_candidates"][0]["delta"] == -71.0)

r = client.get("/tools/bit-cash-ar/recon/feedbead0001")
check("sandbox narrative discloses the arithmetic total",
      r.status_code == 200 and "Total of evidences is XAF" in r.text
      and "523,571" in r.text)
check("sandbox shows the manual-plug section", "Manual plug" in r.text)

r = client.post("/tools/bit-cash-ar/recon/feedbead0001/plug",
                data={"amount": "71", "account": "", "note": ""})
check("plug without a G/L account is refused",
      r.status_code == 303 and "error=" in r.headers["location"])
r = client.post("/tools/bit-cash-ar/recon/feedbead0001/plug",
                data={"amount": "71", "account": "471000",
                      "note": "BANKED SHORT"})
check("plug saved via the route", r.status_code == 303
      and "Manual+plug+saved" in r.headers["location"].replace("%20", "+"))
view = bitcash.recon_view(bitcash.load_recon("feedbead0001"))
check("residual is 0 after the plug",
      view["plug"]["amount"] == 71.0 and view["residual"] == 0.0)

bitcash.set_status("feedbead0001", True, "tester")
je = bitcash.build_journal(_tmp / "je_v95.xlsx")
check("journal built — 8 AWB pairs + balanced plug pair = 18 lines",
      je and je["count"] == 1 and je["lines"] == 18
      and je["tokens"] == ["feedbead0001"])
wj = openpyxl.load_workbook(_tmp / "je_v95.xlsx")
wsj = next(ws for ws in wj.worksheets if ws.title.startswith("CM01_"))
jrows = [row for row in wsj.iter_rows(min_row=4, max_row=24,
                                      values_only=True)
         if row[0] is not None]
check("bank side split per AWB (first pair: 40 + 15, 53,639, AWB in Doc.Nr)",
      jrows[0][9] == 1263001293 and jrows[0][10] == 40
      and jrows[0][11] == 53639.0 and jrows[0][6] == "4095823454"
      and jrows[1][9] == 4003025929 and jrows[1][10] == 15
      and jrows[1][11] == 53639.0 and jrows[1][6] == "4095823454")
plug_pair = [r for r in jrows if r[11] == 71.0]
check("plug booked as a balanced pair (BIT G/L key 50 + 471000 key 40)",
      len(plug_pair) == 2
      and plug_pair[0][9] == 1263001293 and plug_pair[0][10] == 50
      and plug_pair[1][9] == 471000 and plug_pair[1][10] == 40
      and plug_pair[1][13] == "BANKED SHORT")
sum40 = sum(r[11] for r in jrows if r[10] == 40)
sum50 = sum(r[11] for r in jrows if r[10] == 50)
sum15 = sum(r[11] for r in jrows if r[10] == 15)
check("workbook CHECK stays balanced (+40s -50s -15s = 0)",
      round(sum40 - sum50 - sum15, 2) == 0.0)
check("BIT G/L nets to the actually banked amount (523,500)",
      round(sum(r[11] if r[10] == 40 else -r[11]
                for r in jrows if r[9] == 1263001293), 2) == 523500.0)

# === 11. v9.6 — evidence reader fallback + async BIT/Cash AR uploads ========
import time  # noqa: E402


def _buea_like(path):
    """Replica of the DHL BUEA layout that broke the reader: decorative rows,
    an Excel table's phantom 'Column1…' header ABOVE the real header, a #REF!
    remnant and a TEXT total cell."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["", "", "", "", "", "", "", "#REF!"])
    ws.append(["", "", "DHL BUEA CASH RECORD SHEET.", "", "", "", "", ""])
    ws.append(["", "", "Column1", "Column2", "   ", "Column4", "Column5", ""])
    ws.append(["", "", "Date", "Waybill", "Weight (kg, gram)", "Country",
               "Amount (Xaf)", ""])
    ws.append(["", "", "07.07.2026", 4095823454, "0.5 KG", "CANADA", 53639, ""])
    ws.append(["", "", "07.07.2026", 1107827641, "0.5 KG", "FRANCE", 56100, ""])
    ws.append(["", "", "07.07.2026", 7177996162, "0.5 KG", "SURINAME",
               119515, ""])
    ws.append(["", "", "", "", "", "", "TOTAL: 229.254", ""])
    wb.save(path)
    return path


st6 = bitcash._read_excel_statement(_buea_like(_tmp / "buea_like.xlsx"))
check("BUEA-style layout read via the row-scan fallback (3 AWBs, 229,254)",
      st6["total"] == 229254.0 and len(st6["lines"]) == 3
      and st6["lines"][0]["reference"] == "4095823454")

bad = _tmp / "no_columns.xlsx"
wbx = openpyxl.Workbook()
wbx.active.append(["Foo", "Bar"])
wbx.active.append(["x", "y"])
wbx.save(bad)
try:
    bitcash._read_excel_statement(bad)
    check("unreadable evidence raises a clear error", False)
except ValueError as exc:
    check("unreadable evidence raises a clear error",
          "Waybill / AWB / Reference" in str(exc))


def _wait_bitcash(timeout=15.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if not bitcash.status().get("processing"):
            return
        time.sleep(0.1)
    raise SystemExit("background BIT/Cash AR processing never finished")


# Processing banner branches (deterministic — flag written directly).
bitcash._save({"current": {}, "history": {},
               "processing": {"started": "2026-07-17 12:00",
                              "beat": time.time(),
                              "files": [{"kind": "bit", "label": "BIT",
                                         "source": "big.xlsx"}],
                              "stage": "storing the rows", "current": "BIT",
                              "pct": 62, "rows_done": 1500,
                              "rows_total": 2400, "eta_seconds": 95}})
r = client.get("/tools/bit-cash-ar")
flat = " ".join(r.text.split())
check("live tracker shown: %, rows and time left",
      "Ingesting" in flat and "location.replace" in r.text
      and "62%" in flat and "1,500 of 2,400 row(s)" in flat
      and "1 min 35 s left" in flat)
# A stale flag (no heartbeat for >15 min — server restarted) self-clears.
bitcash._save({"current": {}, "history": {},
               "processing": {"started": "2026-07-17 12:00", "beat": 1.0,
                              "files": [], "stage": "queued", "pct": 0}})
st_stale = bitcash.status()
check("stale processing flag cleared into an error",
      not st_stale.get("processing")
      and "interrupted" in (st_stale.get("processing_error") or {})
      .get("message", ""))
bitcash._save({"current": {}, "history": {},
               "processing_error": {"at": "2026-07-17 12:01",
                                    "message": "Could not read big.xlsx"}})
r = client.get("/tools/bit-cash-ar")
check("failed upload surfaced on the page",
      "The last upload failed" in r.text and "Could not read big.xlsx" in r.text)
bitcash._save({"current": {}, "history": {}})

# Real async round-trip: instant 303, counts land in the background.
wb_bit = openpyxl.Workbook()
wb_bit.active.append(["Ref", "Amount", "Status"])
wb_bit.active.append(["T1", 100, "Open"])
wb_bit.active.append(["T2", 200, "Closed"])
bit_path = _tmp / "bit_async.xlsx"
wb_bit.save(bit_path)
with open(bit_path, "rb") as fh:
    r = client.post("/tools/bit-cash-ar/upload",
                    files={"bit_file": ("bit_async.xlsx", fh, XLSX)})
check("upload responds instantly with a redirect", r.status_code == 303
      and "received" in r.headers["location"])
_wait_bitcash()
st7 = bitcash.status()
check("counts recorded by the background parser",
      st7["current"]["bit"]["open_items"] == 1
      and not st7.get("processing") and not st7.get("processing_error"))

# Corrupt file -> background error recorded, nothing crashes.
junk = _tmp / "junk.xlsx"
junk.write_bytes(b"this is not a workbook")
bitcash.process_uploads_async([("bit", junk, "junk.xlsx")])
_wait_bitcash()
check("corrupt upload lands as a processing error",
      (bitcash.status().get("processing_error") or {}).get("message", "")
      .startswith("Could not read junk.xlsx"))

# === 12. v9.9 — JE pack, variance email, deposit slip, progress, replace ====
import zipfile  # noqa: E402
from io import BytesIO  # noqa: E402
from urllib.parse import parse_qs, urlencode, urlparse  # noqa: E402

# --- pack: advice + slip + evidences Excel + journal, zipped
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
(bitcash.FILES_DIR / "advice_feedbead0001.pdf").write_bytes(b"%PDF-1.4 advice")
rec1 = bitcash.load_recon("feedbead0001")
rec1["file"] = "advice_feedbead0001.pdf"
bitcash.save_recon(rec1)
(bitcash.FILES_DIR / "slip_feedbead0001.pdf").write_bytes(b"%PDF-1.4 slip")
check("deposit slip recorded on the sandbox",
      bitcash.set_slip("feedbead0001", "slip_feedbead0001.pdf",
                       "BUEA deposit slip.pdf") is not None)

pack = _tmp / "pack.zip"
bitcash.build_pack(pack, _tmp / "je_v95.xlsx",
                   "CM01_PC TEMPLATE_17.07.26_PC.xlsx", ["feedbead0001"])
with zipfile.ZipFile(pack) as zf:
    names = zf.namelist()
    check("pack carries the journal at the root",
          "CM01_PC TEMPLATE_17.07.26_PC.xlsx" in names)
    check("pack carries the advice copy, the slip and the evidences Excel",
          any("Payment advice - " in n for n in names)
          and any("Deposit slip - " in n for n in names)
          and any(n.endswith("Evidences_feedbead0001.xlsx") for n in names))
    ev_name = next(n for n in names if n.endswith(".xlsx") and "Evidences" in n)
    wse = openpyxl.load_workbook(BytesIO(zf.read(ev_name)))["Evidences"]
    ev_text = " ".join(str(c.value) for row in wse.iter_rows()
                       for c in row if c.value is not None)
    check("evidences Excel lists the AWBs, the total and the plug",
          "4095823454" in ev_text and "TOTAL OF EVIDENCES" in ev_text
          and "523571" in ev_text.replace(",", "").replace(".0", "")
          and "471000" in ev_text)

# --- journal route now redirects with journal + pack downloads
r = client.post("/tools/bit-cash-ar/journal")
check("journal route -> redirect with journal + pack params",
      r.status_code == 303 and "journal=" in r.headers["location"]
      and "pack=" in r.headers["location"])
qs2 = parse_qs(urlparse(r.headers["location"]).query)
page = client.get("/tools/bit-cash-ar?" + urlencode(
    {"journal": qs2["journal"][0], "jname": qs2["jname"][0],
     "pack": qs2["pack"][0]}, doseq=True))
check("home page offers the pack + journal downloads",
      "Full pack" in page.text and "Journal entry only" in page.text)
(main.OUTPUT_DIR / qs2["journal"][0]).unlink(missing_ok=True)
(main.OUTPUT_DIR / qs2["pack"][0]).unlink(missing_ok=True)

# --- variance email: excess payment far above the 5,000 threshold
st2 = dict(st, lines=[dict(ln) for ln in st["lines"]])
rec2 = {"token": "feedbead0002", "status": "open",
        "uploaded": "2026-07-17 11:00", "uploaded_by": "tester",
        "source": "EXCESS CASE.xlsx", "statement": st2,
        "ar_selected": list(range(8)), "bit_candidates": [0, 1],
        "bit_selected": 1, "bit_duplicate": False, "error": "",
        "file": ""}
bitcash.save_recon(rec2)
view2 = bitcash.recon_view(rec2)
check("excess payment above threshold flags the email trigger",
      view2["difference"] == -476429.0 and view2["needs_email"])
page = client.get("/tools/bit-cash-ar/recon/feedbead0002")
check("sandbox shows the variance clarification prompt",
      "Variance clarification" in page.text
      and "excess payment" in page.text
      and "5,000" in page.text and "account holder" in page.text)
r = client.post("/tools/bit-cash-ar/recon/feedbead0002/email",
                data={"recipient": "client@example.com",
                      "subject": "Clarification",
                      "body": "An excess payment of 476,429 XAF was received."})
check("email route falls back to a ready-to-send .eml (SMTP off in tests)",
      r.status_code == 303 and "ready-to-send" in
      r.headers["location"].replace("+", " ").replace("%20", " "))
ve = (bitcash.load_recon("feedbead0002") or {}).get("variance_email") or {}
check("variance email recorded on the sandbox (eml mode)",
      ve.get("mode") == "eml" and ve.get("to") == "client@example.com"
      and (main.OUTPUT_DIR / ve.get("eml", "x")).exists())
page = client.get("/tools/bit-cash-ar/recon/feedbead0002")
check("sandbox discloses the prepared clarification",
      "Clarification" in page.text and "prepared" in page.text)
(main.OUTPUT_DIR / ve["eml"]).unlink(missing_ok=True)

# --- slip route + evidence download
r = client.post("/tools/bit-cash-ar/recon/feedbead0002/slip",
                files={"slip": ("Deposit BUEA.pdf", b"%PDF-1.4 slip2",
                                "application/pdf")})
check("slip upload attaches to the sandbox", r.status_code == 303
      and (bitcash.FILES_DIR / "slip_feedbead0002.pdf").exists())
r = client.get("/tools/bit-cash-ar/recon/feedbead0002/evidence/slip")
check("slip is viewable via the evidence route",
      r.status_code == 200 and r.content.startswith(b"%PDF"))
r = client.get("/tools/bit-cash-ar/recon/feedbead0002/evidence/advice",
               follow_redirects=False)
check("missing advice degrades to a friendly redirect",
      r.status_code == 303 and "error=" in r.headers["location"])

# --- single-parse progress ticks + replace cleanup + history continuity
data_bc = bitcash._load()
data_bc["history"]["2000-01-01"] = {"bit_open": 5}
bitcash._save(data_bc)

wbp = openpyxl.Workbook()
wbp.active.append(["Ref", "Amount", "Status"])
for i in range(3):
    wbp.active.append([f"P{i}", 100 + i, "Open"])
p1 = bitcash.UPLOAD_DIR / "bitcash_bit_first.xlsx"
wbp.save(p1)
ticks = []
bitcash.record_upload("bit", p1, "first.xlsx",
                      progress=lambda **kw: ticks.append(kw))
check("progress ticks cover the stages up to 100%",
      any(t.get("stage") == "reading the workbook" for t in ticks)
      and any(t.get("pct") == 100 for t in ticks)
      and any(t.get("rows_total") == 3 for t in ticks))
check("current position remembers the stored file",
      bitcash.status()["current"]["bit"]["stored"] == "bitcash_bit_first.xlsx")

wbp.save(bitcash.UPLOAD_DIR / "bitcash_bit_second.xlsx")
bitcash.record_upload("bit", bitcash.UPLOAD_DIR / "bitcash_bit_second.xlsx",
                      "second.xlsx")
check("replaced upload file cleaned up, new one is the reference",
      not p1.exists()
      and bitcash.status()["current"]["bit"]["stored"]
      == "bitcash_bit_second.xlsx")
hist_days = dict(bitcash.status()["history"])
check("daily graph keeps prior-day figures after replacement",
      "2000-01-01" in hist_days and hist_days["2000-01-01"]["bit_open"] == 5)

# === 13. v10.0 — reconciliations fixed in time + slip-anchored matching =====
# Simulate a file replacement: the row store gets NEW generation stamps.
FAKE_ROWS["gen_bit"] = "genB1"
FAKE_ROWS["gen_cash"] = "genC1"

# (a) The approved sandbox is FROZEN: even mutating the live rows (as a
# replaced file would) cannot change its figures — the exact bug reported.
view1 = bitcash.recon_view(bitcash.load_recon("feedbead0001"))
check("approved sandbox reads from the frozen snapshot",
      view1["frozen"] and not view1["stale"] and view1["ar_total"] == 523571.0)
_saved_amounts = [r["amount"] for r in FAKE_ROWS["cash"]]
for r_ in FAKE_ROWS["cash"]:
    r_["amount"] = 999999.0
view1b = bitcash.recon_view(bitcash.load_recon("feedbead0001"))
check("replacing the live rows can NOT change an approved reconciliation",
      view1b["ar_total"] == 523571.0 and view1b["bit_amount"] == 523500.0)
for r_, amt in zip(FAKE_ROWS["cash"], _saved_amounts):
    r_["amount"] = amt

# (b) An OPEN sandbox matched against the OLD files reads as stale — it
# resolves nothing instead of silently mixing sources.
st3 = {"label": "STALE CASE", "date": "", "total": 523571.0,
       "lines": [{"reference": a, "amount": amt, "description": "",
                  "matched_ids": []} for a, amt in AWBS]}
rec3 = {"token": "feedbead0003", "status": "open",
        "uploaded": "2026-07-17 09:00", "uploaded_by": "tester",
        "source": "OLD FILES.xlsx", "statement": st3,
        "rows_gen": {"bit": "OLDGEN", "cash": "OLDGEN"},
        "ar_selected": [0, 1, 2], "bit_candidates": [0], "bit_selected": 0,
        "bit_duplicate": False, "error": "", "file": ""}
bitcash.save_recon(rec3)
view3 = bitcash.recon_view(rec3)
check("stale sandbox resolves nothing (no mixed-up rows)",
      view3["stale"] and view3["ar_rows"] == []
      and view3["bit_candidates"] == [] and view3["ar_total"] == 0)
page = client.get("/tools/bit-cash-ar/recon/feedbead0003")
check("stale sandbox shows the replacement warning + re-run button",
      "files on record were replaced" in page.text
      and "Re-run" in page.text)
check("approval refused while stale",
      bitcash.set_status("feedbead0003", True, "t") == "stale")
check("selection refused while stale",
      bitcash.set_selection("feedbead0003", ["0"], "0") == "stale")

# (c) One click re-matches against the CURRENT files and heals the sandbox.
rec3b = bitcash.rematch("feedbead0003")
check("re-match rebuilds selections against the current generation",
      rec3b and rec3b["rows_gen"] == {"bit": "genB1", "cash": "genC1"}
      and rec3b["ar_selected"] and rec3b["bit_selected"] == 0
      and not bitcash.recon_view(rec3b)["stale"])

# (d) A legacy APPROVED sandbox without a frozen snapshot whose files were
# replaced is SKIPPED by the journal, never written from wrong rows.
rec4 = dict(rec3, token="feedbead0004", status="approved",
            rows_gen={"bit": "OLDGEN", "cash": "OLDGEN"},
            approved_by="t", approved_at="2026-07-17 09:30")
rec4.pop("frozen", None)
bitcash.save_recon(rec4)
je2 = bitcash.build_journal(_tmp / "je_v100.xlsx")
check("journal writes frozen sandboxes and SKIPS stale legacy ones",
      je2 and "feedbead0001" in je2["tokens"]
      and "feedbead0004" not in je2["tokens"] and je2["skipped"] >= 1)

# (e) The deposit slip's banked total anchors the BIT search.
st5 = {"label": "SLIP CASE", "total": 999999.0,
       "lines": [{"reference": "5550001112", "amount": 999999.0,
                  "description": ""}]}
_l5, _a5, cands5, sel5 = bitcash.automatch(st5, slip_total=523500.0)
check("slip total wins the BIT selection over the evidences total",
      sel5 == 0 and 0 in cands5)

# (f) Slip uploaded TOGETHER with the statement drives the whole flow.
wbs = openpyxl.Workbook()
wbs.active.append(["AWB", "Amount"])
wbs.active.append(["4095823454", 53639])
stmt_path = _tmp / "slipflow_stmt.xlsx"
wbs.save(stmt_path)
wsl = openpyxl.Workbook()
wsl.active.append(["Ref", "Amount"])
wsl.active.append(["1", 523500])
slip_path = _tmp / "slipflow_slip.xlsx"
wsl.save(slip_path)
with open(stmt_path, "rb") as f1, open(slip_path, "rb") as f2:
    r = client.post("/tools/bit-cash-ar/recon/upload",
                    files={"statement": ("advice.xlsx", f1, XLSX),
                           "deposit_slip": ("slip.xlsx", f2, XLSX)})
check("statement + slip upload accepted", r.status_code == 303)
tok5 = r.headers["location"].rstrip("/").split("/")[-1]
for _ in range(100):
    rec5 = bitcash.load_recon(tok5)
    if rec5 and rec5["status"] in ("open", "error"):
        break
    time.sleep(0.1)
check("slip stored with the sandbox and its banked total read",
      rec5["status"] == "open"
      and (rec5.get("slip") or {}).get("name") == f"slip_{tok5}.xlsx"
      and rec5.get("slip_total") == 523500.0)
check("BIT auto-selected from the slip's banked amount",
      rec5.get("bit_selected") == 0)
page = client.get(f"/tools/bit-cash-ar/recon/{tok5}")
flat5 = " ".join(page.text.split())
check("sandbox narrative discloses the deposit-slip total",
      "Deposit slip total: 523,500" in flat5)

# === 13. v10.1 — match precision, date columns, monthly momo, delete, sync ==
from app.services import bank_statement  # noqa: E402

check("account-stop matches at strict 0.90 precision",
      account_stop.MATCH_RATIO == 0.90)

# The strict ratio must reach every matcher call (bank, cheque, momo paths).
_seen_ratios = []
_orig_best = bank_statement.best_customer_for


def _spy_best(text, custs, min_ratio=0.86):
    _seen_ratios.append(min_ratio)
    return _orig_best(text, custs, min_ratio=min_ratio)


bank_statement.best_customer_for = _spy_best
try:
    account_stop.build_view()
finally:
    bank_statement.best_customer_for = _orig_best
check("every account-stop matcher call uses the strict ratio",
      _seen_ratios and all(r == 0.90 for r in _seen_ratios))

r = client.get("/tools/account-stop")
check("transaction-date column present in all three groups",
      r.text.count("Transaction date on statement") == 3)

r = client.get("/tools/ongoing-ctp-monitoring")
check("credit-hold graph sits at the TOP of the portal",
      "Accounts on credit hold" in r.text
      and r.text.index("Accounts on credit hold")
      < r.text.index("1. AR transaction file"))

# Orange monthly pivot: two months + one duplicated transaction across files.
orange._write_upload({
    "token": "aaaa00000001", "uploaded_at": "2026-06-05 09:00",
    "source": "june.xls",
    "meta": {"account": "658902134", "company": "", "month_label":
             "Juin 2026", "period_start": "2026-06-01",
             "period_end": "2026-06-30"},
    "totals": {"count": 2, "credited": 30000.0, "correspondant_count": 1},
    "correspondants": [{"correspondant": "699111222", "name": "NOVIA SARL",
                        "account": "4003025929", "count": 2,
                        "total": 30000.0}],
    "collections": [
        {"correspondant": "699111222", "date": "05/06/2026",
         "reference": "PP260605.1000.X1", "credit": 10000.0},
        {"correspondant": "699111222", "date": "12/06/2026",
         "reference": "PP260612.1200.X2", "credit": 20000.0}]})
orange._write_upload({
    "token": "aaaa00000002", "uploaded_at": "2026-07-03 09:00",
    "source": "july.xls",
    "meta": {"account": "658902134", "company": "", "month_label":
             "Juillet 2026", "period_start": "2026-07-01",
             "period_end": "2026-07-31"},
    "totals": {"count": 2, "credited": 25000.0, "correspondant_count": 1},
    "correspondants": [{"correspondant": "699111222", "name": "NOVIA SARL",
                        "account": "4003025929", "count": 2,
                        "total": 25000.0}],
    "collections": [
        {"correspondant": "699111222", "date": "12/06/2026",
         "reference": "PP260612.1200.X2", "credit": 20000.0},
        {"correspondant": "699111222", "date": "02/07/2026",
         "reference": "PP260702.0900.X3", "credit": 5000.0}]})
rep = orange.monthly_report()
mrow = next(r_ for r_ in rep["rows"] if r_["correspondant"] == "699111222")
check("monthly pivot: months ordered and labelled",
      [m["key"] for m in rep["months"]][-2:] == ["2026-06", "2026-07"]
      and any(m["label"] == "Juin 2026" for m in rep["months"]))
check("monthly pivot: duplicate transaction counted once",
      mrow["months"]["2026-06"] == 30000.0
      and mrow["months"]["2026-07"] == 5000.0 and mrow["total"] == 35000.0)
check("monthly pivot row carries name, account and phone",
      mrow["name"] == "NOVIA SARL" and mrow["account"] == "4003025929")
r = client.get("/tools/orange-cameroun/monthly")
check("monthly page renders the pivot", r.status_code == 200
      and "NOVIA SARL" in r.text and "Juin 2026" in r.text
      and "35,000" in r.text)
r = client.get("/tools/orange-cameroun/monthly/export")
check("monthly Excel export downloads", r.status_code == 200
      and r.content[:2] == b"PK")

# Processed-file delete (treated twice -> remove one).
r = client.post("/tools/orange-cameroun/history/aaaa00000002/delete",
                follow_redirects=False)
check("processed file removed via the sandbox", r.status_code == 303
      and orange.load_upload("aaaa00000002") is None)
check("monthly report recomputes without the removed file",
      "2026-07" not in {m["key"]
                        for m in orange.monthly_report()["months"]})
r = client.post("/tools/orange-cameroun/history/ffff00000000/delete",
                follow_redirects=False)
check("unknown processed file -> friendly error",
      r.status_code == 303 and "error=" in r.headers["location"])
r = client.get("/tools/orange-cameroun")
check("upload page offers monthly report + remove buttons",
      "payments by month" in r.text and "🗑 Remove" in r.text)

# Bank upload -> cheque register re-match + daily stats snapshot.
_called = []
_orig_refresh = cheques.refresh_all
_orig_record = cheques.record_daily_stats
cheques.refresh_all = lambda rows, summary: _called.append("refresh") or 0
cheques.record_daily_stats = lambda stats: _called.append("record")
try:
    bank.sync_cheque_register()
finally:
    cheques.refresh_all = _orig_refresh
    cheques.record_daily_stats = _orig_record
check("bank sync re-matches the register and snapshots the daily point",
      _called == ["refresh", "record"])

_orig_sync = bank.sync_cheque_register
_orig_build = bank.build_report
_synced = []
bank.sync_cheque_register = lambda: _synced.append(True)
bank.build_report = lambda files, **kw: {
    "token": kw.get("token", "feed00000001"), "status": "done",
    "banks": [], "summary": {}, "statement_lines": [],
    "created_at": "2026-07-19 09:00", "source": "x"}
try:
    tok = bank.start_report([(_tmp / "none.pdf", "x.pdf")],
                            token="feed00000001")
    deadline = time.time() + 10
    while time.time() < deadline and not _synced:
        time.sleep(0.05)
finally:
    bank.sync_cheque_register = _orig_sync
    bank.build_report = _orig_build
check("background statement reports trigger the cheque sync",
      bool(_synced) and tok == "feed00000001")

# === 14. v10.3 — IRO Statements & Returns (operator self-service) ===========
from app.tools import iro  # noqa: E402

iro.IRO_DIR = _tmp / "iro"
iro.MAIL_LOG = iro.IRO_DIR / "_mail.json"

check("operator links are public (no login)",
      auth.is_public("/operator/abc123def"))

FAKE2 = {
    "bit_header": [], "gen_bit": "gen2b", "gen_cash": "gen2c",
    "bit": [
        {"id": 0, "amount": -115700.0, "gl_account": "1263001293",
         "assignment": "0548407000018", "posting_date": "07.07.2026",
         "reference": "DEP-4471-BICEC", "text": "CASH DEPOSIT DLA",
         "doc_no": "410009", "posting_key": "40", "raw": []},
        {"id": 1, "amount": -115700.0, "gl_account": "1263001293",
         "assignment": "0548407000019", "posting_date": "07.07.2026",
         "reference": "OTHER-REF", "text": "CASH DEPOSIT YDE",
         "doc_no": "410010", "posting_key": "40", "raw": []}],
    "cash": [
        {"id": 0, "sap_acct": "415774002", "awb": "8525614075",
         "assignment": "8525614075", "reference": "8525614075",
         "amount": 59600.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900001"},
        {"id": 1, "sap_acct": "415774002", "awb": "2080666324",
         "assignment": "2080666324", "reference": "2080666324",
         "amount": 56100.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900002"},
        {"id": 2, "sap_acct": "415774002", "awb": "9605628896",
         "assignment": "9605628896", "reference": "9605628896",
         "amount": 60011.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900003"},
        {"id": 3, "sap_acct": "4003025929", "awb": "4095823454",
         "assignment": "4095823454", "reference": "4095823454",
         "amount": 53639.0, "customer": "DHL SERVICE POINT BUEA",
         "doc_no": "900004"}],
}
bitcash.rows_store = lambda: FAKE2

groups = iro.operator_accounts()
by_acct = {g["account"]: g for g in groups}
check("operators grouped from the Cash AR by SAP account",
      "415774002" in by_acct and by_acct["415774002"]["count"] == 3
      and by_acct["415774002"]["name"] == "ETS NEW SERVICE TJR")
check("matched AWBs are no longer open — never resent to an IRO",
      "4003025929" not in by_acct)      # its only AWB is frozen in a sandbox

rec = iro.ensure_token("415774002", "ETS NEW SERVICE TJR")
tok1 = rec["token"]
iro.set_email("415774002", "tjr@example.com")
check("operator record: token issued + email remembered",
      len(tok1) == 32 and iro.find_by_token(tok1)["email"]
      == "tjr@example.com")
old = tok1
tok1 = iro.regenerate_token("415774002")["token"]
check("regenerated link invalidates the old one",
      iro.find_by_token(old) is None
      and iro.find_by_token(tok1)["account"] == "415774002")

stmt_path = _tmp / "iro_stmt.xlsx"
iro.build_statement_xlsx(stmt_path, by_acct["415774002"])
wsx = openpyxl.load_workbook(stmt_path)["Statement"]
stxt = " ".join(str(c.value) for row in wsx.iter_rows()
                for c in row if c.value is not None)
check("statement Excel: AWBs + blank Payment-reference column",
      "8525614075" in stxt and "Payment reference" in stxt
      and "175711" in stxt.replace(",", "").replace(".0", ""))

r = client.get(f"/operator/{tok1}")
check("operator page shows only that account's open AWBs",
      r.status_code == 200 and "8525614075" in r.text
      and "4095823454" not in r.text)
r = client.get("/operator/deadbeef00000000000000000000dead")
check("unknown operator link -> 404", r.status_code == 404)

r = client.post(f"/operator/{tok1}/submit",
                data={"awb": ["8525614075", "2080666324"],
                      "ref_8525614075": "DEP-4471-BICEC",
                      "ref_2080666324": "DEP-4471-BICEC",
                      "payment_date": "2026-07-17",
                      "amount_banked": "115700"},
                files=[("deposit_slip",
                        ("slip.png", b"\x89PNG fake", "image/png"))])
check("portal submission accepted", r.status_code == 200
      and "Submission received" in r.text)
sub = iro.load_record("415774002")["submissions"][-1]
check("submission recorded with references + channel",
      sub["channel"] == "portal" and "DEP-4471-BICEC" in sub["references"])
deadline = time.time() + 15
recx = None
while time.time() < deadline:
    recx = bitcash.load_recon(sub["recon_token"])
    if recx and recx.get("status") != "reading":
        break
    time.sleep(0.1)
check("sandbox opened from the operator return",
      recx and recx["status"] == "open"
      and recx["uploaded_by"] == "operator:415774002"
      and len(recx["statement"]["lines"]) == 2)
check("reference-first BIT match beats the duplicated amount",
      recx["bit_selected"] == 0 and recx.get("ref_hits") == [0])
viewx = bitcash.recon_view(recx)
check("candidate flagged 'matched by reference'",
      viewx["bit_candidates"][0].get("by_ref") is True)
check("slip stored with the sandbox",
      (recx.get("slip") or {}).get("source") == "slip.png")
r = client.get(f"/operator/{tok1}")
check("submitted AWBs leave the operator's statement immediately",
      "8525614075" not in r.text and "9605628896" in r.text)

# Email channel — the same funnel, guarded by the registered sender.
mime_bad = (b"From: intruder@example.com\r\nTo: payref@x\r\n"
            b"Subject: PAYREF 415774002\r\nMessage-ID: <m1@x>\r\n\r\n"
            b"AWB 9605628896 60011\r\n")
check("mail from an unregistered sender is quarantined",
      iro.process_message(mime_bad)["status"] == "quarantined"
      and "not the registered email"
      in iro.quarantine_list()[0]["reason"])
mime_ok = (b"From: tjr@example.com\r\nTo: payref@x\r\n"
           b"Subject: PAYREF 415774002\r\nMessage-ID: <m2@x>\r\n\r\n"
           b"REF: OTHER-REF\r\nDATE: 17/07/2026\r\nTOTAL: 60011\r\n"
           b"AWB 9605628896 60011\r\n")
out = iro.process_message(mime_ok)
check("formatted mail creates a submission", out["status"] == "created")
check("duplicate mail is ignored",
      iro.process_message(mime_ok)["status"] == "duplicate")
deadline = time.time() + 15
recm = None
while time.time() < deadline:
    recm = bitcash.load_recon(out["recon_token"])
    if recm and recm.get("status") != "reading":
        break
    time.sleep(0.1)
check("email return matched via its reference too",
      recm and recm["status"] == "open" and recm["bit_selected"] == 1
      and recm["statement"]["lines"][0]["reference"] == "9605628896")

r = client.get("/tools/bit-cash-ar/operators")
check("staff panel lists operators + quarantine",
      r.status_code == 200 and "415774002" in r.text
      and "intruder@example.com" in r.text)
r = client.get("/tools/bit-cash-ar")
check("BIT & Cash AR page links the operator returns",
      "Operator" in r.text and "return(s) awaiting review" in r.text)

# === 15. v10.4 candidate — cheque register × BIT cross-scan =================
FAKE3 = {
    "bit_header": [], "gen_bit": "gen3b", "gen_cash": "gen3c",
    "bit": [
        {"id": 0, "amount": -250000.0, "gl_account": "1263001293",
         "assignment": "0548407000018", "posting_date": "15.07.2026",
         "reference": "REM SGBC", "text": "REMISE CHQ 4471002 SGBC DLA",
         "doc_no": "410021", "posting_key": "40", "raw": []},
        {"id": 1, "amount": -85000.0, "gl_account": "1263001293",
         "assignment": "0548407000019", "posting_date": "16.07.2026",
         "reference": "", "text": "ENC CHEQUE 5580031 BICEC",
         "doc_no": "410022", "posting_key": "40", "raw": []}],
    # The reseller's open rows again (fresh generation) so the portal
    # upgrades below can submit against a live statement.
    "cash": [
        {"id": 0, "sap_acct": "415774002", "awb": "8525614075",
         "assignment": "8525614075", "reference": "8525614075",
         "amount": 59600.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900001"},
        {"id": 1, "sap_acct": "415774002", "awb": "2080666324",
         "assignment": "2080666324", "reference": "2080666324",
         "amount": 56100.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900002"},
        {"id": 2, "sap_acct": "415774002", "awb": "9605628896",
         "assignment": "9605628896", "reference": "9605628896",
         "amount": 60011.0, "customer": "ETS NEW SERVICE TJR",
         "doc_no": "900003"}],
}
bitcash.rows_store = lambda: FAKE3

cdir = cheques.BATCH_DIR / "bb22cc33dd44"
cdir.mkdir(parents=True, exist_ok=True)
(cdir / "results.json").write_text(json.dumps({
    "status": "done", "uploaded_by": "finance",
    "created_at": "2026-07-19 09:00", "banks": {}, "bank_line_count": 0,
    "cheques": [
        {"idx": 0, "filename": "c1.jpg", "stored": "000.jpg",
         "media": "image/jpeg", "status": "done", "appearances": [],
         "error": "", "matched_at": None, "treated": None,
         "result": {"cheque_number": "4471002",
                    "customer_name": "NOVIA SARL", "amount": 250000.0,
                    "amount_currency": "XAF", "cheque_date": "10/07/2026",
                    "drawee_bank": "SGBC", "readability": "good"}},
        {"idx": 1, "filename": "c2.jpg", "stored": "001.jpg",
         "media": "image/jpeg", "status": "done", "appearances": [],
         "error": "", "matched_at": None, "treated": None,
         "result": {"cheque_number": "5580031",
                    "customer_name": "GOOD LOGISTICS", "amount": 90000.0,
                    "amount_currency": "XAF", "cheque_date": "11/07/2026",
                    "drawee_bank": "BICEC", "readability": "good"}},
        {"idx": 2, "filename": "c3.jpg", "stored": "002.jpg",
         "media": "image/jpeg", "status": "done", "appearances": [],
         "error": "", "matched_at": None, "treated": None,
         "result": {"cheque_number": "6692240",
                    "customer_name": "LA TIGRITUDE", "amount": 50000.0,
                    "amount_currency": "XAF", "cheque_date": "12/07/2026",
                    "drawee_bank": "UBA", "readability": "good"}},
    ]}), encoding="utf-8")

reg3, bmeta = main._register_with_bit(
    [r_ for r_ in cheques.register_rows() if r_["batch"] == "bb22cc33dd44"])
by_chq = {r_["cheque_number"]: r_ for r_ in reg3}
check("cheque found on the BIT with the amount aligned",
      by_chq["4471002"]["bit"] and by_chq["4471002"]["bit"]["aligned"]
      and by_chq["4471002"]["bit"]["date"] == "15.07.2026"
      and by_chq["4471002"]["bit"]["doc_no"] == "410021")
check("cheque on the BIT but the amount differs (85,000 vs 90,000)",
      by_chq["5580031"]["bit"] and not by_chq["5580031"]["bit"]["aligned"]
      and by_chq["5580031"]["bit"]["amount"] == 85000.0)
check("cheque absent from the BIT", by_chq["6692240"]["bit"] is None)

r = client.get("/tools/cheque-processing")
flat3 = " ".join(r.text.split())
check("register page: BIT Reference group + sub-columns render",
      "BIT Reference" in r.text
      and "Transaction reference in BIT file" in flat3
      and "Transaction date in BIT file" in flat3
      and "Amount in BIT file" in flat3)
check("register page: BIT reference/date/amount values shown",
      "REM SGBC" in r.text and "15.07.2026" in r.text
      and "85,000" in flat3 and "⚠" in flat3)
check("register page: no stale colspan left behind",
      'colspan="15"' not in r.text and 'colspan="17"' not in r.text)
check("footnote names the BIT scan",
      "BIT Reference" in flat3 and "scans the BIT file on record" in flat3)

r = client.get("/tools/cheque-processing/register/export")
check("register Excel downloads with the BIT columns",
      r.status_code == 200 and r.content[:2] == b"PK")
wbx3 = openpyxl.load_workbook(io.BytesIO(r.content))
wsx3 = wbx3["Cheque register"]
hdr3 = [c.value for c in wsx3[1]]
check("Excel header carries the BIT reference/date/amount columns",
      hdr3[19:22] == ["BIT transaction reference", "BIT transaction date",
                      "BIT amount"])
rows3 = {str(row[2].value): row for row in wsx3.iter_rows(min_row=2)}
check("Excel BIT values per cheque",
      rows3["4471002"][19].value == "REM SGBC"
      and rows3["4471002"][20].value == "15.07.2026"
      and rows3["4471002"][21].value == 250000.0
      and rows3["5580031"][19].value == "410022"
      and rows3["5580031"][21].value == 85000.0
      and not rows3["6692240"][19].value)

# --- v10.4 portal upgrades: slip pre-read, locked amount, evidence copy ----
from app.services import ai_ocr  # noqa: E402

_orig_extract = ai_ocr.extract_payment_statement
ai_ocr.extract_payment_statement = lambda blob, media, cfg: {
    "label": "DHL EXPRESS CAMEROON SARL", "date": "17/07/2026",
    "total": 115700.0, "lines": []}
try:
    r = client.post(f"/operator/{tok1}/read-slip",
                    files=[("slip", ("dep.png", b"\x89PNG fake2",
                                     "image/png"))])
    side = r.json()
    check("slip pre-read returns the banked total + DHL check",
          r.status_code == 200 and side["total"] == 115700.0
          and side["dhl"] is True and side["slip_id"].startswith("iroslip_"))
finally:
    ai_ocr.extract_payment_statement = _orig_extract

r = client.post(f"/operator/{tok1}/submit",
                data={"awb": ["9605628896"],
                      "ref_9605628896": "DEP-9999-UBA",
                      "slip_id": side["slip_id"],
                      "payment_date": "2026-07-19"})
check("submission with a pre-read slip accepted", r.status_code == 200
      and "banked amount read from your deposit slip" in
      " ".join(r.text.split()).lower())
sub2 = iro.load_record("415774002")["submissions"][-1]
check("submission entry carries the locked banked amount",
      sub2["amount"] == 115700.0)
deadline = time.time() + 15
recy = None
while time.time() < deadline:
    recy = bitcash.load_recon(sub2["recon_token"])
    if recy and recy.get("status") != "reading":
        break
    time.sleep(0.1)
check("sandbox carries the pre-read slip total (no second read)",
      recy and recy["status"] == "open"
      and recy.get("slip_total") == 115700.0
      and (recy.get("slip") or {}).get("source") == "dep.png")
check("portal page has search/filter and locked amount, no Excel upload",
      (lambda t: "Search / filter" in t
       and "read from the slip, locked" in t
       and "awb_excel" not in t)(client.get(f"/operator/{tok1}").text))
check("evidence-copy email skipped cleanly when SMTP is off",
      iro.send_evidence_copy({"smtp": {"enabled": False}}, "415774002",
                             "X", [], [], "t", []) == "skipped")
check("poller ignores our own evidence copies (no quarantine noise)",
      iro.process_message(
          b"From: a@b\r\nSubject: Evidence copy - IRO 415774002 - t\r\n"
          b"Message-ID: <c1@x>\r\n\r\nx")["status"] == "copy")

print("\nALL v9.4—v10.4c TESTS PASSED")
