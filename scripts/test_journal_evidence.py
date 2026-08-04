"""JOURNAL EVIDENCE TABS — the deposit slip must be VISIBLE, including PDFs.

Reported from production: after generating the journals, the Excel evidence
tab showed no deposit slip. Cause: the tab only embedded a slip when it was a
raster image, and Excel cannot embed a PDF — so a PDF slip (the common case,
since banks issue PDF bordereaux) was silently skipped, with nothing on the
tab even saying a slip existed.

Fix under test: PDFs are rasterised (first page) and embedded, and every
evidence file is listed by name whether or not it can be previewed.
"""
import io
import sys
import tempfile
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from PIL import Image  # noqa: E402
from reportlab.pdfgen import canvas  # noqa: E402

from app.tools import bitcash  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="jevid_"))
bitcash.RECON_DIR = _tmp / "recons"
bitcash.RECON_DIR.mkdir(parents=True, exist_ok=True)
bitcash.FILES_DIR = _tmp / "files"
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
bitcash.ROWS_PATH = _tmp / "rows.json"
bitcash.STORE_PATH = _tmp / "bitcash.json"

_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


# --- evidence files on disk: a PDF slip, a photo slip, an Excel deposit list
pdf_slip = bitcash.FILES_DIR / "slip_aaa111aaa111.pdf"
c = canvas.Canvas(str(pdf_slip))
c.drawString(80, 700, "BICEC - BORDEREAU DE VERSEMENT")
c.drawString(80, 680, "523,500 XAF  ref 012372")
c.save()

photo_slip = bitcash.FILES_DIR / "slip_bbb222bbb222.jpg"
Image.new("RGB", (1400, 900), (240, 240, 250)).save(photo_slip)

xls_slip = bitcash.FILES_DIR / "slip_ccc333ccc333.xlsx"
openpyxl.Workbook().save(xls_slip)

advice = bitcash.FILES_DIR / "advice_aaa111aaa111.xlsx"
openpyxl.Workbook().save(advice)

# === 1. The renderer itself ================================================
png = bitcash._evidence_preview_png(pdf_slip)
check("a PDF deposit slip renders to a PNG preview",
      png is not None and png.getvalue()[:8] == b"\x89PNG\r\n\x1a\n")
check("a photo slip is re-encoded to PNG for Excel",
      (bitcash._evidence_preview_png(photo_slip) or io.BytesIO()
       ).getvalue()[:4] == b"\x89PNG")
check("an Excel deposit list has no preview (listed, not embedded)",
      bitcash._evidence_preview_png(xls_slip) is None)
check("a missing file never raises",
      bitcash._evidence_preview_png(bitcash.FILES_DIR / "nope.pdf") is None)
big = bitcash._evidence_preview_png(photo_slip)
check("previews are scaled down for the sheet",
      Image.open(big).width <= 900)


def _recon(token, slip_name, extra=()):
    bitcash.save_recon({
        "token": token, "status": "approved", "uploaded": "2026-08-04",
        "uploaded_by": "operator:4003025705", "source": f"IRO stmt {token}",
        "account": "4003025705", "payment_refs": ["DEP-1"],
        "statement": {"label": f"IRO {token}", "date": "2026-08-01",
                      "total": 523500.0,
                      "lines": [{"awb": "4095823454", "amount": 523500.0,
                                 "matched_ids": [0]}]},
        "ar_selected": [0], "bit_selected": 0, "bit_candidates": [0],
        "rows_gen": {"bit": "g", "cash": "g"},
        "frozen": {"ar_rows": [{"id": 0, "sap_acct": "4003025705",
                                "awb": "4095823454",
                                "assignment": "4095823454",
                                "amount": 523500.0}],
                   "bit_row": {"id": 0, "gl_account": "512000",
                               "assignment": "PAY1", "amount": 523500.0,
                               "raw": ["PAY1", "523500"]},
                   "rows_gen": {"bit": "g", "cash": "g"},
                   "at": "2026-08-04 10:00"},
        "slip_total": 523500.0,
        "slip_info": {"bank": "BICEC", "depositor": "FONTEM FONKI A",
                      "amount": 523500.0, "slip_reference": "012372",
                      "date": "2026-08-01"},
        "slip": {"name": slip_name, "source": Path(slip_name).name,
                 "uploaded": "2026-08-04 10:00"} if slip_name else None,
        "extra_slips": [{"name": n, "source": Path(n).name} for n in extra],
        "file": advice.name})
    return token


_recon("aaa111aaa111", pdf_slip.name)
_recon("bbb222bbb222", photo_slip.name)
_recon("ccc333ccc333", xls_slip.name)
_recon("ddd444ddd444", "")

out = _tmp / "journal.xlsx"
res = bitcash.build_journal(out)
check("the journal builds with all four reconciliations",
      res is not None and res["count"] == 4)

wb = openpyxl.load_workbook(out)
tabs = {n: wb[n] for n in wb.sheetnames if n.startswith("EV ")}
check("one evidence tab per reconciliation", len(tabs) == 4)


def _flat(ws):
    return "\n".join(str(c.value) for row in ws.iter_rows()
                     for c in row if c.value is not None)


def _tab_for(token_frag):
    for ws in tabs.values():
        if token_frag in _flat(ws):
            return ws
    return None


# === 2. THE REPORTED BUG: a PDF slip now appears on the tab ================
pdf_ws = _tab_for("IRO stmt aaa111aaa111")
check("the PDF reconciliation has a tab", pdf_ws is not None)
check("the PDF deposit slip is EMBEDDED as a picture (was: missing)",
      pdf_ws is not None and len(pdf_ws._images) >= 1)
check("the PDF slip is also named on the tab",
      pdf_ws is not None and "slip_aaa111aaa111.pdf" in _flat(pdf_ws))
check("the tab says where the file lives",
      pdf_ws is not None and "in the journal pack ZIP" in _flat(pdf_ws))

img_ws = _tab_for("IRO stmt bbb222bbb222")
check("a photo slip is still embedded",
      img_ws is not None and len(img_ws._images) >= 1)

# === 3. Formats that cannot be previewed are still declared ================
xls_ws = _tab_for("IRO stmt ccc333ccc333")
check("an Excel slip is listed by name even though it cannot be shown",
      xls_ws is not None and "slip_ccc333ccc333.xlsx" in _flat(xls_ws))

none_ws = _tab_for("IRO stmt ddd444ddd444")
check("a reconciliation with no slip says so explicitly",
      none_ws is not None and "EVIDENCE FILES" in _flat(none_ws))

# === 4. The readings stay on the tab, and the pack still carries the file ==
check("the slip readings are still on the tab",
      pdf_ws is not None and "BICEC" in _flat(pdf_ws)
      and "012372" in _flat(pdf_ws))
pack = _tmp / "pack.zip"
bitcash.build_pack(pack, out, res["name"], res["tokens"])
names = zipfile.ZipFile(pack).namelist()
check("the original PDF is still in the pack ZIP",
      any(n.endswith(".pdf") and "Deposit slip" in n for n in names))
check("the journal itself is in the pack",
      any(n.endswith(".xlsx") and "CM01" in n for n in names))

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nALL JOURNAL-EVIDENCE TESTS PASSED")
