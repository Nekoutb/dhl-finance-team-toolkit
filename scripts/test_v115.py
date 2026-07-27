"""v11.5 — IRO statement columns, bank/payment-method selection, apply-ref-
across-AWBs, lumpsum/plug defaulting to the IRO account, and evidence tabs in
the journal. Isolated to a temp data dir; leaves real data/ untouched."""
import io
import json
import sys
import tempfile
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from app.tools import bitcash, iro  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="v115_"))
iro.IRO_DIR = _tmp / "iro"
iro.IRO_DIR.mkdir(parents=True, exist_ok=True)
iro.UPLOAD_DIR = _tmp / "uploads"
iro.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
bitcash.RECON_DIR = _tmp / "recons"
bitcash.RECON_DIR.mkdir(parents=True, exist_ok=True)
bitcash.FILES_DIR = _tmp / "files"
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
bitcash.UPLOAD_DIR = _tmp / "uploads"
bitcash.ROWS_PATH = _tmp / "bitcash_rows.json"
bitcash.STORE_PATH = _tmp / "bitcash.json"
_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


# === 1. Statement Excel — the three new columns + difference formula ========
entry = {"account": "4003026257", "name": "ETS ALINE", "total": 150000.0,
         "rows": [{"awb": "2414645273", "reference": "MTN-1",
                   "doc_no": "18000", "amount": 150000.0}]}
p = _tmp / "stmt.xlsx"
iro.build_statement_xlsx(p, entry)
ws = openpyxl.load_workbook(p).active
hdr = [c.value for c in ws[5]]
check("statement adds paid / difference / comments columns",
      hdr[5] == "Amount actually paid"
      and hdr[6] == "Difference (paid - statement)"
      and hdr[7] == "Comments")
# the difference cell is a formula referencing the paid + statement cells
diff_cell = ws.cell(row=6, column=7).value
check("difference column is an auto-computing formula",
      isinstance(diff_cell, str) and diff_cell.startswith("=")
      and "F6" in diff_cell and "D6" in diff_cell)

# === 2. Operator record remembers bank + payment method =====================
iro.set_payment_details("4003026257", bank="BICEC",
                        payment_method="Credit card")
rec = iro.load_record("4003026257")
check("operator record stores bank + payment method",
      rec.get("bank") == "BICEC" and rec.get("payment_method") == "Credit card")
check("PAYMENT_METHODS offers Bank deposit + Credit card",
      "Bank deposit" in iro.PAYMENT_METHODS
      and "Credit card" in iro.PAYMENT_METHODS)

# === 3. recon_account resolves the IRO account every way ====================
check("recon_account from stored field",
      bitcash.recon_account({"account": "777"}) == "777")
check("recon_account from uploaded_by operator:<acct>",
      bitcash.recon_account({"uploaded_by": "operator:4003026257"})
      == "4003026257")
check("recon_account from 'IRO <acct>' source label",
      bitcash.recon_account({"source": "IRO 999 — DEP1"}) == "999")
check("recon_account blank for a staff upload",
      bitcash.recon_account({"uploaded_by": "finance"}) == "")

# === 4. recon_view: plug defaults to the IRO account + auto description ======
rv = bitcash.recon_view({
    "uploaded_by": "operator:4003026257", "payment_refs": ["DEP-4471"],
    "status": "open", "statement": {"lines": [], "total": 0},
    "ar_selected": [], "bit_candidates": []})
check("plug pre-fills the reseller's account",
      rv["plug_account_default"] == "4003026257")
check("plug pre-fills 'payment difference on payment reference <ref>'",
      rv["plug_note_default"]
      == "payment difference on payment reference DEP-4471")
rv_staff = bitcash.recon_view({
    "uploaded_by": "finance", "status": "open",
    "statement": {"lines": [], "total": 0},
    "ar_selected": [], "bit_candidates": []})
check("no plug default for a staff (non-IRO) upload",
      rv_staff["plug_account_default"] == ""
      and rv_staff["plug_note_default"] == "")

# === 5. Journal evidence tabs — one named sheet per approved reconciliation ==
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment"], "gen_bit": "g1", "gen_cash": "g1",
    "bit": [{"id": 0, "gl_account": "512000", "assignment": "PAY1",
             "amount": 150000.0, "posting_date": "2026-07-12",
             "reference": "DEP-4471", "text": "DEP-4471", "doc_no": "1900045",
             "posting_key": "40", "raw": ["PAY1", "150000"]}],
    "cash": [{"id": 0, "sap_acct": "4003026257", "awb": "2414645273",
              "assignment": "2414645273", "reference": "MTN-1",
              "amount": 150000.0, "customer": "ETS ALINE", "doc_no": "18000",
              "date": "2026-07-01"}]}, ensure_ascii=False), encoding="utf-8")
_gen = bitcash.rows_generation()
bitcash.save_recon({
    "token": "abc123abc123", "status": "approved", "uploaded": "2026-07-20",
    "uploaded_by": "operator:4003026257", "source": "IRO 4003026257 — DEP-4471",
    "account": "4003026257", "operator_bank": "BICEC",
    "payment_method": "Credit card", "payment_refs": ["DEP-4471"],
    "statement": {"label": "IRO 4003026257", "date": "2026-07-01",
                  "total": 150000.0,
                  "lines": [{"awb": "2414645273", "amount": 150000.0,
                             "matched_ids": [0]}]},
    "ar_selected": [0], "bit_selected": 0, "bit_candidates": [0],
    "rows_gen": _gen,
    "frozen": {"ar_rows": [{"id": 0, "sap_acct": "4003026257",
                            "awb": "2414645273", "assignment": "2414645273",
                            "amount": 150000.0}],
               "bit_row": {"id": 0, "gl_account": "512000",
                           "assignment": "PAY1", "amount": 150000.0,
                           "raw": ["PAY1", "150000"]},
               "rows_gen": _gen, "at": "2026-07-20 10:00"},
    "slip_total": 150000.0,
    "slip_info": {"bank": "BICEC", "depositor": "ETS ALINE",
                  "amount": 150000.0, "slip_reference": "SR-1"},
    "slip": None, "extra_slips": [], "file": ""})
out = _tmp / "journal.xlsx"
res = bitcash.build_journal(out)
check("journal builds from the approved reconciliation",
      res and res["count"] >= 1 and out.exists())
jwb = openpyxl.load_workbook(out)
ev_tabs = [n for n in jwb.sheetnames if n.startswith("EV ")]
check("journal has an evidence tab named by summary",
      len(ev_tabs) == 1 and "4003026257" in ev_tabs[0]
      and "DEP-4471" in ev_tabs[0])
ev = jwb[ev_tabs[0]]
flat = "\n".join(str(c.value) for row in ev.iter_rows() for c in row
                 if c.value is not None)
check("evidence tab carries the payment ref, bank, account + AWB",
      "DEP-4471" in flat and "BICEC" in flat and "4003026257" in flat
      and "2414645273" in flat)
check("evidence tab name is within Excel's 31-char limit",
      all(len(n) <= 31 for n in jwb.sheetnames))

# === 6. Excel sheet-name sanitiser: forbidden chars + dedupe ================
n1 = bitcash._xl_sheet_name("EV a/b:c*?[x]" * 4, set())
check("sheet name strips forbidden chars + caps 31",
      len(n1) <= 31 and not any(ch in n1 for ch in '[]:*?/\\'))
n2 = bitcash._xl_sheet_name("Same", {"Same"})
check("sheet name de-duplicates", n2 != "Same" and n2.startswith("Same"))

# === 7. REVIEW FIX: a plug on the reseller's CUSTOMER account journals with a
# customer posting key (06 short / 16 excess), never a G/L key (40/50) ========
bitcash.save_recon({
    "token": "def456def456", "status": "approved", "uploaded": "2026-07-20",
    "uploaded_by": "operator:4003026257",
    "source": "IRO 4003026257 — DEP-SHORT", "account": "4003026257",
    "payment_refs": ["DEP-SHORT"],
    "statement": {"label": "IRO short", "date": "2026-07-01", "total": 150000.0,
                  "lines": [{"awb": "2414645273", "amount": 150000.0,
                             "matched_ids": [0]}]},
    "ar_selected": [0], "bit_selected": 0, "bit_candidates": [0],
    "rows_gen": _gen,
    "frozen": {"ar_rows": [{"id": 0, "sap_acct": "4003026257",
                            "awb": "2414645273", "assignment": "2414645273",
                            "amount": 150000.0}],
               "bit_row": {"id": 0, "gl_account": "512000",
                           "assignment": "PAY1", "amount": 130000.0,
                           "raw": ["PAY1", "130000"]},
               "rows_gen": _gen, "at": "2026-07-20 10:00"},
    # short payment of 20,000 plugged onto the reseller's own account
    "plug": {"amount": 20000.0, "account": "4003026257",
             "note": "payment difference on payment reference DEP-SHORT"},
    "slip_total": 130000.0, "slip": None, "extra_slips": [], "file": ""})
# drop the balanced recon so only the short one journals (clean assert)
(bitcash.RECON_DIR / "abc123abc123.json").unlink()
out2 = _tmp / "journal2.xlsx"
bitcash.build_journal(out2)
jwb2 = openpyxl.load_workbook(out2)
cm = next(s for s in jwb2.sheetnames if s.startswith("CM01_"))
ws2 = jwb2[cm]
plug_line = None
for row in range(4, ws2.max_row + 1):
    acct = str(ws2.cell(row=row, column=10).value or "")
    asg = str(ws2.cell(row=row, column=14).value or "")
    if acct == "4003026257" and "difference" in asg.lower():
        plug_line = ws2.cell(row=row, column=11).value    # posting key
# v11.8: only two posting keys are ever used — 40 and 15 (see test_v118 for
# the full pair/sign/balance checks).
check("plug on the reseller account posts on key 15", plug_line == 15)
gl_keys = [ws2.cell(row=row, column=11).value
           for row in range(4, ws2.max_row + 1)
           if str(ws2.cell(row=row, column=10).value or "") == "512000"]
check("plug BIT side posts on key 40", 40 in gl_keys)
all_keys = {ws2.cell(row=row, column=11).value
            for row in range(4, ws2.max_row + 1)
            if ws2.cell(row=row, column=10).value not in (None, "")}
check("the journal uses no posting key other than 40 and 15",
      all_keys == {40, 15})

# === 8. REVIEW FIX: bank dropdown wins over stale free-text; no "__other__" ==
from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402

# a fresh account whose AWB no approved recon covers → it stays "open"
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment"], "gen_bit": "g2", "gen_cash": "g2", "bit": [],
    "cash": [{"id": 0, "sap_acct": "9999", "awb": "8888888888",
              "assignment": "8888888888", "reference": "R", "amount": 5000.0,
              "customer": "NINE", "doc_no": "1", "date": "2026-07-01"}]},
    ensure_ascii=False), encoding="utf-8")
iro.set_payment_details("9999", bank="Ecoban")   # a typo saved first time
tok9 = iro.ensure_token("9999")["token"]
client = TestClient(main.app)

# the PORTAL table shows the paid / difference / comments columns (checked
# BEFORE any submit, while the AWB is still open)
r = client.get(f"/operator/{tok9}")
check("portal table shows the new columns",
      "Amount actually paid" in r.text and ">Difference<" in r.text
      and ">Comments<" in r.text
      and 'name="paid_8888888888"' in r.text
      and 'name="comment_8888888888"' in r.text
      and 'class="iro-diff"' in r.text)

# operator corrects to a listed bank via the dropdown; stale free-text lingers
r = client.post(f"/operator/{tok9}/submit", data={
    "awb": "8888888888", "ref_8888888888": "DEP-9", "reference": "DEP-9",
    "paid_8888888888": "5000",
    "bank": "BICEC", "bank_other": "Ecoban", "payment_method": "Bank deposit"},
    follow_redirects=False)
check("submit accepted", r.status_code == 200)
check("dropdown pick wins over stale free-text (correction kept)",
      iro.load_record("9999").get("bank") == "BICEC")
# picking "Other" with an empty text box never stores the sentinel
r = client.post(f"/operator/{tok9}/submit", data={
    "awb": "8888888888", "ref_8888888888": "DEP-9", "reference": "DEP-9",
    "paid_8888888888": "5000",
    "bank": "__other__", "bank_other": "", "payment_method": "Credit card"},
    follow_redirects=False)
check("'__other__' sentinel is never stored as the bank",
      iro.load_record("9999").get("bank") == "BICEC"    # unchanged, not sentinel
      and iro.load_record("9999").get("payment_method") == "Credit card")

# === 9. Per-AWB paid amount + comment carried into the evidence file =========
# submit with a per-AWB paid amount + comment → carried into the evidence file
ev = _tmp / "ev.xlsx"
iro.build_evidence_xlsx(ev, [{"awb": "8888888888", "amount": 5000.0,
                              "reference": "DEP-9", "amount_paid": 4800.0,
                              "comment": "short by 200"}])
ewb = openpyxl.load_workbook(ev).active
ehdr = [c.value for c in ewb[1]]
check("evidence file carries Actually paid + Comments columns",
      ehdr[3] == "Actually paid" and ehdr[4] == "Comments")
check("evidence file records the per-AWB paid amount + comment",
      ewb.cell(row=2, column=4).value == 4800.0
      and ewb.cell(row=2, column=5).value == "short by 200")

# === 10. Finance deletes a reported deposit so it is no longer a duplicate ===
iro.DEPOSITS_PATH = _tmp / "deposits.json"
iro.claim_deposit("shaA", "4003025705", 523500.0, "2026-07-07", "DEP-A", "t1")
iro.claim_deposit("shaB", "4003025705", 100000.0, "2026-07-08", "DEP-B", "t2")
check("list_deposits returns reported deposits newest-first",
      [d["sha"] for d in iro.list_deposits()] == ["shaB", "shaA"])
# the operators page shows the delete section
op = client.get("/tools/bit-cash-ar/operators")
check("operators page shows the Reported deposits delete section",
      "Reported deposits" in op.text and "DEP-A" in op.text
      and "/operators/deposit/delete" in op.text)
# delete DEP-A via the route → it can be reported again
r = client.post("/tools/bit-cash-ar/operators/deposit/delete",
                data={"sha": "shaA", "reference": "DEP-A",
                      "account": "4003025705", "at": ""},
                follow_redirects=False)
check("delete-deposit route removes it and redirects",
      r.status_code == 303 and "message=" in r.headers["location"]
      and iro.find_prior_deposit(sha="shaA") is None
      and iro.find_prior_deposit(sha="shaB") is not None)
check("the removed deposit is no longer treated as a duplicate",
      iro.claim_deposit("shaA", "4003025705", 523500.0, "2026-07-07",
                        "DEP-A2", "t3") is None)
r = client.post("/tools/bit-cash-ar/operators/deposit/delete",
                data={"sha": "nope"}, follow_redirects=False)
check("deleting a missing deposit reports not-found",
      r.status_code == 303 and "error=" in r.headers["location"])

# === 11. Operator portal: 'ticked to top' + 'only ticked' controls present ===
r = client.get(f"/operator/{tok9}") if iro.account_entry("9999") \
    else type("x", (), {"text": ""})
# (re)seed an open AWB for the portal render
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment"], "gen_bit": "g3", "gen_cash": "g3", "bit": [],
    "cash": [{"id": 0, "sap_acct": "7777", "awb": "7777777777",
              "assignment": "7777777777", "reference": "R", "amount": 9000.0,
              "customer": "SEV", "doc_no": "1", "date": "2026-07-01"}]},
    ensure_ascii=False), encoding="utf-8")
tok7 = iro.ensure_token("7777")["token"]
r = client.get(f"/operator/{tok7}")
check("portal offers 'Ticked to top' + 'Show only ticked' controls",
      'id="iro-ticktop"' in r.text and 'id="iro-onlyticked"' in r.text
      and "Ticked to top" in r.text)

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nALL v11.5 TESTS PASSED")
