"""GOLDEN JOURNAL — the safety net for the multi-tenant refactor (v12 phase 0).

The CM01 journal entry is the app's highest-stakes output: it is uploaded into
SAP. The multi-tenant work moves the company code, currency, doc type, template
path, tab naming, date encoding, column map and posting keys out of literals and
into per-country configuration. NOTHING else proves that output is unchanged.

This test builds a journal from a FIXED fixture at a FIXED timestamp and
compares it cell-by-cell — value, python type, number_format and font — against
a committed golden workbook. Run it before and after every phase.

  python scripts/test_golden_journal.py            # compare (fails on drift)
  python scripts/test_golden_journal.py --bless    # re-record the golden file
                                                   # (ONLY with a reviewed diff)

Isolated to a temp data dir; the real data/ is never touched.
"""
import json
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from app.tools import bitcash  # noqa: E402

GOLDEN = ROOT / "samples" / "golden" / "cm01_journal_golden.xlsx"
BLESS = "--bless" in sys.argv

# A fixed instant — the journal embeds it in the tab name, the post/doc dates
# and the output filename, so it MUST be pinned for the comparison to mean
# anything. 23 Jun 2026 is the date the shipped CM01 template was cut from.
NOW = datetime(2026, 6, 23, 9, 30, 0)

_tmp = Path(tempfile.mkdtemp(prefix="golden_je_"))
bitcash.RECON_DIR = _tmp / "recons"
bitcash.RECON_DIR.mkdir(parents=True, exist_ok=True)
bitcash.FILES_DIR = _tmp / "files"
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
bitcash.ROWS_PATH = _tmp / "bitcash_rows.json"
bitcash.STORE_PATH = _tmp / "bitcash.json"

_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


# --------------------------------------------------------------------------- #
# The fixture: two approved reconciliations exercising every journal branch —
# a clean multi-AWB payment, and a short payment carrying a manual plug.
# --------------------------------------------------------------------------- #
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment", "Amount", "Posting Date", "Document Number"],
    "gen_bit": "goldbit", "gen_cash": "goldcash",
    "bit": [
        {"id": 0, "gl_account": "1263001293", "assignment": "REM SGBC 0106",
         "amount": 523500.0, "posting_date": "23.06.2026",
         "reference": "DEP-4471-BICEC", "text": "REMISE CHQ", "doc_no": "410021",
         "posting_key": "40",
         "raw": ["REM SGBC 0106", "523500", "23.06.2026", "410021"]},
        {"id": 1, "gl_account": "1263001293", "assignment": "REM UBA 0206",
         "amount": 130000.0, "posting_date": "23.06.2026",
         "reference": "DEP-9999-UBA", "text": "VIREMENT", "doc_no": "410022",
         "posting_key": "40",
         "raw": ["REM UBA 0206", "130000", "23.06.2026", "410022"]},
    ],
    "cash": [
        {"id": 0, "sap_acct": "4003025929", "awb": "4095823454",
         "assignment": "4095823454", "reference": "D53104", "amount": 53639.0,
         "customer": "ACME LOGISTICS", "doc_no": "1800001",
         "date": "01.06.2026"},
        {"id": 1, "sap_acct": "4003025929", "awb": "8525614075",
         "assignment": "8525614075", "reference": "D53105", "amount": 469861.0,
         "customer": "ACME LOGISTICS", "doc_no": "1800002",
         "date": "02.06.2026"},
        {"id": 2, "sap_acct": "4003026257", "awb": "2414645273",
         "assignment": "2414645273", "reference": "MTN-17596262359",
         "amount": 150000.0, "customer": "ETS ALINE", "doc_no": "1800003",
         "date": "03.06.2026"},
    ]}, ensure_ascii=False), encoding="utf-8")
_GEN = bitcash.rows_generation()


def _recon(token, bit, ars, total, plug=None):
    rec = {
        "token": token, "status": "approved", "uploaded": "2026-06-22 10:00",
        "uploaded_by": "operator:4003025929",
        "source": f"IRO 4003025929 - {token}", "account": "4003025929",
        "payment_refs": ["DEP-4471-BICEC"],
        "statement": {"label": f"ADVICE {token}", "date": "2026-06-22",
                      "total": total,
                      "lines": [{"reference": a["awb"], "amount": a["amount"],
                                 "description": a["customer"],
                                 "matched_ids": [a["id"]]} for a in ars]},
        "ar_selected": [a["id"] for a in ars], "bit_selected": bit["id"],
        "bit_candidates": [bit["id"]], "rows_gen": _GEN,
        "frozen": {"ar_rows": ars, "bit_row": bit, "rows_gen": _GEN,
                   "at": "2026-06-22 11:00"},
        "slip_total": bit["amount"], "slip": None, "extra_slips": [],
        "file": "",
        "slip_info": {"bank": "BICEC", "depositor": "ACME LOGISTICS",
                      "beneficiary": "DHL INTERNATIONAL CAMEROUN",
                      "account_credited": "00123101000-83",
                      "slip_reference": "012372", "date": "2026-06-22",
                      "amount": bit["amount"]},
        "operator_bank": "BICEC", "payment_method": "Bank deposit"}
    if plug:
        rec["plug"] = plug
    bitcash.save_recon(rec)


_rows = json.loads(bitcash.ROWS_PATH.read_text(encoding="utf-8"))
_bit = {r["id"]: r for r in _rows["bit"]}
_cash = {r["id"]: r for r in _rows["cash"]}

# (a) exact payment across two AWBs — no plug
_recon("d01d0000aaaa", _bit[0], [_cash[0], _cash[1]], 523500.0)
# (b) short payment — 150,000 invoiced vs 130,000 banked -> 20,000 plug onto
#     the reseller's own account (posting keys 40 / 15, signed negative)
_recon("d01d0000bbbb", _bit[1], [_cash[2]], 150000.0,
       plug={"amount": 20000.0, "account": "4003026257",
             "note": "payment difference on payment reference DEP-9999-UBA"})

out = _tmp / "journal.xlsx"
res = bitcash.build_journal(out, now=NOW)
check("journal built from the fixture (2 reconciliations)",
      res is not None and res["count"] == 2 and out.exists())
check("output filename convention unchanged",
      res["name"] == "CM01_PC TEMPLATE_23.06.26_PC.xlsx")


# --------------------------------------------------------------------------- #
# Bless / compare
# --------------------------------------------------------------------------- #
def _fingerprint(path):
    """Every cell that carries meaning for SAP: coordinate -> (value, type,
    number_format, font name, font size). Sheet order included."""
    wb = openpyxl.load_workbook(path)
    snap = {"__sheets__": list(wb.sheetnames)}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cells[c.coordinate] = [
                    c.value if not isinstance(c.value, datetime)
                    else c.value.isoformat(),
                    type(c.value).__name__,
                    c.number_format,
                    (c.font.name if c.font else None),
                    (c.font.size if c.font else None),
                ]
        snap[name] = cells
    return snap


if BLESS or not GOLDEN.exists():
    GOLDEN.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(out, GOLDEN)
    (GOLDEN.with_suffix(".json")).write_text(
        json.dumps(_fingerprint(out), indent=1, ensure_ascii=False,
                   sort_keys=True), encoding="utf-8")
    print(f"\nGOLDEN RECORDED -> {GOLDEN.relative_to(ROOT)}")
    print("Commit it. Re-run without --bless to compare from now on.")
    if _fail:
        sys.exit(1)
    sys.exit(0)

want = json.loads(GOLDEN.with_suffix(".json").read_text(encoding="utf-8"))
got = _fingerprint(out)

check("sheet set + order unchanged", got["__sheets__"] == want["__sheets__"])

diffs = []
for sheet in want["__sheets__"]:
    if sheet not in got:
        diffs.append(f"{sheet}: SHEET MISSING")
        continue
    w, g = want.get(sheet, {}), got.get(sheet, {})
    for coord in sorted(set(w) | set(g)):
        if w.get(coord) != g.get(coord):
            diffs.append(f"{sheet}!{coord}: golden={w.get(coord)!r} "
                         f"now={g.get(coord)!r}")
check(f"every cell identical to the golden journal "
      f"({sum(len(want.get(s, {})) for s in want['__sheets__'])} cells)",
      not diffs)
if diffs:
    print(f"\n  {len(diffs)} DIFFERENCE(S) — first 25:")
    for d in diffs[:25]:
        print("   ", d)
    print("\n  If the change is INTENDED, review each line above, then re-record:")
    print("      python scripts/test_golden_journal.py --bless")

# Explicit assertions on the values that matter most, so a failure names the
# thing that broke rather than just a coordinate.
wb = openpyxl.load_workbook(out)
cm = wb[next(s for s in wb.sheetnames if s.startswith("CM01_"))]
rows = [r for r in cm.iter_rows(min_row=4, values_only=True)
        if r[9] not in (None, "")]
check("company code CM01 on every line", all(r[1] == "CM01" for r in rows))
check("document type DZ on every line", all(r[2] == "DZ" for r in rows))
check("currency XAF on every line", all(r[5] == "XAF" for r in rows))
check("header text 'MANUAL RECIEPT ' (sic) on every AWB line",
      all(r[7] == "MANUAL RECIEPT " for r in rows if r[11] != -20000.0))
check("the plug pair says SHORT PAYMENT and carries the payment reference",
      all(r[7] == "SHORT PAYMENT" for r in rows if r[11] == -20000.0)
      and any(r[13] == "DEP-4471-BICEC" and r[10] == 15
              for r in rows if r[11] == -20000.0))
check("SAP compact date 23626 in post + doc date",
      all(r[3] == 23626 and r[4] == 23626 for r in rows))
check("posting keys are only 40 and 15", {r[10] for r in rows} == {40, 15})
check("tab named CM01_23.06.2026_PC_001",
      "CM01_23.06.2026_PC_001" in wb.sheetnames)
# NB: the shipped template's tab is literally "BANK DETAILS " — with a
# TRAILING SPACE. build_journal matches it with .strip().upper(); any future
# refactor must keep that tolerance or the bank dump silently stops being
# written.
check("BANK DETAILS sheet present (trailing space preserved)",
      any(s.strip().upper() == "BANK DETAILS" for s in wb.sheetnames))
check("evidence tab per reconciliation, 31-char capped + de-duplicated",
      len([s for s in wb.sheetnames if s.startswith("EV ")]) == 2
      and all(len(s) <= 31 for s in wb.sheetnames))
check("the plug posts -20,000 on both 40 and 15",
      sorted(r[10] for r in rows if r[11] == -20000.0) == [15, 40])
d40 = round(sum(r[11] for r in rows if r[10] == 40), 2)
d15 = round(sum(r[11] for r in rows if r[10] == 15), 2)
check("document balances (40 side == 15 side)", d40 == d15)
check("bank side nets to the amount actually banked (523,500 + 130,000)",
      d40 == 653500.0)

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nGOLDEN JOURNAL UNCHANGED — SAP output is byte-identical")
