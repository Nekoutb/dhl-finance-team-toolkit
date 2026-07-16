"""BIT & Cash AR reconciliation sandboxes + CM01 journal-entry generation.

Synthetic BIT / Cash AR / payment statements exercise: row persistence,
auto-matching (AWB → Cash AR, total → BIT incl. the duplicate-amount warning),
selection changes, approve/unapprove, the approved-only generation gate, and
the generated workbook's structure (tab name, LI. grouping, 40/15 keys,
accounts, string assignments, compact dates, BANK DETAILS rows, patched
formulas, removed tabs).
"""
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from testutil import smtp_guard  # noqa: E402
smtp_guard()

import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402
from app.tools import bitcash  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="bitrecon_"))
bitcash.STORE_PATH = _tmp / "bitcash.json"
bitcash.ROWS_PATH = _tmp / "rows.json"
bitcash.RECON_DIR = _tmp / "recons"


def check(label, cond):
    print(f"[{'OK ' if cond else 'FAIL'}] {label}")
    if not cond:
        raise SystemExit(1)


def _wb(path, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


# --- 1. seed BIT + Cash AR files (with a DUPLICATE amount in the BIT) --------
BIT_HDR = ["Company Code", "Fiscal year/period", "Group Account Number",
           "G/L Account", "G/L Account: Long Text", "Posting Date",
           "Document Date", "Reference", "Assignment", "Text", "Posting Key",
           "Company Code Currency Value", "Company Code Currency Key",
           "Document Currency Value", "Currency", "Group Currency Value",
           "Group Currency Key", "Document Type", "Document Number", "Item"]
bit = _wb(_tmp / "bit.xlsx", BIT_HDR, [
    ["CM01", "006.2026", "126300", "1263001293", "3P Bank", "2026-07-10",
     "2026-07-10", "CMSGBCM001", "0534527700018", "VERSEMENT DHL CASH", "50",
     -294500, "XAF", -294500, "XAF", -449, "EUR", "SA", "100031871", 2],
    ["CM01", "006.2026", "126300", "1263001293", "3P Bank", "2026-07-12",
     "2026-07-12", "CMSGBCM002", "0537010000026", "VERSEMENT DHL CASH", "50",
     -294500, "XAF", -294500, "XAF", -449, "EUR", "SA", "100033277", 2],
    ["CM01", "006.2026", "126300", "1263001294", "3P Bank", "2026-07-01",
     "2026-07-01", "ORANGE C4321", "7594392514", "single amount", "40",
     73111, "XAF", 73111, "XAF", 111, "EUR", "DZ", "1400018139", 1],
])
CASH_HDR = ["C Code", "User Name", "MAC Code", "SAP Acct", "IBS Acct",
            "Customer Account: Name", "Mnth/Yr", "Doc. Date", "Doc. No.",
            "Doc. Type", "Assignment", "Reference", "Text", "Amount", "Dr/Cr"]
cash = _wb(_tmp / "cash.xlsx", CASH_HDR, [
    ["CM01", "U", "SP3", "4003025929", "415048444", "FLORENCE NIGHTINGALE",
     "7.2026", "2026-06-25", "90019001", "X4", "4596301334", "DLAR000291001",
     "", 59600, "S"],
    ["CM01", "U", "SP3", "4003025929", "415048444", "FLORENCE NIGHTINGALE",
     "7.2026", "2026-06-25", "90019002", "X4", "2718466096", "DLAR000291002",
     "", 59600, "S"],
    ["CM01", "U", "SP3", "4003025930", "415048445", "KETUM KENNETH",
     "7.2026", "2026-06-25", "90019003", "X4", "7178018330", "DLAR000291003",
     "", 59600, "S"],
    ["CM01", "U", "SP3", "4003025931", "415048446", "UNIVERSITY OF BAMENDA",
     "7.2026", "2026-06-25", "90019004", "X4", "1420812842", "DLAR000291004",
     "", 59600, "S"],
    ["CM01", "U", "SP3", "4003025932", "415048447", "ADELINE TITA",
     "7.2026", "2026-06-25", "90019005", "X4", "3032431695", "DLAR000291005",
     "", 56100, "S"],
    ["CM01", "U", "SP3", "4003026000", "415048448", "OTHER CUSTOMER",
     "7.2026", "2026-06-25", "90019006", "X4", "9999999999", "DLAR000291006",
     "", 12345, "S"],
])
bitcash.record_upload("bit", bit, "bit.xlsx")
bitcash.record_upload("cash", cash, "cash.xlsx")
store = bitcash.rows_store()
check("BIT rows persisted with GL/assignment/amount",
      len(store["bit"]) == 3 and store["bit"][0]["gl_account"] == "1263001293"
      and store["bit"][0]["assignment"] == "0534527700018"
      and store["bit"][0]["amount"] == -294500)
check("Cash AR rows persisted with SAP acct + AWB digits",
      len(store["cash"]) == 6 and store["cash"][0]["sap_acct"] == "4003025929"
      and store["cash"][0]["awb"] == "4596301334")
check("counts/history behaviour unchanged",
      bitcash.status()["current"]["bit"]["items"] == 3)

# --- 2. auto-match: AWBs -> Cash AR; total -> BIT (duplicate flagged) ---------
STMT = {"label": "EDMUND NGOCHI", "date": "25/06/2026", "total": 294500.0,
        "lines": [
            {"reference": "4596301334", "description": "FLORENCE", "amount": 59600},
            {"reference": "2718466096", "description": "FLORENCE", "amount": 59600},
            {"reference": "7178018330", "description": "KETUM", "amount": 59600},
            {"reference": "1420812842", "description": "UNIBA", "amount": 59600},
            {"reference": "3032431695", "description": "ADELINE", "amount": 56100},
            {"reference": "0000000001", "description": "GHOST", "amount": 1},
        ]}
lines, ar_sel, cands, bit_sel = bitcash.automatch(STMT)
check("5 of 6 statement lines matched in the Cash AR",
      sum(1 for l in lines if l["matched_ids"]) == 5)
check("unmatched line flagged", not lines[5]["matched_ids"])
check("5 Cash AR invoices pre-selected", len(ar_sel) == 5)
check("BIT candidates: the duplicate 294,500 appears twice",
      len(cands) == 2 and bit_sel is None)

# --- 3. sandbox via the Excel-statement path (no AI needed) -------------------
xl_stmt = _wb(_tmp / "stmt.xlsx", ["AWB Number", "Consignor", "Amount"], [
    ["4596301334", "FLORENCE NIGHTINGALE", 59600],
    ["2718466096", "FLORENCE NIGHTINGALE", 59600],
    ["7178018330", "KETUM KENNETH", 59600],
    ["1420812842", "UNIVERSITY OF BAMENDA", 59600],
    ["3032431695", "ADELINE TITA", 56100],
])
client = TestClient(main.app, follow_redirects=False)
with open(xl_stmt, "rb") as fh:
    r = client.post("/tools/bit-cash-ar/recon/upload",
                    files={"statement": ("stmt.xlsx", fh,
                                         "application/vnd.ms-excel")})
check("statement upload -> redirect to the sandbox", r.status_code == 303)
token = r.headers["location"].rstrip("/").split("/")[-1]
for _ in range(60):
    rec = bitcash.load_recon(token)
    if rec and rec["status"] in ("open", "error"):
        break
    time.sleep(0.1)
check("sandbox opened (background read finished)",
      rec and rec["status"] == "open")
check("statement total derived from the lines", rec["statement"]["total"] == 294500)
check("duplicate BIT amount flagged on the sandbox",
      rec["bit_duplicate"] and len(rec["bit_candidates"]) == 2
      and rec["bit_selected"] is None)
check("uploader recorded", rec.get("uploaded_by") == "")

page = client.get(f"/tools/bit-cash-ar/recon/{token}")
check("sandbox page renders invoice + payment sections",
      page.status_code == 200 and "Invoice lines" in page.text
      and "Payment lines" in page.text)
check("duplicate warning shown", "appears" in page.text
      and "times" in page.text)
check("matched lines highlighted", "auto-match" in page.text)

# --- 4. approval requires a selected BIT line --------------------------------
r = client.post(f"/tools/bit-cash-ar/recon/{token}/approve")
check("approve blocked until the BIT payment is selected",
      r.status_code == 303 and "error=" in r.headers["location"])

# select the first BIT candidate + keep the 5 invoices
rec = bitcash.load_recon(token)
r = client.post(f"/tools/bit-cash-ar/recon/{token}/select",
                data={"ar": [str(i) for i in rec["ar_selected"]],
                      "bit": str(rec["bit_candidates"][0])})
check("selection saved", r.status_code == 303 and "message=" in r.headers["location"])
view = bitcash.recon_view(bitcash.load_recon(token))
check("difference is zero (5 invoices = payment)",
      view["ar_total"] == 294500 and view["bit_amount"] == 294500
      and view["difference"] == 0)

r = client.post(f"/tools/bit-cash-ar/recon/{token}/approve")
check("approve OK once complete", "message=" in r.headers["location"])
check("status approved", bitcash.load_recon(token)["status"] == "approved")

# approved sandboxes can't be deleted; unapprove works and reopens
r = client.post(f"/tools/bit-cash-ar/recon/{token}/delete")
check("approved sandbox can't be deleted",
      "error=" in r.headers["location"]
      and bitcash.load_recon(token) is not None)
client.post(f"/tools/bit-cash-ar/recon/{token}/unapprove")
check("unapprove reopens", bitcash.load_recon(token)["status"] == "open")
client.post(f"/tools/bit-cash-ar/recon/{token}/approve")

# --- 5. journal generation (approved only) -----------------------------------
out = _tmp / "journal.xlsx"
now = datetime(2026, 7, 16, 10, 0)
result = bitcash.build_journal(out, now=now)
check("journal built for the approved sandbox",
      result and result["count"] == 1 and result["lines"] == 6)
check("download name per the team's convention",
      result["name"] == "CM01_PC TEMPLATE_16.07.26_PC.xlsx")

wb = openpyxl.load_workbook(out)
tab = "CM01_16.07.2026_PC_001"
check("CM01 tab renamed to today's date", tab in wb.sheetnames)
check("Guide + IC tabs removed",
      "Guide" not in wb.sheetnames and "List of IC Accounts" not in wb.sheetnames)
ws = wb[tab]
rows = [[ws.cell(row=r, column=c).value for c in range(1, 15)]
        for r in range(4, 10)]
check("LI. repeats 1 on all 6 lines", all(r[0] == 1 for r in rows))
check("constants CM01/DZ/XAF + MANUAL RECIEPT",
      rows[0][1] == "CM01" and rows[0][2] == "DZ" and rows[0][5] == "XAF"
      and rows[0][6] == "MANUAL RECIEPT ")
check("compact date 16.07.2026 -> 16726",
      rows[0][3] == 16726 and rows[0][4] == 16726)
check("line 1 = BIT: GL account, key 40, abs amount, BIT assignment",
      rows[0][9] == 1263001293 and rows[0][10] == 40
      and rows[0][11] == 294500 and rows[0][13] == "0534527700018")
check("lines 2-6 = AR: SAP accounts, key 15, per-AWB amounts",
      all(r[10] == 15 for r in rows[1:])
      and rows[1][9] == 4003025929 and rows[1][11] == 59600
      and rows[5][11] == 56100)
check("assignments are strings preserving digits",
      all(isinstance(r[13], str) for r in rows))
amounts_40 = sum(r[11] for r in rows if r[10] == 40)
amounts_15 = sum(r[11] for r in rows if r[10] == 15)
check("document balances (40 total == 15 total)", amounts_40 == amounts_15)

# formatting harmonised: frozen headers, Arial 10, no stray fills on data rows
check("CM01 headers frozen at row 4", ws.freeze_panes == "A4")
c4 = ws.cell(row=4, column=2)
check("journal lines written in Arial 10",
      c4.font.name == "Arial" and c4.font.size == 10)
check("no leftover colour on data rows", ws.cell(row=5, column=12).fill.fill_type is None)

bd = wb["BANK DETAILS "]
check("BANK DETAILS lists the journal's BIT line",
      str(bd.cell(row=2, column=9).value) == "0534527700018")
check("BANK DETAILS header frozen + Arial",
      bd.freeze_panes == "A2" and bd.cell(row=2, column=1).font.name == "Arial")
ck = wb["CHECK"]
f = ck.cell(row=5, column=3).value
check("CHECK formulas re-pointed at the renamed tab",
      tab in f and "23.06.2026" not in f)
dl = wb["Doc Level Check"]
f2 = dl.cell(row=2, column=2).value
check("Doc Level Check formulas re-pointed too",
      tab in f2 and "23.06.2026" not in f2)

# nothing approved -> no file
client.post(f"/tools/bit-cash-ar/recon/{token}/unapprove")
check("generation gate: nothing approved -> refused",
      bitcash.build_journal(_tmp / "none.xlsx") is None)
r = client.post("/tools/bit-cash-ar/journal")
check("journal route redirects with a friendly error when nothing is approved",
      r.status_code == 303 and "error=" in r.headers["location"])

print("\nALL BIT & CASH AR RECONCILIATION TESTS PASSED")
