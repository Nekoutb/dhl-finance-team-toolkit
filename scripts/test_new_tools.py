"""Quick Account Statement (one-click customer statement, Excel only) and
BIT & Cash AR (open-items counting + daily graph + file freshness)."""
import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from testutil import smtp_guard  # noqa: E402
smtp_guard()

import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402
from app.services import ar_master  # noqa: E402
from app.tools import bitcash, ongoing_ctp as ctp  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
MASTER = ROOT / "samples" / "ar_master_sample.xlsx"
TXN = ROOT / "samples" / "ar_breakdown_features_sample.xlsx"
_tmp = Path(tempfile.mkdtemp(prefix="newtools_"))


def check(label, cond):
    print(f"[{'OK ' if cond else 'FAIL'}] {label}")
    if not cond:
        raise SystemExit(1)


client = TestClient(main.app, follow_redirects=False)

# === 1. Quick Account Statement (Excel only, same-name combining) ===========
from urllib.parse import parse_qs, urlparse  # noqa: E402

from app.tools import quickstmt  # noqa: E402

res = ctp.analyze(TXN, as_of=date(2026, 6, 13), default_rank=60,
                  master=ar_master.parse_master(MASTER))
res["created_at"] = "2099-01-01 00:00"     # ensure it is the latest analysis
res["source"] = "features.xlsx"
# A second account with the SAME customer name -> must be combined (iff rule);
# give it one open invoice so the statement carries both account numbers.
acme = next(c for c in res["customers"] if c["key"] == "1004000001")
twin = dict(acme, key="1004999999", total_ar=111000.0)
res["customers"].append(twin)
res["invoices"].append(dict(
    next(i for i in res["invoices"]
         if str(i.get("account")) == "1004000001" and i.get("kind") == "invoice"),
    account="1004999999", invoice_no="TWIN-0001", amount=111000.0))
ctp.save_result("aab0c5522", res)
try:
    page = client.get("/tools/quick-statement")
    check("quick statement page 200", page.status_code == 200)
    check("customer list offered", "qs-customers" in page.text
          and "1004000001" in page.text)
    check("same-name sibling flagged in the picker",
          "same name — combined" in page.text)

    # iff rule: same name -> combined; different name -> alone
    d = quickstmt.statement_data(res, "1004000001")
    check("same-name accounts combined (iff rule)",
          d["combined"] and set(d["accounts"]) == {"1004000001", "1004999999"})
    d2 = quickstmt.statement_data(res, "1004000002")
    check("different names are NEVER combined",
          d2 and not d2["combined"] and d2["accounts"] == ["1004000002"])
    check("rows carry the requested detail",
          all(k in d["rows"][0] for k in ("reference", "issue_date", "due_date",
                                          "amount", "payment_terms", "ageing",
                                          "account")))
    check("ageing = days since the invoice issue date",
          quickstmt.ageing_days("2026-05-20", date(2026, 6, 13)) == 24
          and quickstmt.ageing_days("2026-01-01", date(2026, 6, 13)) == 163
          and quickstmt.ageing_days("") is None)
    check("row ageing is numeric (days) and issue date filled",
          isinstance(d["rows"][0]["ageing"], int) and d["rows"][0]["issue_date"])
    check("payment terms carried on every row (TB value)",
          all(r_["payment_terms"] for r_ in d["rows"]))

    r = client.get("/tools/quick-statement/generate?key=1004000001",
                   follow_redirects=False)
    check("generate -> redirect with the Excel only", r.status_code == 303
          and "xlsx=" in r.headers["location"]
          and "pdf=" not in r.headers["location"])
    qs = parse_qs(urlparse(r.headers["location"]).query)
    xlsx_name = qs["xlsx"][0]

    from io import BytesIO
    rx = client.get(f"/download/{xlsx_name}")
    check("Excel downloads", rx.status_code == 200 and rx.content[:2] == b"PK")
    import openpyxl as _px
    wbx = _px.load_workbook(BytesIO(rx.content))
    wsx = wbx["Statement"]
    all_text = " ".join(str(c.value) for row in wsx.iter_rows()
                        for c in row if c.value is not None)
    check("Excel carries the details + both accounts",
          "Référence facture" in all_text and "Date d'émission" in all_text
          and "Ageing (jours)" in all_text and "Termes de paiement" in all_text
          and "1004999999" in all_text and "TWIN-0001" in all_text)
    check("no letter, no logo — clean statement",
          "Cher Client" not in all_text and not getattr(wsx, "_images", []))
    check("Excel header frozen", wsx.freeze_panes not in (None, "A1"))

    r = client.get("/tools/quick-statement/generate?key=nope",
                   follow_redirects=False)
    check("unknown customer -> friendly redirect",
          r.status_code == 303 and "error=" in r.headers["location"])

    # payment terms flow from the master into the analysis customers
    check("payment_term attached from the master",
          any(c.get("payment_term") for c in res["customers"]))
finally:
    (ctp.STORE_DIR / "aab0c5522.json").unlink(missing_ok=True)
print("ok: quick account statement — Excel only, days ageing, same-name combining")

# === 2. BIT & Cash AR ========================================================
bitcash.STORE_PATH = _tmp / "bitcash.json"


def _wb(rows, header):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r_ in rows:
        ws.append(r_)
    p = _tmp / f"{header[0]}_{len(rows)}.xlsx"
    wb.save(p)
    return p

# BIT file WITH a status column: 2 open, 1 closed
bit = _wb([["T1", 100, "Open"], ["T2", 200, "Closed"], ["T3", 300, "ouvert"]],
          ["Ref", "Amount", "Status"])
# Cash AR file WITHOUT a status column: every row is an open item
cash = _wb([["C1", 10], ["C2", 20], ["C3", 30], ["C4", 40]],
           ["Item", "Amount"])

c1 = bitcash.count_items(bit)
check("status column respected (2 open of 3)",
      c1 == {"items": 3, "open_items": 2, "status_col": "Status"})
c2 = bitcash.count_items(cash)
check("no status column -> all rows open (4)",
      c2["items"] == 4 and c2["open_items"] == 4)

with open(bit, "rb") as f1, open(cash, "rb") as f2:
    r = client.post("/tools/bit-cash-ar/upload",
                    files={"bit_file": ("bit.xlsx", f1,
                                        "application/vnd.ms-excel"),
                           "cash_file": ("cash_ar.xlsx", f2,
                                         "application/vnd.ms-excel")})
check("upload both files -> redirect", r.status_code == 303)
st = bitcash.status()
check("current BIT position stored", st["current"]["bit"]["open_items"] == 2)
check("current Cash AR position stored", st["current"]["cash"]["open_items"] == 4)
check("today's history point recorded",
      st["history"] and st["history"][-1][1] == {"bit_open": 2, "cash_open": 4})

page = client.get("/tools/bit-cash-ar")
check("page shows open counts", ">2<" in page.text and ">4<" in page.text)
check("daily graph plotted", "<svg" in page.text
      and "Open BIT items" in page.text)
check("file freshness banners shown, no age bracket",
      "BIT on record:" in page.text and "Cash AR on record:" in page.text
      and "(today)" not in page.text and "day old" not in page.text)

r = client.post("/tools/bit-cash-ar/upload")
check("no file -> friendly 400", r.status_code == 400)
print("ok: BIT & Cash AR — open items counted per day + daily graph")

print("\nALL NEW-TOOLS TESTS PASSED")
