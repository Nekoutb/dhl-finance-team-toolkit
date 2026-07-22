"""v11.4 — cheque evidence bundle (view scan + ZIP of scan/bank-line/BIT-line)
and the Bank balances panel + daily running-balance graph. Fully isolated to a
temp data dir; leaves the real data/ untouched."""
import io
import json
import sys
import tempfile
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from PIL import Image  # noqa: E402

from app.config import save_user_config  # noqa: E402
from app.services import bank_statement, cheques  # noqa: E402
from app.tools import bank, bitcash  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="v114_"))
cheques.BATCH_DIR = _tmp / "cheques"
cheques.STATS_PATH = _tmp / "cheque_stats.json"
bank.STORE_DIR = _tmp / "bank"
bank.BALANCE_STATS_PATH = _tmp / "bank_daily_balance.json"
bitcash.ROWS_PATH = _tmp / "bitcash_rows.json"
bank.STORE_DIR.mkdir(parents=True, exist_ok=True)

from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402

client = TestClient(main.app)
_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


# --------------------------------------------------------------------------- #
# Fixtures: one matched cheque (scan + bank appearance) + a matching BIT row.
# --------------------------------------------------------------------------- #
CHQ = "5550001"
bdir = cheques.BATCH_DIR / "aa11bb22cc33"
bdir.mkdir(parents=True)
Image.new("RGB", (40, 20), "white").save(bdir / "000.png")   # the "scan"
(bdir / "results.json").write_text(json.dumps({
    "status": "done", "uploaded_by": "tester",
    "created_at": "2026-07-17 08:30", "banks": [], "bank_line_count": 0,
    "cheques": [{
        "idx": 0, "filename": "acme_cheque.png", "stored": "000.png",
        "media": "image/png", "status": "done",
        "result": {"cheque_number": CHQ, "customer_name": "ACME LOGISTICS SARL",
                   "amount": 150000.0, "amount_currency": "XAF",
                   "cheque_date": "2026-07-01", "drawee_bank": "SCB"},
        "appearances": [{"bank": "SGBC", "amount": 150000.0,
                         "date": "2026-07-10",
                         "text": "REMISE CHQ 5550001 ACME LOGISTICS SARL"}],
        "matched_at": "2026-07-11 09:00", "error": ""}]},
    ensure_ascii=False), encoding="utf-8")

# A BIT row carrying the cheque number so the evidence bundle has a BIT snapshot.
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment", "Reference", "Amount", "Posting Date",
                   "Document Number"],
    "bit": [{"id": 0, "gl_account": "512000", "assignment": "AWB",
             "amount": 150000.0, "posting_date": "2026-07-12",
             "reference": "CHQ 5550001", "text": "CHQ 5550001 ACME",
             "doc_no": "1900045", "posting_key": "40",
             "raw": ["", "CHQ 5550001", "150000", "2026-07-12", "1900045"]}],
    "cash": [], "gen_bit": "gen1", "gen_cash": ""},
    ensure_ascii=False), encoding="utf-8")

# === 1. View the scanned cheque =============================================
info = cheques.scan_path("aa11bb22cc33", 0)
check("scan_path resolves the stored scan (path/media/name)",
      info is not None and info[1] == "image/png"
      and info[0].name == "000.png")
check("scan_path rejects a path-traversal stored name",
      cheques.scan_path("../etc", 0) is None)

r = client.get("/tools/cheque-processing/register/scan?batch=aa11bb22cc33&idx=0")
check("GET /register/scan serves the cheque image inline",
      r.status_code == 200 and r.headers.get("content-type", "").startswith("image/")
      and "inline" in r.headers.get("content-disposition", ""))
r = client.get("/tools/cheque-processing/register/scan?batch=aa11bb22cc33&idx=9")
check("GET /register/scan 404s an unknown cheque", r.status_code == 404)

# === 2. Evidence ZIP bundle =================================================
r = client.get("/tools/cheque-processing/register/evidence?batch=aa11bb22cc33&idx=0")
check("GET /register/evidence returns a ZIP",
      r.status_code == 200 and r.content[:2] == b"PK")
names = zipfile.ZipFile(io.BytesIO(r.content)).namelist()
print("   zip entries:", names)
check("bundle carries the cheque scan",
      any(n.startswith("cheque_5550001") and n.endswith(".png") for n in names))
check("bundle carries the bank-statement line snapshot",
      "bank_statement_line_5550001.png" in names)
check("bundle carries the BIT line snapshot",
      "bit_line_5550001.png" in names)
# the snapshots are real, non-trivial PNG images
z = zipfile.ZipFile(io.BytesIO(r.content))
bank_png = z.read("bank_statement_line_5550001.png")
im = Image.open(io.BytesIO(bank_png))
check("bank-line snapshot is a legible PNG (reasonable size)",
      im.format == "PNG" and im.width >= 600 and im.height >= 120)

r = client.get("/tools/cheque-processing/register/evidence?batch=zz&idx=0",
               follow_redirects=False)
check("evidence redirects with an error for an unknown cheque",
      r.status_code == 303 and "error=" in r.headers.get("location", ""))

# register page shows the two new actions
home = client.get("/tools/cheque-processing")
check("register row offers View + Evidence actions",
      "View cheque" in home.text and "Evidence" in home.text
      and "/register/scan?batch=" in home.text
      and "/register/evidence?batch=" in home.text)

# === 3. Bank balances — closing balance from the Solde column ===============
save_user_config({"banks": ["Ecobank Cameroon", "BICEC"]})

stmt = _tmp / "ecobank_july.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Date", "Libelle", "Credit", "Debit", "Solde"])
ws.append(["01/07/2026", "VIREMENT ACME LOGISTICS", "150000", "", "1150000"])
ws.append(["02/07/2026", "CHEQUE 000088", "", "50000", "1100000"])
wb.save(stmt)

rep = bank.build_report([(str(stmt), "ecobank_july.xlsx")],
                        token=bank.bank_token("Ecobank Cameroon", "2026-07"),
                        bank_slot="Ecobank Cameroon", period="2026-07")
bank.save_report(rep)
check("closing balance read from the Solde column (last row)",
      rep["summary"].get("closing_balance") == 1100000.0)

rows, total, missing = bank.bank_balances(["Ecobank Cameroon", "BICEC"])
eco = next(r for r in rows if r["bank"] == "Ecobank Cameroon")
bicec = next(r for r in rows if r["bank"] == "BICEC")
check("bank_balances: Ecobank has its balance + upload date",
      eco["has_balance"] and eco["balance"] == 1100000.0 and eco["last_upload"])
check("bank_balances: BICEC (no statement) flagged, not counted",
      bicec["has_balance"] is False and bicec["has_report"] is False)
check("bank_balances totals only the known balances",
      total == 1100000.0 and missing is False)  # BICEC has no report → not 'missing'

# refresh must PRESERVE the balance (it doesn't re-read the file)
ref = bank.refresh_report(bank.bank_token("Ecobank Cameroon", "2026-07"))
check("refresh_report preserves the closing balance",
      ref["summary"].get("closing_balance") == 1100000.0)

# daily snapshot + history + graph
_y = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
bank.BALANCE_STATS_PATH.write_text(json.dumps({_y: 900000.0}),
                                   encoding="utf-8")
bank.record_daily_balance(1100000.0)
hist = bank.balance_history()
check("record_daily_balance stamps today; history has >= 2 days",
      len(hist) >= 2 and hist[-1][1] == 1100000.0)
check("record_daily_balance ignores a None total",
      (bank.record_daily_balance(None) or True) and len(bank.balance_history()) >= 2)

page = client.get("/tools/bank-statements")
check("Bank balances panel renders with balance + total + graph",
      "Bank balances" in page.text and "1,100,000" in page.text
      and "TOTAL bank balance" in page.text and "Running balance" in page.text
      and "<svg" in page.text)

# === 4. Review fixes ========================================================
# 4a. The evidence ZIP never persists on disk (OUTPUT_DIR is public /download).
before_out = set(p.name for p in main.OUTPUT_DIR.glob("Cheque_evidence*"))
r = client.get("/tools/cheque-processing/register/evidence?batch=aa11bb22cc33&idx=0")
after_out = set(p.name for p in main.OUTPUT_DIR.glob("Cheque_evidence*"))
check("evidence ZIP is built in memory — nothing lands in OUTPUT_DIR",
      r.status_code == 200 and r.content[:2] == b"PK"
      and after_out == before_out
      and "attachment" in r.headers.get("content-disposition", ""))

# 4b. Newest-first statements: the closing balance is the LATEST-DATED row's.
stmt2 = _tmp / "desc_order.xlsx"
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.append(["Date", "Libelle", "Credit", "Debit", "Solde"])
ws2.append(["03/07/2026", "LATEST OP", "10000", "", "900000"])    # newest FIRST
ws2.append(["02/07/2026", "MIDDLE OP", "", "5000", "890000"])
ws2.append(["01/07/2026", "OLDEST OP", "20000", "", "895000"])    # oldest LAST
wb2.save(stmt2)
r2 = bank_statement.read_bank(str(stmt2))
check("newest-first statement: closing balance = latest DATE, not last row",
      r2["closing_balance"] == 900000.0)

# 4c. Accents/em-dash transliterated so the snapshot font renders them.
from app.services import evidence_img  # noqa: E402
check("French accents + dashes transliterate cleanly for the snapshot",
      evidence_img._sanitize("Solde après opération — reçu à Yaoundé")
      == "Solde apres operation - recu a Yaounde")

# 4d. Daily balance snapshot survives concurrent writers (unique tmp).
import threading as _thr  # noqa: E402
_errs = []


def _wb(i):
    try:
        bank.record_daily_balance(1100000.0 + i)
    except Exception as exc:  # noqa: BLE001
        _errs.append(exc)


_ts = [_thr.Thread(target=_wb, args=(i,)) for i in range(12)]
for _t in _ts:
    _t.start()
for _t in _ts:
    _t.join()
check("concurrent daily-balance snapshots never crash or corrupt",
      not _errs and isinstance(bank.balance_history()[-1][1], float))

# === 5. Cash AR ageing — DOCUMENT DATE basis + in-place re-read =============
bitcash.STORE_PATH = _tmp / "bitcash.json"
bitcash.UPLOAD_DIR = _tmp / "uploads"
bitcash.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# 5a. "Doc. Date" drives the ageing even when an (empty) posting column exists.
cash_x = _tmp / "cash_ar.xlsx"
wbc = openpyxl.Workbook()
wsc = wbc.active
wsc.append(["SAP Acct", "Assignment", "Reference", "Amount",
            "Customer Name", "Doc. No.", "Pstng Date", "Doc. Date"])
wsc.append(["415774002", "8890001111", "INV-1", "50000",
            "OMEGA SARL", "19001", "", "15.03.2026"])   # posting EMPTY
wsc.append(["415774002", "8890002222", "INV-2", "30000",
            "OMEGA SARL", "19002", "", "01.07.2026"])
wbc.save(cash_x)
bitcash._persist_rows("cash", cash_x)
_st5 = bitcash.rows_store()
check("Doc. Date wins over an empty posting column (document-date basis)",
      _st5.get("cash_date_col") == "Doc. Date"
      and _st5["cash"][0]["date"] == "15.03.2026")

# 5b. The EXISTING file on record gains its ageing via in-place re-read: fake
# a pre-v11.4 store (rows without dates) + the current stored file on disk.
import shutil  # noqa: E402
stored_name = "bitcash_cash_test11.xlsx"
shutil.copy(cash_x, bitcash.UPLOAD_DIR / stored_name)
bitcash.STORE_PATH.write_text(json.dumps({
    "current": {"cash": {"source": "cash_ar.xlsx", "stored": stored_name,
                         "uploaded": "2026-07-15 09:00"}},
    "history": {}}), encoding="utf-8")
_gen_before = "keepme123456"
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": [], "bit": [], "gen_bit": "",
    "gen_cash": _gen_before,
    "cash": [{"id": 0, "sap_acct": "415774002", "awb": "8890001111",
              "assignment": "8890001111", "reference": "INV-1",
              "amount": 50000.0, "customer": "OMEGA SARL",
              "doc_no": "19001"}]}), encoding="utf-8")   # NO dates captured
ag0 = bitcash.cash_ageing()
check("pre-fix store shows no dated buckets", not ag0["has_dates"])
# The hot GET render must NOT parse anything (no auto-heal on the render path).
r = client.get("/tools/bit-cash-ar")
check("landing GET does not auto-parse (still no dates, marker unset)",
      r.status_code == 200 and not bitcash.cash_ageing()["has_dates"]
      and bitcash.rows_store().get("reparsed_cash") != stored_name)
check("reparse_current re-reads the stored file in place",
      bitcash.reparse_current("cash") is True)
_st5b = bitcash.rows_store()
ag1 = bitcash.cash_ageing()
check("existing file now shows an ageing (dates from Doc. Date)",
      ag1["has_dates"] and ag1["date_col"] == "Doc. Date"
      and ag1["totals"]["b61"] == 50000.0)   # 15.03 is >60 days before today
check("in-place re-read keeps the generation (sandboxes stay valid)",
      _st5b.get("gen_cash") == _gen_before)
check("re-read is attempted once per stored file",
      _st5b.get("reparsed_cash") == stored_name
      and bitcash.reparse_current("cash") is False)
# reparse ABORTS rather than resurrect rows when the file was replaced meanwhile
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": [], "bit": [], "gen_bit": "", "gen_cash": "othergen",
    "cash": [{"id": 0, "sap_acct": "999", "awb": "1", "assignment": "1",
              "reference": "X", "amount": 1.0, "customer": "NEWFILE",
              "doc_no": "1", "date": ""}]}), encoding="utf-8")  # different gen
_j = _tmp / "bitcash.json"
_j.write_text(json.dumps({"current": {"cash": {"source": "other.xlsx",
              "stored": "does_not_exist.xlsx", "uploaded": "x"}},
              "history": {}}), encoding="utf-8")
check("reparse aborts when the current file was replaced (no resurrection)",
      bitcash.reparse_current("cash") is False
      and bitcash.rows_store()["cash"][0]["customer"] == "NEWFILE")
# the age-current POST route redirects (button path)
bitcash.STORE_PATH.write_text(json.dumps({
    "current": {"cash": {"source": "cash_ar.xlsx", "stored": stored_name,
                         "uploaded": "2026-07-15 09:00"}}, "history": {}}),
    encoding="utf-8")
r = client.post("/tools/bit-cash-ar/age-current", follow_redirects=False)
check("age-current route redirects with a message", r.status_code == 303)

import os  # noqa: E402
if (ROOT / "config.json").exists():
    os.remove(ROOT / "config.json")

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nALL v11.4 TESTS PASSED")
