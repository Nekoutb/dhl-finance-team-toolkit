"""Prove legacy Excel 97-2003 (.xls) files work across the app.

Builds an .xls version of the Orange sample with xlwt, then runs it through
the shared reader AND a real upload to the Orange tool.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import xlwt
from fastapi.testclient import TestClient

from app.main import app
from app.services import excel_reader

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "samples" / "orange_cameroun_sample.xlsx"
DST = ROOT / "samples" / "orange_cameroun_sample.xls"


def check(label, cond):
    print(f"[{'OK ' if cond else 'FAIL'}] {label}")
    if not cond:
        raise SystemExit(1)


# Convert the xlsx sample to a genuine .xls
wb_in = openpyxl.load_workbook(SRC, data_only=True)
ws_in = wb_in.active
wb_out = xlwt.Workbook()
ws_out = wb_out.add_sheet("Transactions")
for r, row in enumerate(ws_in.iter_rows(values_only=True)):
    for c, val in enumerate(row):
        if val is not None:
            ws_out.write(r, c, val)
wb_out.save(str(DST))
print(f"built {DST.name}")

# Shared reader parses it identically
parsed = excel_reader.read_transactions(DST)
check(".xls parsed: header detected", "MSISDN" in parsed["header"])
check(".xls parsed: 4 transaction rows", len(parsed["rows"]) == 4)
check(".xls parsed: metadata kept", any("Orange" in m for m in parsed["metadata"]))
amounts = [r["data"].get("Amount (XAF)") for r in parsed["rows"]]
check(".xls numbers are clean ints", amounts[0] == 150000)

# Real upload through the web app
client = TestClient(app)
with open(DST, "rb") as fh:
    r = client.post("/tools/orange-cameroun/upload",
                    files={"file": ("orange_cameroun_sample.xls", fh,
                                    "application/vnd.ms-excel")})
check(".xls upload -> review 200", r.status_code == 200)
check(".xls review shows transactions", "transaction(s) found" in r.text
      and "MSISDN" in r.text)

print("\nALL .XLS CHECKS PASSED")
for p in (ROOT / "data" / "uploads").glob("*.xls"):
    p.unlink()
