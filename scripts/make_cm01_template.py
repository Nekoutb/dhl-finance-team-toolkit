"""Build the bundled CM01 journal-entry template from the team's master file.

Reads the original CM01_PC TEMPLATE workbook and produces
samples/cm01_pc_template.xlsx with:
  - the CM01_* tab's sample data cleared (rows 4+),
  - colours harmonised: leftover fills stripped from the data area, every
    cell standardised to Arial 10 (headers keep their fills, bold retained),
  - headers frozen (CM01 tab at row 4, BANK DETAILS below its header),
  - BANK DETAILS cleared below its header row,
  - Sheet1 (sample paste area) cleared,
  - "Guide" and "List of IC Accounts" tabs removed,
  - CHECK + Doc Level Check formulas kept untouched.

Run (one-off, whenever the team's master template changes):
    python scripts/make_cm01_template.py "<path to the master template>"
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "samples" / "cm01_pc_template.xlsx"

DEFAULT_SRC = r"C:\Users\UltraBook 3.1\Desktop\CM01_PC TEMPLATE_.23.06.26_PC.xlsx"

_NO_FILL = PatternFill(fill_type=None)


def _arial(cell, bold=None):
    old = cell.font
    cell.font = Font(name="Arial", size=10,
                     bold=old.bold if bold is None else bold,
                     italic=old.italic, color=old.color)


def main(src):
    wb = openpyxl.load_workbook(src)          # keep formulas (no data_only)
    main_tab = next(n for n in wb.sheetnames if n.startswith("CM01_"))

    # 1. CM01 tab: clear the sample JE lines (row 4 down), strip their leftover
    #    fills, standardise Arial 10 everywhere, freeze the header rows.
    ws = wb[main_tab]
    for row in ws.iter_rows(min_row=1, max_row=3):
        for c in row:
            _arial(c)                          # headers keep fill + bold
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for c in row:
            c.value = None
            c.fill = _NO_FILL                  # harmonise: no stray colours
            _arial(c, bold=False)
    ws.freeze_panes = "A4"

    # 2. BANK DETAILS: clear below the header, same harmonisation
    bd = next((n for n in wb.sheetnames if n.strip().upper() == "BANK DETAILS"),
              None)
    if bd:
        ws = wb[bd]
        for c in ws[1]:
            _arial(c)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row:
                c.value = None
                c.fill = _NO_FILL
                _arial(c, bold=False)
        ws.freeze_panes = "A2"

    # 3. clear Sheet1 (sample paste area)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for row in ws.iter_rows():
            for c in row:
                c.value = None
                c.fill = _NO_FILL

    # 4. standardise the check/fields tabs to Arial 10 (fills untouched)
    for name in ("CHECK", "Doc Level Check", "Fields"):
        if name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for c in row:
                    _arial(c)

    # 5. drop the tabs the team asked to remove
    for name in ("Guide", "List of IC Accounts"):
        if name in wb.sheetnames:
            del wb[name]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print("  tabs:", openpyxl.load_workbook(OUT).sheetnames)


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC)
