"""v11.8 — pick which journals to generate, plug on posting keys 40/15 only,
IRO must state the amount actually paid (and explain any variance), IRO can
delete unsent payment evidence, and the IRO emails are three lines.
Isolated to a temp data dir; leaves the real data/ untouched."""
import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from app.tools import bitcash, iro  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="v118_"))
iro.IRO_DIR = _tmp / "iro"
iro.IRO_DIR.mkdir(parents=True, exist_ok=True)
iro.UPLOAD_DIR = _tmp / "uploads"
iro.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
iro.DEPOSITS_PATH = _tmp / "deposits.json"
bitcash.RECON_DIR = _tmp / "recons"
bitcash.RECON_DIR.mkdir(parents=True, exist_ok=True)
bitcash.FILES_DIR = _tmp / "files"
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
bitcash.ROWS_PATH = _tmp / "bitcash_rows.json"
bitcash.STORE_PATH = _tmp / "bitcash.json"
_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": ["Assignment"], "gen_bit": "g1", "gen_cash": "g1",
    "bit": [{"id": 0, "gl_account": "512000", "assignment": "PAY1",
             "amount": 130000.0, "posting_date": "2026-07-12",
             "reference": "DEP-1", "text": "DEP-1", "doc_no": "19",
             "posting_key": "40", "raw": ["PAY1", "130000"]},
            {"id": 1, "gl_account": "512000", "assignment": "PAY2",
             "amount": 50000.0, "posting_date": "2026-07-13",
             "reference": "DEP-2", "text": "DEP-2", "doc_no": "20",
             "posting_key": "40", "raw": ["PAY2", "50000"]}],
    "cash": [{"id": 0, "sap_acct": "4003025705", "awb": "1005475881",
              "assignment": "1005475881", "reference": "D53104",
              "amount": 150000.0, "customer": "SMIC", "doc_no": "1",
              "date": "2026-07-01"},
             {"id": 1, "sap_acct": "4003025705", "awb": "1011864556",
              "assignment": "1011864556", "reference": "R2",
              "amount": 50000.0, "customer": "SMIC", "doc_no": "2",
              "date": "2026-07-02"}]}, ensure_ascii=False), encoding="utf-8")
_gen = bitcash.rows_generation()


def _recon(token, bit_id, ar_id, ar_amt, bit_amt, plug=None, generated=""):
    rec = {
        "token": token, "status": "approved", "uploaded": "2026-07-20",
        "uploaded_by": "operator:4003025705",
        "source": f"IRO 4003025705 — {token}", "account": "4003025705",
        "payment_refs": ["DEP-1"],
        "statement": {"label": f"IRO {token}", "date": "2026-07-01",
                      "total": ar_amt,
                      "lines": [{"awb": "1005475881", "amount": ar_amt,
                                 "matched_ids": [ar_id]}]},
        "ar_selected": [ar_id], "bit_selected": bit_id,
        "bit_candidates": [bit_id], "rows_gen": _gen,
        "frozen": {"ar_rows": [{"id": ar_id, "sap_acct": "4003025705",
                                "awb": "1005475881",
                                "assignment": "1005475881",
                                "amount": ar_amt}],
                   "bit_row": {"id": bit_id, "gl_account": "512000",
                               "assignment": f"PAY{bit_id + 1}",
                               "amount": bit_amt,
                               "raw": [f"PAY{bit_id + 1}", str(bit_amt)]},
                   "rows_gen": _gen, "at": "2026-07-20 10:00"},
        "slip_total": bit_amt, "slip": None, "extra_slips": [], "file": ""}
    if plug:
        rec["plug"] = plug
    if generated:
        rec["generated_at"] = generated
    bitcash.save_recon(rec)
    return token


# A short-paid reconciliation (150,000 invoiced, 130,000 banked → plug 20,000)
# and a second, already-journalled one.
_recon("aaa111aaa111", 0, 0, 150000.0, 130000.0,
       plug={"amount": 20000.0, "account": "4003025705",
             "note": "payment difference on payment reference DEP-1"})
_recon("bbb222bbb222", 1, 1, 50000.0, 50000.0, generated="2026-07-19 08:00")

# === 1. Only the SELECTED reconciliations are journalled =====================
check("journalable_recons lists both approved sandboxes",
      {r["token"] for r in bitcash.journalable_recons()}
      == {"aaa111aaa111", "bbb222bbb222"})
out1 = _tmp / "j_one.xlsx"
res1 = bitcash.build_journal(out1, tokens=["aaa111aaa111"])
check("journal built for the picked reconciliation only",
      res1 and res1["tokens"] == ["aaa111aaa111"] and res1["count"] == 1)
out_all = _tmp / "j_all.xlsx"
res_all = bitcash.build_journal(out_all)
check("the service default (tokens=None) still means every approved sandbox",
      res_all and res_all["count"] == 2)
check("an unknown/empty selection journals nothing",
      bitcash.build_journal(_tmp / "j_none.xlsx", tokens=["zzz"]) is None)

# === 2. The plug pair uses ONLY posting keys 40 and 15 ======================
wb1 = openpyxl.load_workbook(out1)
ws1 = wb1[next(s for s in wb1.sheetnames if s.startswith("CM01_"))]
rows = []
for r in range(4, ws1.max_row + 1):
    acct = ws1.cell(row=r, column=10).value
    if acct in (None, ""):
        continue
    rows.append({"account": str(acct), "key": ws1.cell(row=r, column=11).value,
                 "amount": ws1.cell(row=r, column=12).value,
                 "assignment": str(ws1.cell(row=r, column=14).value or "")})
keys = {r["key"] for r in rows}
check("journal uses ONLY posting keys 40 and 15", keys == {40, 15})
plug_lines = [r for r in rows if "difference" in r["assignment"].lower()
              or r["amount"] in (-20000.0, -20000)]
check("the plug is a balanced 40/15 pair",
      len(plug_lines) == 2
      and {r["key"] for r in plug_lines} == {40, 15})
check("plug direction rides on the sign (short payment posts negative)",
      all(r["amount"] == -20000.0 for r in plug_lines))
check("the plug's customer side is the reseller's account",
      any(r["account"] == "4003025705" and r["key"] == 15
          for r in plug_lines))
# the document still balances: sum of key-40 lines == sum of key-15 lines
d40 = round(sum(r["amount"] for r in rows if r["key"] == 40), 2)
d15 = round(sum(r["amount"] for r in rows if r["key"] == 15), 2)
check("journal document balances (40 side == 15 side)", d40 == d15)
check("the bank side nets to the amount actually banked", d40 == 130000.0)

# === 3. IRO cannot submit without the amount paid / a variance reason =======
from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402

client = TestClient(main.app)
tok = iro.ensure_token("4003025705")["token"]
# (rendered BEFORE the submits below, which close the last open airwaybill)
r = client.get(f"/operator/{tok}")
check("portal offers a Remove control for attached evidence",
      "discard-slip" in r.text and "Remove" in r.text)
base = {"awb": "1011864556", "ref_1011864556": "DEP-9", "reference": "DEP-9",
        "bank": "BICEC", "payment_method": "Bank deposit"}
r = client.post(f"/operator/{tok}/submit", data=dict(base))
check("submit refused when the amount actually paid is missing",
      r.status_code == 200 and "amount actually paid" in r.text)
r = client.post(f"/operator/{tok}/submit",
                data=dict(base, paid_1011864556="45000"))
check("submit refused when a variance has no explanation",
      r.status_code == 200 and "Explain the difference" in r.text)
r = client.post(f"/operator/{tok}/submit",
                data=dict(base, paid_1011864556="45000",
                          comment_1011864556="bank charges deducted"))
check("submit accepted once the variance is explained",
      r.status_code == 200 and "Explain the difference" not in r.text
      and "amount actually paid" not in r.text)

# === 4. IRO deletes unsent evidence; a submitted one is protected ===========
side = {"slip_id": "iroslip_abcd1234", "name": "iroslip_abcd1234.jpg",
        "account": "4003025705", "sha": "shaX", "total": 5000.0,
        "date": "2026-07-07", "source": "slip.jpg"}
(iro.UPLOAD_DIR / "iroslip_abcd1234.json").write_text(json.dumps(side),
                                                      encoding="utf-8")
(iro.UPLOAD_DIR / "iroslip_abcd1234.jpg").write_bytes(b"img")
check("an unsent evidence can be removed by its own operator",
      iro.discard_preread("iroslip_abcd1234", "4003025705") is True
      and not (iro.UPLOAD_DIR / "iroslip_abcd1234.jpg").exists())
# re-create it, then CLAIM it (as a submission does) — now it is protected
(iro.UPLOAD_DIR / "iroslip_abcd1234.json").write_text(json.dumps(side),
                                                      encoding="utf-8")
(iro.UPLOAD_DIR / "iroslip_abcd1234.jpg").write_bytes(b"img")
iro.claim_deposit("shaX", "4003025705", 5000.0, "2026-07-07", "DEP-X", "tk")
check("evidence already sent to finance can NOT be removed",
      iro.discard_preread("iroslip_abcd1234", "4003025705") is False
      and (iro.UPLOAD_DIR / "iroslip_abcd1234.jpg").exists())
check("another operator can never remove someone else's evidence",
      iro.discard_preread("iroslip_abcd1234", "9999") is False)
r = client.post(f"/operator/{tok}/discard-slip",
                json={"slip_id": "iroslip_abcd1234"})
check("the discard route refuses an already-sent evidence",
      r.status_code == 409)

# === 5. Journal UI wiring ===================================================
r = client.get("/tools/bit-cash-ar")
check("journal section lets finance tick which reconciliations to journal",
      'id="jeform"' in r.text and 'name="token"' in r.text
      and "ticked reconciliation" in r.text)

# === 6. The IRO statement email is three lines ==============================
subject, body = iro.statement_email(
    {"name": "SMIC LUXURY SHOP", "total": 10568327.0, "count": 186},
    "https://www.rtrotc.com/operator/abc", "4003025705")
lines = [ln for ln in body.splitlines() if ln.strip()]
print("   email body:", " | ".join(lines))
check("statement email is at most three lines", len(lines) <= 3)
check("email still carries account, total and the link",
      "4003025705" in body and "10,568,327" in body
      and "/operator/abc" in body)

# === 7. REVIEW FIX: ticking NOTHING must never re-journal everything ========
_stamp_before = bitcash.load_recon("bbb222bbb222").get("generated_at")
r = client.post("/tools/bit-cash-ar/journal", data={"picker": "1"},
                follow_redirects=False)
check("empty selection is refused, not treated as 'all'",
      r.status_code == 303 and "error=" in r.headers["location"]
      and "Nothing+ticked" in r.headers["location"].replace("%20", "+"))
# nothing was journalled, so no sandbox was re-stamped
check("an empty pick re-journals (and re-stamps) nothing",
      bitcash.load_recon("bbb222bbb222").get("generated_at")
      == _stamp_before)
# picking one explicitly still works through the route
r = client.post("/tools/bit-cash-ar/journal",
                data={"picker": "1", "token": "aaa111aaa111"},
                follow_redirects=False)
check("an explicit pick journals through the route",
      r.status_code == 303 and "message=" in r.headers["location"])
r = client.get("/tools/bit-cash-ar")
check("the journal form posts the picker marker", 'name="picker"' in r.text)

# === 8. REVIEW FIX: a typed 0 is a real answer, not 'missing' ===============
tok2 = iro.ensure_token("4003025705")["token"]
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit_header": [], "gen_bit": "g9", "gen_cash": "g9", "bit": [],
    "cash": [{"id": 0, "sap_acct": "4003025705", "awb": "7010101010",
              "assignment": "7010101010", "reference": "R9", "amount": 9000.0,
              "customer": "SMIC", "doc_no": "9", "date": "2026-07-02"}]},
    ensure_ascii=False), encoding="utf-8")
z = {"awb": "7010101010", "ref_7010101010": "DEP-Z", "reference": "DEP-Z"}
r = client.post(f"/operator/{tok2}/submit", data=dict(z, paid_7010101010="0"))
check("a typed 0 is NOT reported as a missing amount",
      "Enter the amount actually paid for every ticked" not in r.text)
check("a typed 0 against a 9,000 invoice demands an explanation",
      "Explain the difference" in r.text)

# === 9. REVIEW FIX: a refusal re-fills the operator's typed work ============
r = client.post(f"/operator/{tok2}/submit",
                data=dict(z, paid_7010101010="8500",
                          comment_7010101010="", payment_date="2026-07-09",
                          slip_id="iroslip_deadbeef"))
flat = " ".join(r.text.split())
check("the refused page keeps the amount paid, reference and date",
      'value="8500"' in flat and 'value="DEP-Z"' in flat
      and 'value="2026-07-09"' in flat)
check("the refused page keeps the airwaybill ticked",
      'value="7010101010" class="iro-tick"' in flat.replace('  ', ' ')
      or ('7010101010' in flat and 'checked' in flat))
check("the refused page keeps the attached evidence",
      'name="slip_id" value="iroslip_deadbeef"' in flat)
r = client.post(f"/operator/{tok2}/submit",
                data=dict(z, paid_7010101010="8500",
                          comment_7010101010="bank charge"))
check("submitting with the explanation now succeeds",
      "Submission received" in r.text)

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nALL v11.8 TESTS PASSED")
