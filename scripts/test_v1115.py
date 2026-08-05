"""v11.15 — the IRO statement gains a per-AWB mode of payment + provider,
ticks pre-fill the amount paid with running totals, a confirmation summary
gates the send, the per-mode Bank total anchors the BIT hunt, and the plug's
journal lines carry SHORT PAYMENT / OVERPAYMENT with the payment reference.

Isolated to a temp data dir; the real data/ is never touched.
"""
import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app import config  # noqa: E402
from app.tools import bitcash, iro  # noqa: E402

_tmp = Path(tempfile.mkdtemp(prefix="v1115_"))
config.CONFIG_PATH = _tmp / "config.json"
config.invalidate_config_cache()
iro.IRO_DIR = _tmp / "iro"
iro.IRO_DIR.mkdir(parents=True, exist_ok=True)
iro.UPLOAD_DIR = _tmp / "uploads"
iro.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
iro.DEPOSITS_PATH = _tmp / "deposits.json"
bitcash.RECON_DIR = _tmp / "recons"
bitcash.RECON_DIR.mkdir(parents=True, exist_ok=True)
bitcash.FILES_DIR = _tmp / "recon_files"
bitcash.FILES_DIR.mkdir(parents=True, exist_ok=True)
bitcash.ROWS_PATH = _tmp / "rows.json"
bitcash.STORE_PATH = _tmp / "bitcash.json"

from fastapi.testclient import TestClient  # noqa: E402

from app import main  # noqa: E402

main.OUTPUT_DIR = _tmp / "out"
main.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
main.UPLOAD_DIR = iro.UPLOAD_DIR

client = TestClient(main.app)
_fail = 0


def check(label, cond):
    global _fail
    print(("[OK ] " if cond else "[FAIL] ") + label)
    if not cond:
        _fail += 1


# === 0. Fixture: one operator, four open AWBs; BIT carries the BANK sum =====
ACCT = "4003025705"
AWBS = ["7010101010", "7010101020", "7010101030", "7010101040"]
AMTS = [100000.0, 50000.0, 30000.0, 20000.0]
bitcash.ROWS_PATH.write_text(json.dumps({
    "bit": [
        # 150,000 = what was paid BY BANK (AWBs 1+2) — not the 200,000 grand
        # total, and not on any slip.
        {"id": 0, "amount": -150000.0, "gl_account": "1263001293",
         "assignment": "0548407000018", "posting_date": "01.07.2026",
         "reference": "CMDLA", "text": "CASH DEPOSIT DLA", "doc_no": "410001",
         "raw": []},
        {"id": 1, "amount": -200000.0, "gl_account": "1263001293",
         "assignment": "0548407000019", "posting_date": "01.07.2026",
         "reference": "X", "text": "OTHER", "doc_no": "410002", "raw": []}],
    "bit_header": [], "gen_bit": "g1", "gen_cash": "g2",
    "cash_date_col": "Doc. Date",
    "cash": [{"id": i, "sap_acct": ACCT, "ibs_acct": "415048444",
              "awb": a, "assignment": a, "reference": a, "amount": m,
              "customer": "OP", "doc_no": str(i), "date": "01.07.2026"}
             for i, (a, m) in enumerate(zip(AWBS, AMTS))]},
    ensure_ascii=False), encoding="utf-8")

tok = iro.ensure_token(ACCT)["token"]

# === 1. The portal page: mode + provider columns, no Reference column =======
r = client.get(f"/operator/{tok}")
check("portal renders", r.status_code == 200)
check("Reference column replaced by Mode of payment",
      "<th>Reference</th>" not in r.text
      and "<th>Mode of payment</th>" in r.text
      and "<th>Provider</th>" in r.text)
check("all four modes offered per row",
      all(f'value="{m}"' in r.text for m in iro.PAYMENT_MODES)
      and r.text.count('class="iro-mode"') == len(AWBS))
check("provider select + cash deposit-reference box per row",
      r.text.count('class="iro-provider"') == len(AWBS)
      and r.text.count('class="iro-cashref"') == len(AWBS))
check("providers piped from config to the script",
      "BICEC" in r.text and "MTN" in r.text
      and ("Société Générale Cameroun" in r.text
           or "Soci\\u00e9t\\u00e9 G\\u00e9n\\u00e9rale Cameroun" in r.text))
check("tick pre-fills the amount and keeps a running paid total",
      'id="iro-total-paid"' in r.text and "dataset.auto" in r.text)
check("a confirmation step gates the send",
      'id="iro-confirm"' in r.text and "Kindly confirm before sending"
      in r.text and "Confirm & send" in r.text)
check("the old global payment-method select is gone",
      'name="payment_method"' not in r.text)

# === 2. Drafts carry modes + providers ======================================
r = client.post(f"/operator/{tok}/draft", json={
    "ticks": {AWBS[0]: "DEP-1"}, "modes": {AWBS[0]: "Mobile Money"},
    "providers": {AWBS[0]: "MTN"}, "reference": "DEP-1",
    "payment_date": "2026-08-01"})
check("draft accepts modes and providers", r.status_code == 200
      and r.json().get("ok") is True)
d = iro.load_record(ACCT).get("draft") or {}
check("draft stores them", d.get("modes", {}).get(AWBS[0]) == "Mobile Money"
      and d.get("providers", {}).get(AWBS[0]) == "MTN")
r = client.get(f"/operator/{tok}")
check("the page restores the drafted mode",
      'value="Mobile Money" selected' in r.text)

# === 3. Submit with mixed modes — NO deposit slip attached ==================
r = client.post(f"/operator/{tok}/submit", data={
    "awb": AWBS,
    f"ref_{AWBS[0]}": "DEP-BANK-1", f"ref_{AWBS[1]}": "DEP-BANK-1",
    f"ref_{AWBS[2]}": "", f"ref_{AWBS[3]}": "",
    f"paid_{AWBS[0]}": "100000", f"paid_{AWBS[1]}": "50000",
    f"paid_{AWBS[2]}": "30000", f"paid_{AWBS[3]}": "20000",
    f"mode_{AWBS[0]}": "Bank", f"mode_{AWBS[1]}": "Bank",
    f"mode_{AWBS[2]}": "Mobile Money", f"mode_{AWBS[3]}": "Cash",
    f"provider_{AWBS[0]}": "BICEC", f"provider_{AWBS[1]}": "BICEC",
    f"provider_{AWBS[2]}": "ORANGE", f"provider_{AWBS[3]}": "CASHDEP-77",
    "reference": "DEP-BANK-1", "bank": "BICEC"})
check("submission accepted without any deposit slip",
      r.status_code == 200 and "Submission received" in r.text)
sub = iro.load_record(ACCT)["submissions"][-1]
rtok = sub["recon_token"]

import time  # noqa: E402

for _ in range(80):
    rec = bitcash.load_recon(rtok) or {}
    if rec.get("status") in ("open", "error"):
        break
    time.sleep(0.25)
check("reconciliation opened", rec.get("status") == "open")

# === 4. The per-mode groups the IRO confirmed are on the sandbox ============
groups = {g["mode"]: g for g in rec.get("mode_totals", [])}
check("Bank group: 150,000 over two AWBs via BICEC",
      groups.get("Bank", {}).get("total") == 150000.0
      and sorted(groups["Bank"]["awbs"]) == sorted(AWBS[:2])
      and groups["Bank"]["providers"] == ["BICEC"])
check("Mobile Money group: 30,000 via ORANGE",
      groups.get("Mobile Money", {}).get("total") == 30000.0
      and groups["Mobile Money"]["providers"] == ["ORANGE"])
check("Cash group: 20,000, deposit reference kept",
      groups.get("Cash", {}).get("total") == 20000.0
      and groups["Cash"]["providers"] == ["CASHDEP-77"])
check("the cash deposit reference joined the payment references",
      "CASHDEP-77" in rec.get("payment_refs", []))

# === 5. The BANK total (not the grand total) anchors the BIT ================
check("BIT selected on the bank-mode sum of 150,000",
      rec.get("bit_selected") == 0)
view = bitcash.recon_view(rec)
check("view discloses the mode breakdown + bank anchor",
      view["bank_total"] == 150000.0 and len(view["mode_totals"]) == 3)
r = client.get(f"/tools/bit-cash-ar/recon/{rtok}")
check("sandbox page shows the declared modes",
      "As declared by the operator" in r.text and "150,000" in r.text)

# unit: bank_total outranks slip + evidence totals
st = {"total": 0, "lines": [{"reference": AWBS[0], "amount": 200000.0}]}
_l, _a, cands, sel = bitcash.automatch(st, slip_total=None,
                                       bank_total=150000.0)
check("automatch picks the bank-mode amount first",
      sel == 0 and cands[0] == 0)

# === 6. mode_groups defaults ================================================
check("a return that never carried modes (email channel) yields no groups —"
      " so no bank anchor is invented",
      iro.mode_groups([{"awb": "1", "amount": 10.0, "amount_paid": None}])
      == [])
g = iro.mode_groups([
    {"awb": "1", "amount": 10.0, "amount_paid": None, "mode": ""},
    {"awb": "2", "amount": 5.0, "amount_paid": 5.0, "mode": "Bank"}])
check("within a moded return, a blank mode defaults to Bank",
      g == [{"mode": "Bank", "total": 15.0, "awbs": ["1", "2"],
             "providers": []}])

# the confirm gate: edits void a shown confirmation; re-entry is guarded
OP = (ROOT / "app" / "templates" / "operator" / "statement.html").read_text(
    encoding="utf-8")
check("editing the form voids a confirmation on display",
      "function unconfirm()" in OP and OP.count("unconfirm();") >= 2)
check("a second submit while sending is swallowed",
      "if (submitted) { e.preventDefault(); return; }" in OP)
check("a failed validation voids the confirmation",
      "confirmed = false;      // whatever was approved is now void" in OP)
check("Confirm & send cannot double-fire",
      "this.disabled = true;" in OP)

# === 7. Journal: SHORT PAYMENT / OVERPAYMENT + payment ref in assignment ====
import openpyxl  # noqa: E402

_gen = {"bit": "g1", "cash": "g2"}


def _plug_recon(token, plug_amt, refs):
    bitcash.save_recon({
        "token": token, "status": "approved", "uploaded": "2026-08-01",
        "uploaded_by": f"operator:{ACCT}", "source": f"IRO {ACCT}",
        "account": ACCT, "payment_refs": refs,
        "statement": {"label": "t", "date": "2026-08-01", "total": 100000.0,
                      "lines": [{"reference": AWBS[0], "amount": 100000.0,
                                 "matched_ids": [0]}]},
        "ar_selected": [0], "bit_selected": 0, "bit_candidates": [0],
        "rows_gen": _gen,
        "frozen": {"ar_rows": [{"id": 0, "sap_acct": ACCT, "awb": AWBS[0],
                                "assignment": AWBS[0], "amount": 100000.0}],
                   "bit_row": {"id": 0, "gl_account": "1263001293",
                               "assignment": "0548407000018",
                               "reference": "CMDLA",
                               "amount": -(100000.0 - plug_amt), "raw": []},
                   "rows_gen": _gen, "at": "2026-08-01 10:00"},
        "plug": {"amount": plug_amt, "account": ACCT, "note": "why"},
        "slip_total": None, "slip": None, "extra_slips": [], "file": ""})


def _journal_rows(tokens):
    out = _tmp / f"j_{'_'.join(tokens)}.xlsx"
    bitcash.build_journal(out, tokens=tokens)
    wb = openpyxl.load_workbook(out)
    ws = wb[next(s for s in wb.sheetnames if s.startswith("CM01_"))]
    return [r for r in ws.iter_rows(min_row=4, values_only=True)
            if r[9] not in (None, "")]


_plug_recon("aaaa0000aaaa", 25000.0, ["PAYREF-SHORT-1"])   # short payment
rows = _journal_rows(["aaaa0000aaaa"])
plug_rows = [r for r in rows if r[11] == -25000.0]
check("short payment: plug pair posts -25,000 on keys 40+15",
      sorted(r[10] for r in plug_rows) == [15, 40])
check("short payment: text column reads SHORT PAYMENT",
      all(r[7] == "SHORT PAYMENT" for r in plug_rows))
check("short payment: customer-side assignment = the payment reference",
      any(r[10] == 15 and r[13] == "PAYREF-SHORT-1" for r in plug_rows))
check("AWB lines keep MANUAL RECIEPT",
      all(r[7] == "MANUAL RECIEPT " for r in rows if r not in plug_rows))

_plug_recon("bbbb0000bbbb", -15000.0, [])                  # overpayment
rows = _journal_rows(["bbbb0000bbbb"])
plug_rows = [r for r in rows if r[11] == 15000.0]
check("overpayment: plug pair posts +15,000 on keys 40+15",
      sorted(r[10] for r in plug_rows) == [15, 40])
check("overpayment: text column reads OVERPAYMENT",
      all(r[7] == "OVERPAYMENT" for r in plug_rows))
check("overpayment with no submitted reference falls back to the BIT's",
      any(r[10] == 15 and r[13] == "CMDLA" for r in plug_rows))

# === 8. The evidence Excel carries mode + provider for finance ==============
ev = iro.UPLOAD_DIR / "ev_check.xlsx"
iro.build_evidence_xlsx(ev, [
    {"awb": "1", "amount": 10.0, "reference": "R", "amount_paid": 10.0,
     "comment": "", "mode": "Mobile Money", "provider": "MTN"}])
wb = openpyxl.load_workbook(ev)
ws = wb.active
check("evidence file has Mode of payment + Provider columns",
      ws.cell(row=1, column=6).value == "Mode of payment"
      and ws.cell(row=1, column=7).value == "Provider"
      and ws.cell(row=2, column=6).value == "Mobile Money"
      and ws.cell(row=2, column=7).value == "MTN")

if _fail:
    print(f"\n{_fail} CHECK(S) FAILED")
    sys.exit(1)
print("\nALL v11.15 TESTS PASSED")
